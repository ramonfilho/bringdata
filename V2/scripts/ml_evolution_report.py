"""
ml_evolution_report.py — Relatório de Evolução de Performance ML

Script único que agrega todas as fontes de dados para gerar a tabela de
evolução de performance ML ao longo dos lançamentos.

FONTES:
  1. Google Sheets (Produção + Backup) — decil por lead, com dedup global
  2. Cloud SQL backup (.sql) — CAPI enviados + decil para períodos pré-Railway
  3. Railway PostgreSQL — CAPI enviados + decil para períodos pós-fev/18
  4. Cloud Run logs — emails + decis para gap entre as planilhas
  5. Validation xlsx reports — AUC, CPA, ROAS, concentrações por decil

PARA ADICIONAR UM NOVO LANÇAMENTO:
  - Adicionar entrada em PERIODS (cap_start, cap_end, vendas_start, vendas_end)
  - O xlsx report é auto-detectado pela pasta de vendas ou pelo glob
  - Rodar: python V2/scripts/ml_evolution_report.py
"""

import os
import re
import json
import subprocess
import tempfile
import gzip
from pathlib import Path
from datetime import datetime

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────

BASE = Path(__file__).parent.parent  # V2/

# Mapeamento vendas_start → nome do lançamento.
# Único lugar a editar ao criar um novo lançamento (apenas o nome, sem datas).
PERIOD_NAMES = {
    '2025-12-08': 'LF40',
    '2025-12-15': 'LF41',
    '2025-12-22': 'LF42',
    '2026-01-19': 'DEV19',
    '2026-02-02': 'LF43',
    '2026-02-09': 'LF44',
    '2026-03-02': 'LF45',
    '2026-03-09': 'LF46',
    '2026-03-16': 'LF47',
    '2026-03-23': 'LF48',
}


def discover_periods() -> list:
    """
    Auto-detecta períodos a partir das pastas de validação.
    Lê 'Performance Geral' de cada relatório para extrair as datas reais.
    Não contém datas hardcoded — adicionar apenas o nome em PERIOD_NAMES ao
    criar um novo lançamento.
    """
    periods = []
    for folder in sorted(VALIDATION_DIR.iterdir()):
        if not folder.is_dir() or ':' not in folder.name:
            continue
        reports = sorted(folder.glob('validation_report_*.xlsx'))
        if not reports:
            continue
        try:
            pg = pd.read_excel(reports[-1], sheet_name='Performance Geral', header=None)
            rows = {
                str(r[0]).strip(): str(r.iloc[1]).strip()
                for _, r in pg.iterrows()
                if pd.notna(r[0]) and len(r) > 1 and pd.notna(r.iloc[1])
            }
            cap_str = rows.get('Período de Captação', '')
            ven_str = rows.get('Período de Vendas', '')
            if ' a ' not in cap_str or ' a ' not in ven_str:
                continue
            cap_start, cap_end       = [s.strip() for s in cap_str.split(' a ')]
            vendas_start, vendas_end = [s.strip() for s in ven_str.split(' a ')]
            name = PERIOD_NAMES.get(vendas_start, folder.name)
            periods.append({
                'name':         name,
                'cap_start':    cap_start,
                'cap_end':      cap_end,
                'vendas_start': vendas_start,
                'vendas_end':   vendas_end,
            })
        except Exception as e:
            print(f"  Aviso: falha ao ler período {folder.name}: {e}")
    periods.sort(key=lambda p: p['vendas_start'])
    return periods

# Google Sheets IDs
SHEETS = {
    'producao': '1VYti8jX277VNMkvzrfnJSR_Ko8L1LQFDdMEeD6D8_Vo',  # [LF] Pesquisa - Produção
    'backup':   '1OqNYA5zU9ix1uf52ovRYIdLhcugzwgfKOheKxE_zgvE',   # [LF] Pesquisa - Backup
}
SHEETS_TAB = '[LF] Pesquisa'

# Cloud SQL backup
CLOUDSQL_PATH = BASE / 'data/backups/cloud-sql-final-export-20260225.sql'

# Railway cutover (primeiro registro real no Railway)
RAILWAY_CUTOVER = pd.Timestamp('2026-02-18')

# Cloud Run
CLOUDRUN_PROJECT = 'smart-ads-451319'
CLOUDRUN_SERVICE = 'bring-data-api'

# Validation xlsx folder
VALIDATION_DIR = BASE / 'outputs/validation'
OUTPUT_DIR     = BASE / 'outputs/validation/historico'

# Períodos auto-detectados das pastas de validação (sem datas hardcoded)
PERIODS = discover_periods()

# ─────────────────────────────────────────────────────────────────────────────
# ENV
# ─────────────────────────────────────────────────────────────────────────────

_env_file = BASE / '.env'
if _env_file.exists():
    with open(_env_file) as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith('#') and '=' in _line:
                _k, _v = _line.split('=', 1)
                os.environ.setdefault(_k.strip(), _v.strip())

# ─────────────────────────────────────────────────────────────────────────────
# FONTE 1 — Google Sheets
# ─────────────────────────────────────────────────────────────────────────────

def _load_sheet_csv(sheet_id: str, gid: int) -> pd.DataFrame:
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
    with tempfile.NamedTemporaryFile(suffix='.csv', delete=False) as tmp:
        subprocess.run(['curl', '-sL', '--max-time', '60', url, '-o', tmp.name],
                       capture_output=True, timeout=65)
        df = pd.read_csv(tmp.name, low_memory=False)
        os.unlink(tmp.name)
    return df


def load_sheets_data() -> pd.DataFrame:
    """
    Carrega [LF] Pesquisa de Produção e Backup, combina e deduplica globalmente.
    Retorna DataFrame com colunas: email, data, decil, lead_score (Score).
    """
    print("  [Sheets] Carregando planilhas...")
    import gspread
    from google.auth import default as gauth_default

    scopes = ['https://www.googleapis.com/auth/spreadsheets.readonly']
    creds, _ = gauth_default(scopes=scopes)
    gc = gspread.authorize(creds)

    frames = []
    for label, sid in SHEETS.items():
        try:
            sp = gc.open_by_key(sid)
            ws = sp.worksheet(SHEETS_TAB)
            df = _load_sheet_csv(sid, ws.id)
            df['_source'] = label
            frames.append(df)
            print(f"    {label}: {len(df):,} linhas")
        except Exception as e:
            print(f"    WARN: falha ao carregar {label}: {e}")

    if not frames:
        return pd.DataFrame()

    combined = pd.concat(frames, ignore_index=True)

    # Normalizar colunas chave
    col_map = {}
    for col in combined.columns:
        cl = col.lower().strip()
        if cl in ('e-mail', 'email'):
            col_map[col] = 'email'
        elif cl == 'data' or cl == 'data do envio':
            col_map[col] = 'data'
        elif cl == 'score':
            col_map[col] = 'lead_score'
        elif cl == 'decil':
            col_map[col] = 'decil'
    combined = combined.rename(columns=col_map)

    for req in ('email', 'data'):
        if req not in combined.columns:
            combined[req] = None

    combined['data'] = pd.to_datetime(combined['data'], errors='coerce')
    combined['email'] = combined['email'].astype(str).str.lower().str.strip()

    has_decil = combined['decil'].notna() & (combined['decil'].astype(str).str.strip() != '')
    print(f"    Total bruto: {len(combined):,} linhas | Com decil: {has_decil.sum():,}")

    # Detectar gap entre as planilhas (para referência e Cloud Run logs)
    b_max = combined[combined['_source'] == 'backup']['data'].max() if 'backup' in combined['_source'].values else None
    p_min = combined[combined['_source'] == 'producao']['data'].min() if 'producao' in combined['_source'].values else None
    if b_max is not None and p_min is not None:
        print(f"    Backup até: {b_max}  |  Produção a partir de: {p_min}")
        print(f"    Gap detectado: {b_max} → {p_min}")

    # Retorna o DataFrame RAW (sem dedup global) —
    # sheets_decil_stats faz filtra-primeiro + dedup por período
    return combined


def _normalize_decil(raw) -> str | None:
    """Normaliza qualquer formato de decil para 'D1'–'D10'.

    Aceita: 'D01', 'D1', 'D10', '01', '1', '10', 1, 10, etc.
    """
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None
    s = str(raw).strip().upper().lstrip('D')  # remove prefixo 'D'
    try:
        return f'D{int(s)}'
    except ValueError:
        return None


def sheets_decil_stats(sheets_raw_df: pd.DataFrame, cap_start: str, cap_end: str) -> dict:
    """
    Calcula distribuição de decis para um período de captação via Sheets.

    Usa o DataFrame RAW (sem dedup global) — filtra por data PRIMEIRO,
    deduplica por email dentro do período, depois conta D10%.
    """
    if sheets_raw_df.empty:
        return {}

    start = pd.Timestamp(cap_start)
    end   = pd.Timestamp(cap_end) + pd.Timedelta(days=1)

    # Filtrar pelo período de captação
    period_df = sheets_raw_df[
        (sheets_raw_df['data'] >= start) & (sheets_raw_df['data'] < end)
    ].copy()

    if period_df.empty:
        return {}

    # Normalizar decil
    period_df['decil_norm'] = period_df['decil'].apply(_normalize_decil)
    has_decil = period_df['decil_norm'].notna()

    # Dedup por email dentro do período — priorizar quem tem decil
    with_d    = period_df[has_decil].drop_duplicates('email', keep='last')
    without_d = period_df[~has_decil & ~period_df['email'].isin(with_d['email'])] \
                         .drop_duplicates('email', keep='last')
    period_deduped = pd.concat([with_d, without_d], ignore_index=True)

    scored = period_deduped[period_deduped['decil_norm'].notna()]
    total  = len(scored)
    if total == 0:
        return {'sheets_scored': 0, 'sheets_decil_dist': {}, 'sheets_d10_pct': None}

    dist = scored['decil_norm'].value_counts().sort_index().to_dict()
    d10  = dist.get('D10', 0)

    return {
        'sheets_scored':     len(period_deduped),
        'sheets_decil_dist': dist,
        'sheets_d10_pct':    round(d10 / total * 100, 1),
    }

# ─────────────────────────────────────────────────────────────────────────────
# FONTE 2 — Cloud SQL backup (.sql)
# ─────────────────────────────────────────────────────────────────────────────

def load_cloudsql_backup() -> pd.DataFrame:
    """Parseia o backup SQL e retorna DataFrame com colunas relevantes."""
    path = CLOUDSQL_PATH
    if not path.exists():
        print(f"  [CloudSQL] Backup não encontrado: {path}")
        return pd.DataFrame()

    print(f"  [CloudSQL] Parsing {path.name}...")
    # Colunas confirmadas:
    # 0=id 1=email 2=name 3=phone 4=fbp 5=fbc 6=event_id 7=user_agent
    # 8=client_ip 9=event_source_url 10=utm_source 11=utm_medium 12=utm_campaign
    # 13=utm_term 14=utm_content 15=tem_comp 16=created_at 17=updated_at
    # 18=first_name 19=last_name 20=capi_sent_at 21=capi_response_status ...
    # 36=lead_score 37=decil 38=scored_at
    C_EMAIL      = 1
    C_CREATED    = 16
    C_CAPI_SENT  = 20
    C_DECIL      = 37

    records = []
    in_copy = False
    opener = gzip.open if str(path).endswith('.gz') else open
    with opener(path, 'rt', encoding='utf-8', errors='replace') as f:
        for line in f:
            if line.startswith('COPY public.leads_capi'):
                in_copy = True; continue
            if in_copy:
                if line.strip() == '\\.':
                    in_copy = False; continue
                parts = line.rstrip('\n').split('\t')
                if len(parts) < 38:
                    continue
                records.append({
                    'email':        parts[C_EMAIL].lower().strip(),
                    'created_at':   parts[C_CREATED][:19]  if parts[C_CREATED]  != '\\N' else None,
                    'capi_sent_at': parts[C_CAPI_SENT][:10] if parts[C_CAPI_SENT] != '\\N' else None,
                    'decil':        parts[C_DECIL]          if parts[C_DECIL]     != '\\N' else None,
                })

    df = pd.DataFrame(records)
    df['created_at'] = pd.to_datetime(df['created_at'], errors='coerce')
    print(f"    {len(df):,} registros | {df['created_at'].min().date()} → {df['created_at'].max().date()}")
    return df

# ─────────────────────────────────────────────────────────────────────────────
# FONTE 3 — Railway PostgreSQL
# ─────────────────────────────────────────────────────────────────────────────

def load_railway() -> pd.DataFrame:
    """Carrega todos os registros do Railway e normaliza decil para 'D10' format."""
    try:
        import pg8000.native
        conn = pg8000.native.Connection(
            host=os.environ.get('RAILWAY_DB_HOST', 'shortline.proxy.rlwy.net'),
            port=int(os.environ.get('RAILWAY_DB_PORT', '11594')),
            database=os.environ.get('RAILWAY_DB_NAME', 'railway'),
            user=os.environ.get('RAILWAY_DB_USER', 'postgres'),
            password=os.environ['RAILWAY_DB_PASSWORD'],
            ssl_context=True,
        )
        rows = conn.run('SELECT email, "createdAt", "capiSentAt", decil, "leadScore" FROM "Lead"')
        conn.close()
        df = pd.DataFrame(rows, columns=['email', 'created_at', 'capi_sent_at', 'decil', 'lead_score'])
        df['email']      = df['email'].astype(str).str.lower().str.strip()
        df['created_at'] = pd.to_datetime(df['created_at'], utc=True).dt.tz_convert(None)
        # Railway armazena decil como integer (1–10) → converter para 'D10'
        df['decil'] = df['decil'].apply(lambda x: f'D{int(x)}' if pd.notna(x) else None)
        print(f"  [Railway] {len(df):,} registros | {df['created_at'].min().date()} → {df['created_at'].max().date()}")
        return df
    except Exception as e:
        print(f"  [Railway] Falha na conexão: {e}")
        return pd.DataFrame()

# ─────────────────────────────────────────────────────────────────────────────
# FONTE 4 — Cloud Run logs (gap entre planilhas)
# ─────────────────────────────────────────────────────────────────────────────

def load_cloudrun_logs(start_ts: str, end_ts: str) -> pd.DataFrame:
    """
    Extrai email + decil dos logs do Cloud Run para o período informado.

    start_ts / end_ts: formato ISO 8601 UTC, ex: '2026-02-08T15:01:18Z'
    Retorna DataFrame com colunas: email, decil, timestamp
    """
    print(f"  [CloudRun] Consultando logs {start_ts} → {end_ts}...")

    filter_str = (
        f'resource.type="cloud_run_revision" '
        f'AND resource.labels.service_name="{CLOUDRUN_SERVICE}" '
        f'AND textPayload=~"LeadQualified enviado" '
        f'AND timestamp>"{start_ts}" '
        f'AND timestamp<"{end_ts}"'
    )
    cmd = [
        'gcloud', 'logging', 'read', filter_str,
        f'--project={CLOUDRUN_PROJECT}',
        '--limit=50000',
        '--format=json',
    ]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        if result.returncode != 0:
            print(f"    WARN: gcloud error: {result.stderr[:200]}")
            return pd.DataFrame()

        entries = json.loads(result.stdout) if result.stdout.strip() else []
        print(f"    {len(entries)} entradas de log encontradas")

        # Regex: "✅ LeadQualified enviado: email@x.com (decil: D10, ..."
        # Excluir LeadQualifiedHighQuality (é evento secundário, para não duplicar)
        pattern = re.compile(
            r'✅ LeadQualified enviado: (\S+@\S+) \(decil: (D\d+)',
            re.IGNORECASE
        )
        records = []
        for entry in entries:
            text = entry.get('textPayload', '')
            # Pular HighQuality para não contar em dobro
            if 'HighQuality' in text:
                continue
            m = pattern.search(text)
            if m:
                records.append({
                    'email':     m.group(1).lower().strip(),
                    'decil':     m.group(2),
                    'timestamp': entry.get('timestamp', ''),
                })

        df = pd.DataFrame(records)
        if not df.empty:
            df = df.drop_duplicates(subset=['email'], keep='last')
            print(f"    {len(df):,} leads únicos com decil extraídos dos logs")
        return df

    except Exception as e:
        print(f"    WARN: falha ao consultar logs: {e}")
        return pd.DataFrame()

# ─────────────────────────────────────────────────────────────────────────────
# FONTE 5 — Validation xlsx reports
# ─────────────────────────────────────────────────────────────────────────────

def find_xlsx_for_period(vendas_start: str, vendas_end: str) -> Path | None:
    """
    Auto-detecta o relatório xlsx de validação para um período de vendas.
    Procura pelo padrão: validation_report_FECHAMENTO_{vendas_start}_to_{vendas_end}_*.xlsx
    """
    pattern = f"**/validation_report_*{vendas_start}*{vendas_end}*.xlsx"
    candidates = sorted(VALIDATION_DIR.glob(pattern))
    if candidates:
        return candidates[-1]  # mais recente

    # Fallback: procura por pasta com datas de vendas (dd:mm)
    vs = pd.Timestamp(vendas_start)
    ve = pd.Timestamp(vendas_end)
    folder_pattern = f"{vs.day:02d}:{vs.month:02d} - {ve.day:02d}:{ve.month:02d}"
    folder = VALIDATION_DIR / folder_pattern
    if folder.exists():
        candidates = sorted(folder.glob('validation_report_*.xlsx'))
        if candidates:
            return candidates[-1]

    return None


def parse_xlsx_report(path: Path) -> dict:
    """Extrai métricas-chave do relatório xlsx de validação."""
    if path is None or not path.exists():
        return {}

    xl = pd.ExcelFile(path)
    metrics = {'xlsx_path': str(path)}

    # Performance Geral
    if 'Performance Geral' in xl.sheet_names:
        pg = xl.parse('Performance Geral', header=None)
        row_dict = {}
        for _, row in pg.iterrows():
            if pd.notna(row[0]) and len(row) > 1 and pd.notna(row.iloc[1]):
                row_dict[str(row[0]).strip()] = row.iloc[1]
        metrics['leads_meta']    = _get_val(row_dict, 'Leads Meta')
        metrics['leads_capi']    = _get_val(row_dict, 'Pessoas únicas (CAPI)')
        metrics['leads_survey']  = _get_val(row_dict, 'Respostas na pesquisa')
        metrics['vendas_total']  = _get_val(row_dict, 'Vendas no Período')
        metrics['vendas_match']  = _get_val(row_dict, 'Vendas identificadas')
        metrics['pct_tracking']  = _get_val(row_dict, '% de trackeamento')
        metrics['gasto_total']   = _get_val(row_dict, 'Gasto Total')

    # Performance ML — linha de resumo (4ª linha após header)
    if 'Performance ML' in xl.sheet_names:
        pm = xl.parse('Performance ML', header=None)
        for i, row in pm.iterrows():
            if row.notna().sum() >= 7:
                vals = [v for v in row if pd.notna(v)]
                try:
                    # Linha: gasto | leads | conv_track | conv_real | cpa | roas | roas_adj_tmb | ...
                    if isinstance(vals[0], (int, float)) and vals[0] > 1000:
                        metrics['gasto_ml']    = float(vals[0])
                        metrics['leads_ml']    = int(vals[1])
                        metrics['conv_track']  = float(vals[2])
                        metrics['conv_real']   = float(vals[3])
                        metrics['cpa_ml']      = float(vals[4])
                        metrics['roas_ml']     = float(vals[5])
                        metrics['roas_adj']    = float(vals[6])
                        break
                except (IndexError, ValueError, TypeError):
                    continue

    # Fallback: Comparação ML (estrutura pós-18/03 — sem aba Performance ML)
    if 'gasto_ml' not in metrics and 'Comparação ML' in xl.sheet_names:
        cm = xl.parse('Comparação ML', header=None)
        in_ml_section = False
        for i, row in cm.iterrows():
            vals = [v for v in row if pd.notna(v)]
            if not vals:
                continue
            label = str(vals[0]).strip()
            if 'COMPARAÇÃO ML' in label.upper():
                if in_ml_section:
                    # Segunda ocorrência = seção Matched Pairs — parar aqui
                    break
                in_ml_section = True
                continue
            if not in_ml_section:
                continue
            if 'ADSETS' in label.upper() or 'MATCHED' in label.upper():
                break
            if len(vals) >= 2 and isinstance(vals[1], (int, float)):
                if label == 'Gasto':
                    metrics['gasto_ml'] = float(vals[1])
                elif label == 'Leads':
                    metrics['leads_ml'] = int(vals[1])
                elif label == 'Conversões':
                    metrics.setdefault('conv_track', float(vals[1]))
                    metrics.setdefault('conv_real',  float(vals[1]))
                elif label == 'ROAS':
                    metrics['roas_ml'] = float(vals[1])
                elif label == 'CPA':
                    metrics['cpa_ml']  = float(vals[1])

    # ML Monitoring — AUC e concentrações
    if 'ML Monitoring' in xl.sheet_names:
        mm = xl.parse('ML Monitoring', header=None)
        for i, row in mm.iterrows():
            vals = [v for v in row if pd.notna(v)]
            if not vals:
                continue
            label = str(vals[0]).strip()
            if label == 'AUC' and len(vals) >= 3:
                try:
                    metrics['auc_prod']  = float(vals[1])
                    metrics['auc_test']  = float(vals[2])
                    metrics['auc_delta'] = float(vals[3]) if len(vals) > 3 else None
                except (ValueError, TypeError):
                    pass
            elif 'Top 3' in label and len(vals) >= 2:
                try:
                    metrics['top3_decis'] = float(vals[1])
                except (ValueError, TypeError):
                    pass
            elif 'Top 5' in label and len(vals) >= 2:
                try:
                    metrics['top5_decis'] = float(vals[1])
                except (ValueError, TypeError):
                    pass
            elif 'Leads com score' in label and len(vals) >= 2:
                try:
                    raw = str(vals[1]).replace(',', '').replace('.', '')
                    metrics['leads_scored_mon'] = int(raw)
                except (ValueError, TypeError):
                    pass
            elif 'CORRELAÇÃO' in label.upper() and ':' in label and len(vals) >= 2:
                try:
                    metrics['corr_tmb_auc'] = float(vals[1])
                except (ValueError, TypeError):
                    pass

    return metrics


def _get_val(d: dict, key: str):
    """Busca chave no dicionário com match parcial."""
    for k, v in d.items():
        if key.lower() in k.lower():
            return v
    return None

# ─────────────────────────────────────────────────────────────────────────────
# CAPI stats por período (Cloud SQL + Railway combinados)
# ─────────────────────────────────────────────────────────────────────────────

def capi_stats_for_period(cap_start: str, cap_end: str,
                           csql_df: pd.DataFrame, rail_df: pd.DataFrame,
                           cloudrun_df: pd.DataFrame) -> dict:
    """
    Combina Cloud SQL, Railway e Cloud Run logs para obter:
    - Total de leads no período
    - CAPI enviados
    - Leads com decil
    - Distribuição por decil / D10%
    """
    start = pd.Timestamp(cap_start)
    end   = pd.Timestamp(cap_end) + pd.Timedelta(days=1)

    frames = []

    # Cloud SQL — cobre até fev/25 (antes do Railway)
    if not csql_df.empty:
        cs_end = min(end, RAILWAY_CUTOVER)
        mask = (csql_df['created_at'] >= start) & (csql_df['created_at'] < cs_end)
        sub = csql_df[mask][['email', 'created_at', 'capi_sent_at', 'decil']].copy()
        sub['_src'] = 'cloudsql'
        frames.append(sub)

    # Railway — a partir do cutover
    if not rail_df.empty:
        r_start = max(start, RAILWAY_CUTOVER)
        if r_start < end:
            mask = (rail_df['created_at'] >= r_start) & (rail_df['created_at'] < end)
            sub = rail_df[mask][['email', 'created_at', 'capi_sent_at', 'decil']].copy()
            sub['_src'] = 'railway'
            frames.append(sub)

    if not frames:
        return {}

    df = pd.concat(frames, ignore_index=True)
    df = df.drop_duplicates(subset=['email'], keep='last')

    # Cloud Run logs NÃO são usados para enriquecer decil aqui:
    # os logs cobrem o gap 08–16/02 e atribuir esse decil a períodos
    # anteriores (DEV19/LF43/LF44) geraria contaminação — seria o decil
    # de um re-cadastro posterior, não do período original.
    # D10% via DB é confiável apenas onde o Railway armazena decil (LF45+).

    total = len(df)
    sent  = df['capi_sent_at'].notna().sum()
    has_d = df['decil'].notna() & (df['decil'] != '')
    scored_df = df[has_d]
    n_scored  = len(scored_df)

    decil_dist = scored_df['decil'].value_counts().to_dict() if n_scored > 0 else {}
    d10 = decil_dist.get('D10', 0)
    # Só reportar D10% se tiver volume mínimo confiável (Railway = LF45+)
    d10_pct = round(d10 / n_scored * 100, 1) if n_scored >= 100 else None

    return {
        'db_leads':   total,
        'capi_sent':  int(sent),
        'db_scored':  n_scored,
        'decil_dist': decil_dist,
        'd10_pct':    d10_pct,
    }

# ─────────────────────────────────────────────────────────────────────────────
# LIFT POR DECIL — cruzamento leads×compradores
# ─────────────────────────────────────────────────────────────────────────────

def compute_decil_lift(xlsx_path: Path, sheets_raw_df: pd.DataFrame,
                       cap_start: str, cap_end: str,
                       rail_df: pd.DataFrame) -> pd.DataFrame:
    """
    Calcula taxa de conversão real por decil cruzando:
      - Compradores: aba 'Detalhes das Conversões' do xlsx de validação
      - Decil por lead: Google Sheets (filtrado por cap_start/cap_end) + Railway

    Retorna DataFrame com colunas:
      decil | leads | buyers | conv_rate_pct | lift
    onde lift = conv_rate / baseline (taxa média de todos os leads com decil).
    """
    if xlsx_path is None or not xlsx_path.exists():
        return pd.DataFrame()

    xl = pd.ExcelFile(xlsx_path)
    if 'Detalhes das Conversões' not in xl.sheet_names:
        return pd.DataFrame()

    # ── 1. Emails dos compradores ──
    det = xl.parse('Detalhes das Conversões', header=None)
    # Localizar linha de header (contém 'E-mail')
    header_idx = None
    for i, row in det.iterrows():
        if any('mail' in str(v).lower() for v in row.values):
            header_idx = i
            break
    if header_idx is None:
        return pd.DataFrame()

    det.columns = det.iloc[header_idx].values
    det = det.iloc[header_idx + 1:].reset_index(drop=True)

    email_col = next((c for c in det.columns if 'mail' in str(c).lower()), None)
    if email_col is None:
        return pd.DataFrame()

    buyer_emails = set(
        det[email_col].dropna().astype(str).str.lower().str.strip()
    )

    # ── 2. Leads com decil no período de captação ──
    start = pd.Timestamp(cap_start)
    end   = pd.Timestamp(cap_end) + pd.Timedelta(days=1)

    leads_df = pd.DataFrame()

    if not sheets_raw_df.empty:
        period = sheets_raw_df[
            (sheets_raw_df['data'] >= start) & (sheets_raw_df['data'] < end)
        ].copy()
        period['decil_norm'] = period['decil'].apply(_normalize_decil)
        has_d = period['decil_norm'].notna()
        with_d    = period[has_d].drop_duplicates('email', keep='last')
        without_d = period[~has_d & ~period['email'].isin(with_d['email'])] \
                          .drop_duplicates('email', keep='last')
        leads_df = pd.concat([with_d, without_d], ignore_index=True)[['email', 'decil_norm']]

    # Railway complementa para períodos com cobertura (LF45+)
    if not rail_df.empty:
        r_mask = (rail_df['created_at'] >= start) & (rail_df['created_at'] < end)
        rail_p  = rail_df[r_mask].copy()
        rail_p['decil_norm'] = rail_p['decil'].apply(_normalize_decil)
        new_rail = rail_p[~rail_p['email'].isin(leads_df['email'])]
        leads_df = pd.concat(
            [leads_df, new_rail[['email', 'decil_norm']]], ignore_index=True
        )

    if leads_df.empty:
        return pd.DataFrame()

    leads_df['bought'] = leads_df['email'].isin(buyer_emails)
    scored = leads_df[leads_df['decil_norm'].notna()].copy()

    if scored.empty:
        return pd.DataFrame()

    baseline = scored['bought'].mean() * 100  # taxa média (todos os decis)

    DECILS = [f'D{i}' for i in range(1, 11)]
    rows = []
    for d in DECILS:
        sub = scored[scored['decil_norm'] == d]
        n   = len(sub)
        b   = int(sub['bought'].sum())
        cr  = round(b / n * 100, 3) if n > 0 else 0.0
        lift = round(cr / baseline, 2) if baseline > 0 and n > 0 else None
        rows.append({'decil': d, 'leads': n, 'buyers': b,
                     'conv_rate_pct': cr, 'lift': lift})

    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL OUTPUT
# ─────────────────────────────────────────────────────────────────────────────

def build_excel(rows: list[dict], output_path: Path):
    """Gera a tabela de evolução em xlsx com formatação."""

    HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
    SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
    ALT_FILL     = PatternFill("solid", fgColor="D9E1F2")
    BEST_FILL    = PatternFill("solid", fgColor="C6EFCE")
    WORST_FILL   = PatternFill("solid", fgColor="FFC7CE")
    WARN_FILL    = PatternFill("solid", fgColor="FFEB9C")
    WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")

    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
    SECTION_FONT = Font(bold=True, color="FFFFFF", size=10)
    NORMAL       = Font(size=10)
    SMALL        = Font(size=9, italic=True, color="595959")

    CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    thin   = Side(style='thin', color='B8CCE4')
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evolução ML"

    names = [r['name'] for r in rows]
    ncols = 1 + len(names)

    def sc(row, col, value, font=None, fill=None, align=None, num_fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        if font:    c.font      = font
        if fill:    c.fill      = fill
        if align:   c.alignment = align
        if num_fmt: c.number_format = num_fmt
        c.border = BORDER
        return c

    def section(row, text):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=text)
        c.font = SECTION_FONT; c.fill = SECTION_FILL
        c.alignment = CENTER; c.border = BORDER
        ws.row_dimensions[row].height = 20

    def data_row(row, label, values, fills=None, fmt=None, label_fill=None):
        sc(row, 1, label, font=NORMAL, fill=label_fill or WHITE_FILL, align=LEFT)
        for i, v in enumerate(values):
            f = fills[i] if fills else None
            sc(row, i+2, v, font=NORMAL, fill=f or WHITE_FILL, align=CENTER,
               num_fmt=fmt[i] if fmt else None)
        ws.row_dimensions[row].height = 18

    def bw(vals, higher=True):
        """Best=green, Worst=red fills for numeric values."""
        nums = [(i, v) for i, v in enumerate(vals) if isinstance(v, (int, float))]
        if len(nums) < 2:
            return [None] * len(vals)
        best_i  = (max if higher else min)(nums, key=lambda x: x[1])[0]
        worst_i = (min if higher else max)(nums, key=lambda x: x[1])[0]
        fills   = [None] * len(vals)
        fills[best_i]  = BEST_FILL
        fills[worst_i] = WORST_FILL
        return fills

    r = 1

    # Title
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1, value="EVOLUÇÃO DE PERFORMANCE ML — DevClub")
    c.font = Font(bold=True, color="FFFFFF", size=14)
    c.fill = HEADER_FILL; c.alignment = CENTER
    ws.row_dimensions[r].height = 30; r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1,
                value=f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
                      f"Modelo: v1_devclub_rf_temporal_leads_single (treinado 30/01/2026)")
    c.font = SMALL; c.alignment = CENTER; c.fill = ALT_FILL
    ws.row_dimensions[r].height = 16; r += 2

    # Column headers
    sc(r, 1, 'Métrica', font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    for i, name in enumerate(names):
        sc(r, i+2, name, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    ws.row_dimensions[r].height = 22; r += 1

    def col(key, default=None):
        return [rd.get(key, default) for rd in rows]

    # ── Períodos ──
    section(r, 'PERÍODOS'); r += 1
    data_row(r, 'Captação início',  col('cap_start'));  r += 1
    data_row(r, 'Captação fim',     col('cap_end'));    r += 1
    data_row(r, 'Vendas início',    col('vendas_start')); r += 1
    data_row(r, 'Vendas fim',       col('vendas_end')); r += 1

    # ── Volume de Leads ──
    section(r, 'VOLUME DE LEADS & COBERTURA'); r += 1
    data_row(r, 'Leads Meta (total)',       col('leads_meta'),
             bw(col('leads_meta')), ['#,##0']*len(names)); r += 1
    data_row(r, 'Leads únicos no banco CAPI (Cloud SQL → Railway)',  col('leads_capi'),
             bw(col('leads_capi')), ['#,##0']*len(names)); r += 1
    data_row(r, 'Respostas na pesquisa',    col('leads_survey'),
             bw(col('leads_survey')), ['#,##0']*len(names)); r += 1
    data_row(r, 'Vendas totais no período', col('vendas_total'),
             bw(col('vendas_total')), ['#,##0']*len(names)); r += 1
    data_row(r, 'Vendas identificadas (matched)', col('vendas_match'),
             bw(col('vendas_match')), ['#,##0']*len(names)); r += 1
    pct = [f"{v:.1%}" if isinstance(v, float) else v for v in col('pct_tracking')]
    data_row(r, '% Trackeamento', pct, bw(col('pct_tracking'))); r += 1

    # ── Helpers financeiros (fonte única: Comparação ML — All Campaigns) ──
    def _mval(key):
        return [rd.get('margem_data', {}).get(key) for rd in rows]

    def _brl(v):
        if v is None:
            return 'n/d'
        try:
            fv = float(v)
            return 'n/d' if np.isnan(fv) else f"R$ {fv:,.0f}"
        except (TypeError, ValueError):
            return 'n/d'

    def _rfmt(v):
        if v is None:
            return 'n/d'
        try:
            fv = float(v)
            return 'n/d' if np.isnan(fv) else f"{fv:.2f}x"
        except (TypeError, ValueError):
            return 'n/d'

    # ── Performance Financeira — Investimento & Resultados ──
    section(r, 'PERFORMANCE FINANCEIRA — INVESTIMENTO & RESULTADOS'); r += 1
    data_row(r, 'Gasto Total Lançamento (R$)', col('gasto_total'),
             None, ['R$#,##0']*len(names)); r += 1
    data_row(r, 'Gasto Campanhas ML (R$)',     col('gasto_ml'),
             None, ['R$#,##0']*len(names)); r += 1
    data_row(r, 'Gasto Controle (R$)',
             [_brl(v) for v in _mval('gasto_ctrl')]); r += 1
    pct_ml = [
        f"{ml/total*100:.1f}%" if isinstance(ml, (int,float)) and isinstance(total, (int,float)) and total > 0 else 'n/d'
        for ml, total in zip(col('gasto_ml'), col('gasto_total'))
    ]
    data_row(r, '% Orçamento em ML', pct_ml,
             bw([float(v.rstrip('%')) if isinstance(v,str) and v!='n/d' else None for v in pct_ml])); r += 1
    data_row(r, 'Leads totais do lançamento',  col('leads_meta'),
             None, ['#,##0']*len(names)); r += 1
    data_row(r, 'Leads em Campanhas ML',       col('leads_ml'),
             bw(col('leads_ml')), ['#,##0']*len(names)); r += 1
    data_row(r, 'Conversões totais',           col('vendas_total'),
             None, ['#,##0']*len(names)); r += 1
    data_row(r, 'Conversões ML Reais (Est.)',  col('conv_real'),
             bw(col('conv_real')), ['#,##0.0']*len(names)); r += 1
    taxa_conv = [
        f"{v/l*100:.2f}%" if isinstance(v, (int,float)) and isinstance(l, (int,float)) and l > 0 else 'n/d'
        for v, l in zip(col('conv_real'), col('leads_ml'))
    ]
    data_row(r, 'Taxa de conversão ML (conv. reais / leads ML)', taxa_conv,
             bw([float(v.rstrip('%')) if isinstance(v,str) and v!='n/d' else None for v in taxa_conv])); r += 1
    data_row(r, 'CPA ML (R$) ← menor é melhor', col('cpa_ml'),
             bw(col('cpa_ml'), higher=False), ['R$#,##0.00']*len(names)); r += 1

    # ── Receita & ROAS ──
    section(r, 'RECEITA & ROAS — ML vs CONTROLE'); r += 1
    data_row(r, 'Receita ML (R$)',
             [_brl(v) for v in _mval('receita_ml')],
             bw(_mval('receita_ml'))); r += 1
    data_row(r, 'Receita Controle (R$)',
             [_brl(v) for v in _mval('receita_ctrl')]); r += 1
    data_row(r, 'Receita Total (R$)',
             [_brl(v) for v in _mval('receita_total')],
             bw(_mval('receita_total'))); r += 1

    # ── Faturamento contratado e recebido (previsão de faturamento) ──
    TICKET           = 2_200.0
    GURU_TICKET      = 1_997.0   # preço real Guru (payment.gross via API) — ≠ ticket contratado
    GURU_REALIZACAO  = 0.87      # fator de realização: ~13% cancelamentos/chargebacks (back-calculado LF42–LF47)
    PCT_CARTAO       = 0.469     # mediana LF44–LF47
    N_PARCELAS       = 12        # entrada + 11x (fonte: contas_a_receber TMB)
    PARCELA_TMB      = TICKET / N_PARCELAS

    def _fat_contratado(vendas):
        if vendas is None or not isinstance(vendas, (int, float)):
            return None
        return vendas * TICKET

    def _fat_recebido(vendas):
        if vendas is None or not isinstance(vendas, (int, float)):
            return None
        v_guru = vendas * PCT_CARTAO
        v_tmb  = vendas * (1 - PCT_CARTAO)
        return v_guru * GURU_TICKET * GURU_REALIZACAO + v_tmb * PARCELA_TMB

    fat_cont_vals = [_fat_contratado(v) for v in col('vendas_total')]
    fat_rec_vals  = [_fat_recebido(v)  for v in col('vendas_total')]

    data_row(r, 'Fat. Contratado (R$) — vendas × R$2.200',
             [_brl(v) for v in fat_cont_vals],
             bw(fat_cont_vals)); r += 1
    data_row(r, 'Fat. Recebido Est. (R$) — cartão integral + 1ª parc. boleto',
             [_brl(v) for v in fat_rec_vals],
             bw(fat_rec_vals)); r += 1

    data_row(r, 'ROAS ML',
             [_rfmt(v) for v in _mval('roas_ml')],
             bw(_mval('roas_ml'))); r += 1
    data_row(r, 'ROAS Controle',
             [_rfmt(v) for v in _mval('roas_ctrl')]); r += 1
    data_row(r, 'ROAS Total do Lançamento',
             [_rfmt(v) for v in _mval('roas_total')],
             bw(_mval('roas_total'))); r += 1

    # ── Margem de Contribuição ──
    section(r, 'MARGEM DE CONTRIBUIÇÃO'); r += 1
    data_row(r, 'Margem ML (R$)',
             [_brl(v) for v in _mval('margem_ml')],
             bw(_mval('margem_ml'))); r += 1
    data_row(r, 'Margem Controle (R$)',
             [_brl(v) for v in _mval('margem_ctrl')]); r += 1
    data_row(r, 'Margem Total (R$)',
             [_brl(v) for v in _mval('margem_total')],
             bw(_mval('margem_total'))); r += 1

    # ── Análise Contrafactual ──
    section(r, 'ANÁLISE CONTRAFACTUAL (e se todo spend fosse ao ROAS Controle?)'); r += 1
    data_row(r, 'Receita Contrafactual (R$)',
             [_brl(v) for v in _mval('receita_cf')]); r += 1
    data_row(r, 'Margem Contrafactual (R$)',
             [_brl(v) for v in _mval('margem_cf')]); r += 1
    _ganho_vals = _mval('ganho_margem')
    _ganho_fills = [
        (BEST_FILL  if isinstance(v, (int, float)) and v > 0 else
         WORST_FILL if isinstance(v, (int, float)) and v < 0 else None)
        for v in _ganho_vals
    ]
    data_row(r, 'Ganho de Margem vs Contrafactual (R$)',
             [_brl(v) for v in _ganho_vals], _ganho_fills); r += 1

    # ── Qualidade do Modelo ML ──
    section(r, 'QUALIDADE DO MODELO ML'); r += 1
    data_row(r, 'AUC Produção ← >0.7 = bom', col('auc_prod'),
             bw(col('auc_prod')), ['0.0000']*len(names)); r += 1
    data_row(r, 'AUC Test Set (referência)',  col('auc_test'),
             None, ['0.0000']*len(names)); r += 1
    top3_fmt = [
        f"{v:.1%}" if isinstance(v, float) and v > 0 else 'n/d'
        for v in col('top3_decis')
    ]
    data_row(r, 'Concentração Top 3 Decis (D8/D9/D10)',
             top3_fmt,
             bw([v if isinstance(v, float) and v > 0 else None for v in col('top3_decis')])); r += 1
    data_row(r, 'Eventos CAPI enviados à Meta (LeadQualified)', col('capi_sent'),
             bw(col('capi_sent')), ['#,##0']*len(names)); r += 1
    data_row(r, 'Leads com decil nas Sheets (por período, dedup)',
             col('sheets_scored'), bw(col('sheets_scored')), ['#,##0']*len(names)); r += 1
    s_d10_vals = [f"{v:.1f}%" if isinstance(v, float) else (v or 'n/d') for v in col('sheets_d10_pct')]
    s_d10_fills = [
        (WORST_FILL if isinstance(v, float) and v >= 40 else
         WARN_FILL  if isinstance(v, float) and v >= 30 else None)
        for v in col('sheets_d10_pct')
    ]
    data_row(r, '% D10 nas Sheets ← ≥40% = alerta loop CAPI', s_d10_vals, s_d10_fills); r += 1

    # ── Lift por Decil — D1–D10 completo ──
    section(r, 'LIFT POR DECIL — CONVERSÃO REAL D1–D10 (leads × compradores)'); r += 1

    def _lift_val(rows_data, decil, field, pct=False):
        results = []
        for rd in rows_data:
            df = rd.get('decil_lift_df')
            if df is None or df.empty:
                results.append('n/d')
                continue
            row_d = df[df['decil'] == decil]
            if row_d.empty:
                results.append('n/d')
                continue
            v = row_d.iloc[0][field]
            if pct and isinstance(v, float):
                results.append(f"{v:.3f}%")
            else:
                results.append(v if pd.notna(v) else 'n/d')
        return results

    def _baseline_conv(rows_data):
        results = []
        for rd in rows_data:
            df = rd.get('decil_lift_df')
            if df is None or df.empty:
                results.append('n/d')
                continue
            total_leads = df['leads'].sum()
            total_buyers = df['buyers'].sum()
            if total_leads > 0:
                results.append(f"{total_buyers / total_leads * 100:.3f}%")
            else:
                results.append('n/d')
        return results

    data_row(r, 'Taxa de conversão baseline (todos os decis com score)',
             _baseline_conv(rows)); r += 1

    for d in [f'D{i}' for i in range(10, 0, -1)]:
        cr_vals   = _lift_val(rows, d, 'conv_rate_pct', pct=True)
        lift_vals = _lift_val(rows, d, 'lift')
        lift_fills = [
            (BEST_FILL if isinstance(v, float) and v >= 2 else
             WARN_FILL if isinstance(v, float) and v >= 1 else
             WORST_FILL if isinstance(v, float) else None)
            for v in lift_vals
        ]
        combined = []
        for cr, lf in zip(cr_vals, lift_vals):
            if cr == 'n/d':
                combined.append('n/d')
            elif lf == 'n/d':
                combined.append(cr)
            else:
                combined.append(f"{cr}  (lift {lf}x)")
        data_row(r, f'Decil {d} — taxa conv. (lift vs baseline)', combined, lift_fills); r += 1

    # ── Notas ──
    r += 1
    section(r, 'FONTES & NOTAS'); r += 1
    notes = [
        "FONTES: Google Sheets ([LF] Pesquisa — Produção + Backup) | Cloud SQL backup (sql) | Railway PostgreSQL | Cloud Run logs | Validation xlsx",
        "Gap entre Sheets: Backup até 08/02/2026 15:01 | Produção a partir de 16/02/2026 07:48 → preenchido via Cloud Run logs",
        "Railway cutover: 18/02/2026 — antes: Cloud SQL; depois: Railway",
        "D10% Sheets: filtro-por-período PRIMEIRO, depois dedup — fonte primária para DEV19/LF43/LF44 (logs Cloud Run expirados após 30d)",
        "D10% DB/Railway: apenas LF45 tem dados confiáveis (Railway) — Cloud SQL tem só 31 registros com decil (Feb 15-16)",
        "DEV19 D10%=18.6% é esperado: primeiro lançamento com CAPI+ML, Meta ainda sem feedback → distribuição quase uniforme D1-D10",
        "Modelo único para todos os períodos: v1_devclub_rf_temporal_leads_single (treinado 30/01/2026, AUC test=0.7311)",
        "LF45 praticamente sem grupo controle (62 leads vs 27.553 ML) — comparação ML vs controle não representativa",
    ]
    for note in notes:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        c = ws.cell(row=r, column=1, value=f"• {note}")
        c.font = SMALL; c.alignment = LEFT
        ws.row_dimensions[r].height = 16; r += 1

    # Column widths
    ws.column_dimensions['A'].width = 55
    for i in range(len(names)):
        ws.column_dimensions[chr(ord('B') + i)].width = 18
    ws.freeze_panes = 'B5'

    # ── Aba Resumo (inserida como primeira aba) ──
    _build_summary_sheet(wb, rows)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ─────────────────────────────────────────────────────────────────────────────
# ABA RESUMO EXECUTIVO
# ─────────────────────────────────────────────────────────────────────────────

def _build_summary_sheet(wb, rows):
    """
    Cria aba 'Resumo' como primeira aba do workbook.
    Métricas-chave por lançamento em formato compacto para leitura rápida.
    """
    ws = wb.create_sheet("Resumo", 0)  # posição 0 = primeira aba

    HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
    SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
    ALT_FILL     = PatternFill("solid", fgColor="D9E1F2")
    BEST_FILL    = PatternFill("solid", fgColor="C6EFCE")
    WORST_FILL   = PatternFill("solid", fgColor="FFC7CE")
    WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")

    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
    SECTION_FONT = Font(bold=True, color="FFFFFF", size=10)
    NORMAL       = Font(size=10)
    SMALL        = Font(size=9, italic=True, color="595959")

    CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT   = Alignment(horizontal='left',   vertical='center')

    thin   = Side(style='thin', color='B8CCE4')
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    names = [r['name'] for r in rows]
    ncols = 1 + len(names)

    def sc(row, col, value, font=None, fill=None, align=None, num_fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        if font:    c.font      = font
        if fill:    c.fill      = fill
        if align:   c.alignment = align
        if num_fmt: c.number_format = num_fmt
        c.border = BORDER
        return c

    def section(row, text):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=text)
        c.font = SECTION_FONT; c.fill = SECTION_FILL
        c.alignment = CENTER; c.border = BORDER
        ws.row_dimensions[row].height = 18

    def data_row(row, label, values, fills=None, fmt=None):
        sc(row, 1, label, font=NORMAL, fill=WHITE_FILL, align=LEFT)
        for i, v in enumerate(values):
            f = fills[i] if fills else None
            sc(row, i + 2, v, font=NORMAL, fill=f or WHITE_FILL, align=CENTER,
               num_fmt=fmt[i] if fmt else None)
        ws.row_dimensions[row].height = 18

    def bw(vals, higher=True):
        nums = [(i, v) for i, v in enumerate(vals) if isinstance(v, (int, float))]
        if len(nums) < 2:
            return [None] * len(vals)
        best_i  = (max if higher else min)(nums, key=lambda x: x[1])[0]
        worst_i = (min if higher else max)(nums, key=lambda x: x[1])[0]
        out = [None] * len(vals)
        out[best_i]  = BEST_FILL
        out[worst_i] = WORST_FILL
        return out

    def col(key):
        return [rd.get(key) for rd in rows]

    def mval(key):
        return [rd.get('margem_data', {}).get(key) for rd in rows]

    def brl_str(v):
        if v is None: return 'n/d'
        try:
            fv = float(v)
            return 'n/d' if np.isnan(fv) else f"R$ {fv:,.0f}"
        except (TypeError, ValueError):
            return 'n/d'

    r = 1

    # Título
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1, value="RESUMO — PERFORMANCE ML por LANÇAMENTO")
    c.font = Font(bold=True, color="FFFFFF", size=14)
    c.fill = HEADER_FILL; c.alignment = CENTER
    ws.row_dimensions[r].height = 28; r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1,
                value=f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  DevClub")
    c.font = SMALL; c.alignment = CENTER; c.fill = ALT_FILL
    ws.row_dimensions[r].height = 15; r += 2

    # Headers de lançamento
    sc(r, 1, 'Métrica', font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    for i, name in enumerate(names):
        sc(r, i + 2, name, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    ws.row_dimensions[r].height = 22; r += 1

    # ── Período ──
    section(r, 'PERÍODO'); r += 1
    data_row(r, 'Captação início', col('cap_start')); r += 1
    data_row(r, 'Captação fim',    col('cap_end'));   r += 1
    data_row(r, 'Vendas início',   col('vendas_start')); r += 1
    data_row(r, 'Vendas fim',      col('vendas_end'));   r += 1

    # ── Investimento ──
    section(r, 'INVESTIMENTO'); r += 1
    data_row(r, 'Gasto total (R$)', col('gasto_total'),
             None, ['R$#,##0'] * len(names)); r += 1
    data_row(r, 'Gasto ML (R$)', col('gasto_ml'),
             None, ['R$#,##0'] * len(names)); r += 1
    pct_ml = [
        f"{ml / total * 100:.1f}%"
        if isinstance(ml, (int, float)) and isinstance(total, (int, float)) and total > 0
        else 'n/d'
        for ml, total in zip(col('gasto_ml'), col('gasto_total'))
    ]
    data_row(r, '% Orçamento em ML', pct_ml,
             bw([float(v.rstrip('%')) if isinstance(v, str) and v != 'n/d' else None
                 for v in pct_ml])); r += 1

    # ── Leads & Conversão ──
    section(r, 'LEADS & CONVERSÃO'); r += 1
    data_row(r, 'Leads', col('leads_meta'),
             bw(col('leads_meta')), ['#,##0'] * len(names)); r += 1
    cpl_vals = [
        round(g / l, 2)
        if isinstance(g, (int, float)) and isinstance(l, (int, float)) and l > 0
        else None
        for g, l in zip(col('gasto_total'), col('leads_meta'))
    ]
    data_row(r, 'CPL (R$)', cpl_vals,
             bw(cpl_vals, higher=False), ['R$#,##0.00'] * len(names)); r += 1
    data_row(r, 'Vendas', col('vendas_total'),
             bw(col('vendas_total')), ['#,##0'] * len(names)); r += 1
    pct_conv = [
        f"{v / l * 100:.2f}%"
        if isinstance(v, (int, float)) and isinstance(l, (int, float)) and l > 0
        else 'n/d'
        for v, l in zip(col('vendas_total'), col('leads_meta'))
    ]
    data_row(r, '% Conversão', pct_conv,
             bw([float(v.rstrip('%')) if isinstance(v, str) and v != 'n/d' else None
                 for v in pct_conv])); r += 1

    # ── Resultado Financeiro ──
    section(r, 'RESULTADO FINANCEIRO'); r += 1
    data_row(r, 'Receita (R$)', [brl_str(v) for v in mval('receita_total')],
             bw(mval('receita_total'))); r += 1
    roas_vals = mval('roas_total')
    roas_fmt  = [f"{v:.2f}x" if isinstance(v, float) else 'n/d' for v in roas_vals]
    data_row(r, 'ROAS', roas_fmt, bw(roas_vals)); r += 1
    data_row(r, 'Margem (R$)', [brl_str(v) for v in mval('margem_total')],
             bw(mval('margem_total'))); r += 1
    cpa_total = [
        round(g / v, 2)
        if isinstance(g, (int, float)) and isinstance(v, (int, float)) and v > 0
        else None
        for g, v in zip(col('gasto_total'), col('vendas_total'))
    ]
    data_row(r, 'CPA (R$)', cpa_total,
             bw(cpa_total, higher=False), ['R$#,##0.00'] * len(names)); r += 1

    # Larguras
    ws.column_dimensions['A'].width = 25
    for i in range(len(names)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i + 2)].width = 15
    ws.freeze_panes = 'B5'


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def run(extra_period: dict = None):
    """
    Gera o relatório de evolução. Pode ser chamado diretamente de outros scripts.

    Args:
        extra_period: dict opcional com keys name, cap_start, cap_end,
                      vendas_start, vendas_end — adicionado ao PERIODS se
                      ainda não estiver na lista (match por vendas_start+vendas_end).
    """
    import copy
    periods = copy.deepcopy(PERIODS)

    if extra_period:
        existing = {(p['vendas_start'], p['vendas_end']) for p in periods}
        key = (extra_period['vendas_start'], extra_period['vendas_end'])
        if key not in existing:
            periods.append(extra_period)
            print(f"  + Período adicionado dinamicamente: {extra_period['name']}")
    else:
        periods = PERIODS

    print("=" * 60)
    print("ml_evolution_report.py — carregando fontes...")
    print("=" * 60)

    # ── Carregar fontes globais ──
    print("\n[1/5] Google Sheets")
    sheets_df = load_sheets_data()

    gap_start = gap_end = None
    if not sheets_df.empty and '_source' in sheets_df.columns:
        b_max = sheets_df[sheets_df['_source'] == 'backup']['data'].max()
        p_min = sheets_df[sheets_df['_source'] == 'producao']['data'].min()
        if pd.notna(b_max) and pd.notna(p_min):
            gap_start = b_max.strftime('%Y-%m-%dT%H:%M:%SZ')
            gap_end   = p_min.strftime('%Y-%m-%dT%H:%M:%SZ')

    print("\n[2/5] Cloud SQL backup")
    csql_df = load_cloudsql_backup()

    print("\n[3/5] Railway")
    rail_df = load_railway()

    print("\n[4/5] Cloud Run logs (gap entre planilhas)")
    if gap_start and gap_end:
        cloudrun_df = load_cloudrun_logs(gap_start, gap_end)
    else:
        print("  Gap não detectado — pulando Cloud Run logs")
        cloudrun_df = pd.DataFrame()

    print("\n[5/5] Validation xlsx reports")

    print("\n" + "=" * 60)
    print("Processando períodos...")
    all_rows = []

    for p in periods:
        name = p['name']
        print(f"\n  {name} ({p['cap_start']} → {p['cap_end']})")

        row = {
            'name':         name,
            'cap_start':    p['cap_start'],
            'cap_end':      p['cap_end'],
            'vendas_start': p['vendas_start'],
            'vendas_end':   p['vendas_end'],
        }

        xlsx_path = find_xlsx_for_period(p['vendas_start'], p['vendas_end'])
        if xlsx_path:
            print(f"    xlsx: {xlsx_path.name}")
            row.update(parse_xlsx_report(xlsx_path))
            try:
                import sys as _sys_m
                _scripts_dir_m = str(Path(__file__).parent)
                if _scripts_dir_m not in _sys_m.path:
                    _sys_m.path.insert(0, _scripts_dir_m)
                from gerar_evolucao_margem import parse_comparacao_ml as _pcm
                row['margem_data'] = _pcm(xlsx_path)
            except Exception as _em:
                print(f"    Aviso: margem_data falhou — {_em}")
                row['margem_data'] = {}
        else:
            print(f"    xlsx: NÃO ENCONTRADO para {p['vendas_start']} → {p['vendas_end']}")
            xlsx_path = None
            row['margem_data'] = {}

        stats = capi_stats_for_period(p['cap_start'], p['cap_end'], csql_df, rail_df, cloudrun_df)
        row.update(stats)
        print(f"    DB leads: {stats.get('db_leads', 0):,} | CAPI sent: {stats.get('capi_sent', 0):,} "
              f"| Scored: {stats.get('db_scored', 0):,} | D10%: {stats.get('d10_pct', 'n/d')}")

        s_stats = sheets_decil_stats(sheets_df, p['cap_start'], p['cap_end'])
        row.update(s_stats)
        if s_stats:
            print(f"    Sheets scored: {s_stats.get('sheets_scored', 0):,} | D10%: {s_stats.get('sheets_d10_pct', 'n/d')}")

        lift_df = compute_decil_lift(xlsx_path, sheets_df, p['cap_start'], p['cap_end'], rail_df)
        row['decil_lift_df'] = lift_df
        if not lift_df.empty:
            d10_cr = lift_df[lift_df['decil'] == 'D10']['conv_rate_pct'].values
            d1_cr  = lift_df[lift_df['decil'] == 'D1']['conv_rate_pct'].values
            print(f"    Lift: D10 conv={d10_cr[0]:.3f}% | D1 conv={d1_cr[0]:.3f}%"
                  if len(d10_cr) and len(d1_cr) else "    Lift: calculado")

        all_rows.append(row)

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = OUTPUT_DIR / f"evolucao_ml_devclub_{ts}.xlsx"
    print(f"\nGerando Excel: {output_path}")
    build_excel(all_rows, output_path)
    print(f"✅ Salvo: {output_path}")

    # Adicionar Síntese Executiva automaticamente
    try:
        import sys as _sys
        _scripts_dir = str(Path(__file__).parent)
        if _scripts_dir not in _sys.path:
            _sys.path.insert(0, _scripts_dir)
        from gerar_evolucao_margem import update_sintese
        print("\nGerando Síntese Executiva...")
        update_sintese(output_path, periods)
    except Exception as _e:
        print(f"  Aviso: erro ao gerar Síntese Executiva: {_e}")

    import subprocess as sp
    sp.run(['open', str(output_path)], check=False)

    return output_path


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Gera relatório de evolução ML DevClub.")
    parser.add_argument('--name',         help='Nome do lançamento extra (ex: LF48)')
    parser.add_argument('--cap-start',    help='Início captação (YYYY-MM-DD)')
    parser.add_argument('--cap-end',      help='Fim captação (YYYY-MM-DD)')
    parser.add_argument('--vendas-start', help='Início vendas (YYYY-MM-DD)')
    parser.add_argument('--vendas-end',   help='Fim vendas (YYYY-MM-DD)')
    args = parser.parse_args()

    extra = None
    if args.name and args.cap_start and args.cap_end and args.vendas_start and args.vendas_end:
        extra = {
            'name':         args.name,
            'cap_start':    args.cap_start,
            'cap_end':      args.cap_end,
            'vendas_start': args.vendas_start,
            'vendas_end':   args.vendas_end,
        }

    run(extra)


if __name__ == '__main__':
    main()
