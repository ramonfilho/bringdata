"""
gerar_evolucao_margem.py

Adiciona nova sheet "Margem & Contrafactual" ao arquivo de evolução ML
e gera gráficos de margem, ROAS total e análise contrafactual.

Pergunta central: o sistema ML aumenta a margem de contribuição total
do negócio, ou apenas substitui ROAS das campanhas Controle?

Uso standalone:
    python V2/scripts/gerar_evolucao_margem.py

Também importável:
    from gerar_evolucao_margem import add_margem_sheet
    add_margem_sheet(xlsx_path, periods)  # periods: list[dict] do ml_evolution_report
"""

import sys
import yaml
from pathlib import Path
import warnings
warnings.filterwarnings('ignore')

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

BASE = Path(__file__).parent.parent  # V2/

GRAFICOS_DIR = BASE / 'outputs' / 'validation' / 'historico' / 'graficos'
GRAFICOS_DIR.mkdir(parents=True, exist_ok=True)


# ─────────────────────────────────────────────────────────────────────────────
# 1. SELEÇÃO DO RELATÓRIO MAIS RECENTE
# ─────────────────────────────────────────────────────────────────────────────

def get_latest_report(folder: str) -> Path | None:
    """Localiza o relatório xlsx para um período.

    Suporta duas convenções:
      - Legada:  pasta 'DD:MM - DD:MM' com 'validation_report_*.xlsx'
      - Atual:   pasta 'YYYY-MM' com 'LF* - DD:MM a DD:MM.xlsx'
    """
    # 1. Convenção legada
    path = BASE / 'outputs' / 'validation' / folder
    if path.exists():
        files = sorted(path.glob('validation_report_*.xlsx'))
        if files:
            return files[-1]

    # 2. Convenção atual: derivar pasta YYYY-MM e tag de datas do folder "DD:MM - DD:MM"
    import re
    m = re.match(r'^(\d{2}):(\d{2}) - (\d{2}):(\d{2})$', folder)
    if m:
        d1, mo1, d2, mo2 = m.groups()
        # Ano: inferir do contexto (2025 ou 2026) — usar ambos
        for year in ('2026', '2025'):
            month_folder = BASE / 'outputs' / 'validation' / f'{year}-{mo1}'
            if month_folder.exists():
                start_tag = f'{d1}:{mo1}'
                end_tag   = f'{d2}:{mo2}'
                candidates = [
                    f for f in sorted(month_folder.glob('*.xlsx'))
                    if start_tag in f.name and end_tag in f.name
                ]
                if candidates:
                    return candidates[-1]
    return None


# ─────────────────────────────────────────────────────────────────────────────
# 2. PARSING DA SHEET "COMPARAÇÃO ML"
# ─────────────────────────────────────────────────────────────────────────────

def _to_float(v, default=0.0):
    if v is None or (isinstance(v, str) and v.strip() in ('—', '-', 'N/A', '')):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def parse_comparacao_ml(xlsx_path: Path) -> dict:
    """
    Extrai métricas da aba 'Comparação ML' (bloco All vs All, não Matched Pairs).
    Suporta formato novo (18/03) e antigo (09/03).
    """
    df = pd.read_excel(xlsx_path, sheet_name='Comparação ML', header=None)

    result = {
        'gasto_ml': None, 'gasto_ctrl': None,
        'receita_ml': None, 'receita_ctrl': None,
        'roas_ml': None, 'roas_ctrl': None,
        'margem_ml': None, 'margem_ctrl': None,
        'formato': None,
    }

    # Detectar formato pelo conteúdo da primeira célula não-nula
    first_label = ''
    for _, row in df.iterrows():
        v = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
        if v:
            first_label = v
            break

    if 'TOTAIS DO LANÇAMENTO' in first_label:
        result['formato'] = 'novo'
        _parse_novo(df, result)
    else:
        result['formato'] = 'antigo'
        _parse_antigo(df, result)

    # Métricas derivadas
    gml  = result['gasto_ml']   or 0
    gctl = result['gasto_ctrl'] or 0
    rml  = result['receita_ml'] or 0
    rctl = result['receita_ctrl'] or 0
    mml  = result['margem_ml']  or 0
    mctl = result['margem_ctrl'] or 0

    # gasto_total: usa TOTAIS DO LANÇAMENTO (cobre TODAS as campanhas)
    result['gasto_total'] = result.get('gasto_all') or (gml + gctl)

    # receita_total para DISPLAY: usa TOTAIS extrapolado quando disponível
    result['receita_total'] = result.get('receita_all') or (rml + rctl)

    # receita para CÁLCULO DE MARGEM/GANHO: usa receita rastreada (ML + Controle)
    # roas_ctrl vem da seção de comparação (base rastreada). Misturar com receita
    # extrapolada do TOTAIS infla o ganho artificialmente — bases incompatíveis.
    tracked_receita = (rml or 0) + (rctl or 0)
    receita_calc = tracked_receita if tracked_receita > 0 else result['receita_total']
    result['margem_total'] = receita_calc - result['gasto_total']

    roas_ctrl = result['roas_ctrl']
    if roas_ctrl and roas_ctrl > 0 and result['gasto_total'] > 0:
        result['receita_cf'] = result['gasto_total'] * roas_ctrl
        result['margem_cf']  = result['receita_cf'] - result['gasto_total']
        result['ganho_margem'] = result['margem_total'] - result['margem_cf']
    else:
        result['receita_cf']   = None
        result['margem_cf']    = None
        result['ganho_margem'] = None

    result['pct_budget_ml'] = (gml / result['gasto_total'] * 100) if result['gasto_total'] > 0 else 100.0

    if result.get('roas_all') is not None:
        result['roas_total'] = result['roas_all']
    elif result['gasto_total'] > 0:
        result['roas_total'] = result['receita_total'] / result['gasto_total']
    else:
        result['roas_total'] = None

    return result


def _parse_novo(df: pd.DataFrame, result: dict):
    """Formato novo (18/03): seção 'COMPARAÇÃO ML vs CONTROLE' com labels em col A."""
    # --- Bloco "TOTAIS DO LANÇAMENTO" (primeiras linhas) ---
    # Estrutura:
    #   Título: 'TOTAIS DO LANÇAMENTO — CAMPANHAS'
    #   Linha de valores: [Leads, Conversões, Taxa, Gasto, CPL]
    #   Linha de labels:  ['Leads', 'Conversões', ...]
    #   Linha de valores: [Receita, ROAS, CPA, Margem, Ticket Médio]
    #   Linha de labels:  ['Receita', 'ROAS', ...]
    in_totais = False
    totais_first_vals = None  # [Leads, Conv, Taxa, Gasto, CPL]
    totais_second_vals = None  # [Receita, ROAS, CPA, Margem, Ticket]
    numeric_block_count = 0

    for _, row in df.iterrows():
        label = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
        if 'TOTAIS DO LANÇAMENTO' in label:
            in_totais = True
            numeric_block_count = 0
            continue
        if in_totais and ('COMPARAÇÃO ML' in label or 'ADSETS' in label.upper()):
            break
        if in_totais:
            # Verifica se a célula é realmente numérica (não string de label)
            raw = row.iloc[0]
            is_numeric = isinstance(raw, (int, float)) and not pd.isna(raw)
            if is_numeric:
                numeric_block_count += 1
                if numeric_block_count == 1:
                    totais_first_vals = [_to_float(row.iloc[i]) for i in range(min(5, len(row)))]
                elif numeric_block_count == 2:
                    totais_second_vals = [_to_float(row.iloc[i]) for i in range(min(5, len(row)))]
                    break

    if totais_first_vals and totais_second_vals:
        # totais_first_vals:  [Leads, Conversões, Taxa, Gasto, CPL]
        # totais_second_vals: [Receita, ROAS, CPA, Margem, Ticket Médio]
        result['receita_all'] = totais_second_vals[0] if totais_second_vals[0] else None
        result['roas_all']    = totais_second_vals[1] if totais_second_vals[1] else None
        result['margem_all']  = totais_second_vals[3] if totais_second_vals[3] else None
        result['gasto_all']   = totais_first_vals[3]  if totais_first_vals[3]  else None

    # --- Bloco "COMPARAÇÃO ML vs CONTROLE" ---
    in_comparacao = False
    for _, row in df.iterrows():
        label = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''

        if 'COMPARAÇÃO ML vs CONTROLE' in label or 'COMPARAÇÃO ML' in label:
            in_comparacao = True
            continue

        # Parar ao chegar na seção de Adsets
        if in_comparacao and ('ADSETS' in label.upper() or 'MATCHED' in label.upper()):
            break

        if not in_comparacao:
            continue

        v1 = _to_float(row.iloc[1])
        v2 = _to_float(row.iloc[2])

        if label == 'Gasto':
            result['gasto_ml'], result['gasto_ctrl'] = v1, v2
        elif label == 'Receita':
            result['receita_ml'], result['receita_ctrl'] = v1, v2
        elif label == 'ROAS':
            result['roas_ml'], result['roas_ctrl'] = v1, v2
        elif label == 'Margem':
            result['margem_ml'], result['margem_ctrl'] = v1, v2
        elif label == 'Conversões':
            result['conv_ml'], result['conv_ctrl'] = v1, v2


def _parse_antigo(df: pd.DataFrame, result: dict):
    """Formato antigo (09/03): labels descritivos, usa Receita Traqueada e ROAS Traqueado."""
    in_allvsall = True
    for _, row in df.iterrows():
        label = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
        if 'ADSETS MATCHED' in label.upper() or 'MATCHED PAIRS' in label.upper():
            break
        v1 = _to_float(row.iloc[1])
        v2 = _to_float(row.iloc[2])

        if 'Receita Total (Traqueada)' in label:
            result['receita_ml'], result['receita_ctrl'] = v1, v2
        elif 'Gasto Total' in label and 'Lançamento' not in label:
            result['gasto_ml'], result['gasto_ctrl'] = v1, v2
        elif 'ROAS (Traqueado)' in label:
            result['roas_ml'], result['roas_ctrl'] = v1, v2
        elif 'Margem Contribuição (Traqueada)' in label:
            result['margem_ml'], result['margem_ctrl'] = v1, v2


# ─────────────────────────────────────────────────────────────────────────────
# 3. FORMATTERS & MÉTRICAS
# ─────────────────────────────────────────────────────────────────────────────

def brl(v, default='—'):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return default
    return f"R$ {v:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")

def pct(v, default='—'):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return default
    return f"{v:.1f}%"

def roas_str(v, default='—'):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return default
    return f"{v:.2f}x"

METRICAS = [
    ('RECEITA & ROAS', None),
    ('Receita ML (R$)',          lambda m: brl(m.get('receita_ml'))),
    ('Receita Controle (R$)',    lambda m: brl(m.get('receita_ctrl'))),
    ('Receita Total (R$)',       lambda m: brl(m.get('receita_total'))),
    ('ROAS ML',                  lambda m: roas_str(m.get('roas_ml'))),
    ('ROAS Controle',            lambda m: roas_str(m.get('roas_ctrl'))),
    ('ROAS Total do Lançamento', lambda m: roas_str(m.get('roas_total'))),
    ('MARGEM DE CONTRIBUIÇÃO', None),
    ('Margem ML (R$)',           lambda m: brl(m.get('margem_ml'))),
    ('Margem Controle (R$)',     lambda m: brl(m.get('margem_ctrl'))),
    ('Margem Total (R$)',        lambda m: brl(m.get('margem_total'))),
    ('ANÁLISE CONTRAFACTUAL', None),
    ('Receita CF — e se todo spend fosse a ROAS Ctrl (R$)', lambda m: brl(m.get('receita_cf'), '* sem ctrl')),
    ('Margem CF (R$)',           lambda m: brl(m.get('margem_cf'), '* sem ctrl')),
    ('Ganho de Margem vs CF (R$)', lambda m: brl(m.get('ganho_margem'), '* sem ctrl')),
    ('ALOCAÇÃO DE BUDGET', None),
    ('Gasto ML (R$)',            lambda m: brl(m.get('gasto_ml'))),
    ('Gasto Controle (R$)',      lambda m: brl(m.get('gasto_ctrl'))),
    ('% Budget em ML',           lambda m: pct(m.get('pct_budget_ml'))),
    ('Gasto Total (R$)',         lambda m: brl(m.get('gasto_total'))),
]


# ─────────────────────────────────────────────────────────────────────────────
# 4. ADD SHEET 'Margem & Contrafactual'
# ─────────────────────────────────────────────────────────────────────────────

def add_margem_sheet(xlsx_path: Path, periods: list) -> None:
    """
    Adiciona/atualiza sheet 'Margem & Contrafactual' no xlsx de evolução.

    Args:
        xlsx_path: caminho para o arquivo xlsx de evolução
        periods: lista de dicts com 'name', 'vendas_start', 'vendas_end'
                 (formato de ml_evolution_report.PERIODS)
    """
    # Coletar dados de cada período
    data = {}
    for p in periods:
        name = p['name']
        vs = pd.Timestamp(p['vendas_start'])
        ve = pd.Timestamp(p['vendas_end'])
        folder = f"{vs.day:02d}:{vs.month:02d} - {ve.day:02d}:{ve.month:02d}"
        xlsx = get_latest_report(folder)
        if not xlsx:
            print(f"  {name}: sem relatório em {folder}")
            continue
        try:
            m = parse_comparacao_ml(xlsx)
            m['name'] = name
            data[name] = m
            ctrl_str = f"ctrl_ROAS={m['roas_ctrl']:.2f}" if m['roas_ctrl'] else "sem ctrl"
            print(f"  {name}: ML_ROAS={m['roas_ml']:.2f} | {ctrl_str} | margem_total=R${m['margem_total']:,.0f}")
        except Exception as e:
            print(f"  {name}: ERRO — {e}")

    lancamentos = [p['name'] for p in periods if p['name'] in data]
    if not lancamentos:
        print("  Nenhum dado disponível para 'Margem & Contrafactual'")
        return

    # Cores
    AZUL_HEADER  = PatternFill('solid', fgColor='1E3A5F')
    CINZA_BLOCO  = PatternFill('solid', fgColor='D0D7E3')
    VERDE_GANHO  = PatternFill('solid', fgColor='C6EFCE')
    VERMELHO_CF  = PatternFill('solid', fgColor='FFCCCC')

    font_header  = Font(bold=True, color='FFFFFF', size=11)
    font_bloco   = Font(bold=True, color='1E3A5F', size=10)
    font_normal  = Font(size=10)

    wb = load_workbook(xlsx_path)
    if 'Margem & Contrafactual' in wb.sheetnames:
        del wb['Margem & Contrafactual']
    ws = wb.create_sheet('Margem & Contrafactual')

    # Cabeçalho
    ws.cell(1, 1, 'MARGEM DE CONTRIBUIÇÃO & ANÁLISE CONTRAFACTUAL — DevClub').font = Font(bold=True, size=13, color='1E3A5F')
    ws.cell(2, 1, f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")} | Dados dos relatórios mais recentes de cada período').font = Font(size=9, italic=True, color='666666')
    ws.cell(3, 1, '* Contrafactual: e se todo o gasto do lançamento (ML + Controle) fosse alocado ao ROAS Controle?').font = Font(size=9, color='666666')

    # Linha de nomes de lançamentos (row 5)
    ws.cell(5, 1, 'Métrica').fill = AZUL_HEADER
    ws.cell(5, 1).font = font_header
    ws.cell(5, 1).alignment = Alignment(horizontal='left')

    for col_idx, name in enumerate(lancamentos, start=2):
        c = ws.cell(5, col_idx, name)
        c.fill = AZUL_HEADER
        c.font = font_header
        c.alignment = Alignment(horizontal='center')

    # Dados
    row = 6
    for metrica_label, fn in METRICAS:
        if fn is None:
            # Bloco header
            c = ws.cell(row, 1, metrica_label)
            c.fill = CINZA_BLOCO
            c.font = font_bloco
            c.alignment = Alignment(horizontal='left')
            for col_idx in range(2, len(lancamentos) + 2):
                ws.cell(row, col_idx).fill = CINZA_BLOCO
            row += 1
            continue

        ws.cell(row, 1, metrica_label).font = font_normal
        ws.cell(row, 1).alignment = Alignment(horizontal='left')

        for col_idx, name in enumerate(lancamentos, start=2):
            m = data.get(name, {})
            val = fn(m)
            c = ws.cell(row, col_idx, val)
            c.font = font_normal
            c.alignment = Alignment(horizontal='center')

            if metrica_label == 'Ganho de Margem vs CF (R$)' and val not in ('—', '* sem ctrl'):
                raw = m.get('ganho_margem')
                if raw is not None:
                    c.fill = VERDE_GANHO if raw > 0 else VERMELHO_CF
            elif metrica_label == 'Margem Total (R$)':
                raw = m.get('margem_total')
                if raw is not None and raw < 0:
                    c.fill = VERMELHO_CF
        row += 1

    # Larguras
    ws.column_dimensions['A'].width = 48
    for col_idx in range(2, len(lancamentos) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # Nota de rodapé
    row += 1
    ws.cell(row, 1, '* Sem ctrl = lançamentos sem campanhas controle com conversões. Contrafactual não calculável.').font = Font(size=8, italic=True, color='888888')

    wb.save(xlsx_path)
    print(f"  Sheet 'Margem & Contrafactual' salva em {xlsx_path.name}")


# ─────────────────────────────────────────────────────────────────────────────
# 5. GRÁFICOS (standalone)
# ─────────────────────────────────────────────────────────────────────────────

def _gerar_graficos(data: dict, lancamentos: list):
    plt.rcParams.update({'font.family': 'sans-serif', 'axes.spines.top': False, 'axes.spines.right': False})

    AZUL    = '#2563EB'
    CINZA   = '#94A3B8'
    LARANJA = '#F97316'
    VERDE   = '#10B981'
    VERM    = '#EF4444'

    com_ctrl = [n for n in lancamentos if data[n].get('roas_ctrl') and data[n]['roas_ctrl'] > 0]
    x_all  = np.arange(len(lancamentos))
    x_ctrl = np.arange(len(com_ctrl))

    # ── Gráfico 1: ROAS ML vs Controle vs Total ─────────────────────────────
    fig, ax = plt.subplots(figsize=(13, 6))
    roas_ml_vals   = [data[n].get('roas_ml')    or np.nan for n in lancamentos]
    roas_ctrl_vals = [data[n].get('roas_ctrl')  or np.nan for n in lancamentos]
    roas_tot_vals  = [data[n].get('roas_total') or np.nan for n in lancamentos]

    ax.plot(x_all, roas_ml_vals,   'o-', color=AZUL,   linewidth=2.2, markersize=8, label='ROAS ML')
    ax.plot(x_all, roas_ctrl_vals, 's--',color=CINZA,  linewidth=2.0, markersize=7, label='ROAS Controle')
    ax.plot(x_all, roas_tot_vals,  '^-', color=LARANJA,linewidth=2.0, markersize=7, label='ROAS Total Lançamento', alpha=0.85)

    for i, (v_ml, v_tot) in enumerate(zip(roas_ml_vals, roas_tot_vals)):
        if not np.isnan(v_ml):
            ax.annotate(f'{v_ml:.2f}x', (i, v_ml), textcoords='offset points', xytext=(0, 10), ha='center', fontsize=8.5, color=AZUL, fontweight='bold')
        if not np.isnan(v_tot):
            ax.annotate(f'{v_tot:.2f}x', (i, v_tot), textcoords='offset points', xytext=(0, -16), ha='center', fontsize=8.5, color=LARANJA)

    ax.axhline(1.0, color='red', linestyle=':', linewidth=1, alpha=0.5)
    ax.set_xticks(x_all)
    ax.set_xticklabels(lancamentos, fontsize=11)
    ax.set_ylabel('ROAS', fontsize=11)
    ax.set_title('ROAS por Lançamento: ML vs Controle vs Total do Negócio\n(ROAS Total = toda a receita / todo o gasto do lançamento)', fontsize=13, fontweight='bold', pad=10)
    ax.legend(fontsize=10)
    ax.grid(axis='y', alpha=0.3)
    valid_ml = [v for v in roas_ml_vals if not np.isnan(v)]
    if valid_ml:
        ax.set_ylim(0, max(valid_ml) * 1.3)

    plt.tight_layout()
    plt.savefig(GRAFICOS_DIR / '01_roas_ml_controle_total.png', dpi=180, bbox_inches='tight')
    plt.close()

    if not com_ctrl:
        return

    # ── Gráfico 2: Margem Real vs Contrafactual ──────────────────────────────
    fig, ax = plt.subplots(figsize=(12, 6))
    W = 0.35
    margem_real_cf = [data[n]['margem_total'] / 1000 for n in com_ctrl]
    margem_cf_vals = [data[n]['margem_cf']    / 1000 for n in com_ctrl]

    b1 = ax.bar(x_ctrl - W/2, margem_real_cf, W, label='Margem Real (ML + Ctrl)', color=AZUL, alpha=0.88)
    b2 = ax.bar(x_ctrl + W/2, margem_cf_vals, W, label='Margem Contrafactual (sem ML)', color=CINZA, alpha=0.80)

    for bar, v in zip(b1, margem_real_cf):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + (1 if v >= 0 else -8),
                f'R${v:.0f}k', ha='center', va='bottom', fontsize=8.5, color=AZUL, fontweight='bold')
    for bar, v in zip(b2, margem_cf_vals):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + (1 if v >= 0 else -8),
                f'R${v:.0f}k', ha='center', va='bottom', fontsize=8.5, color='#555')

    ax.axhline(0, color='black', linewidth=0.8)
    ax.set_xticks(x_ctrl)
    ax.set_xticklabels(com_ctrl, fontsize=11)
    ax.set_ylabel('Margem de Contribuição (R$ mil)', fontsize=11)
    ax.set_title('Margem Real vs Contrafactual\n"e se todo o gasto do lançamento fosse ao ROAS Controle?"', fontsize=13, fontweight='bold', pad=10)
    ax.legend(fontsize=10)
    ax.grid(axis='y', alpha=0.3)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f'R${v:.0f}k'))
    plt.tight_layout()
    plt.savefig(GRAFICOS_DIR / '02_margem_real_vs_contrafactual.png', dpi=180, bbox_inches='tight')
    plt.close()

    # ── Gráfico 3: Ganho de Margem absoluto ──────────────────────────────────
    fig, ax = plt.subplots(figsize=(12, 5.5))
    ganhos = [data[n]['ganho_margem'] / 1000 for n in com_ctrl]
    cores  = [VERDE if g > 0 else VERM for g in ganhos]

    bars = ax.bar(x_ctrl, ganhos, 0.55, color=cores, alpha=0.88)
    for bar, v in zip(bars, ganhos):
        va  = 'bottom' if v >= 0 else 'top'
        off = 1 if v >= 0 else -1
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + off,
                f'R${v:+.0f}k', ha='center', va=va, fontsize=9.5, fontweight='bold',
                color=bar.get_facecolor())

    total_ganho = sum(ganhos)
    ax.axhline(0, color='black', linewidth=0.8)
    ax.set_xticks(x_ctrl)
    ax.set_xticklabels(com_ctrl, fontsize=11)
    ax.set_ylabel('Ganho de Margem (R$ mil)', fontsize=11)
    ax.set_title(f'Ganho de Margem atribuído ao ML vs Contrafactual\nAcumulado: R${total_ganho:+.0f}k', fontsize=13, fontweight='bold', pad=10)
    ax.grid(axis='y', alpha=0.3)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f'R${v:+.0f}k'))
    plt.tight_layout()
    plt.savefig(GRAFICOS_DIR / '03_ganho_margem_vs_contrafactual.png', dpi=180, bbox_inches='tight')
    plt.close()

    # ── Gráfico 4: Budget ML% + ROAS Total ───────────────────────────────────
    fig, ax1 = plt.subplots(figsize=(13, 6))
    ax2 = ax1.twinx()

    gasto_ml_k   = [data[n]['gasto_ml']   / 1000 for n in lancamentos]
    gasto_ctrl_k = [data[n]['gasto_ctrl'] / 1000 if data[n]['gasto_ctrl'] else 0 for n in lancamentos]
    roas_tot     = [data[n]['roas_total'] or np.nan for n in lancamentos]

    b1 = ax1.bar(x_all, gasto_ml_k,   0.55, label='Gasto ML', color=AZUL, alpha=0.88)
    b2 = ax1.bar(x_all, gasto_ctrl_k, 0.55, bottom=gasto_ml_k, label='Gasto Controle', color=CINZA, alpha=0.80)

    ax2.plot(x_all, roas_tot, 'D--', color=LARANJA, linewidth=2.2, markersize=8, label='ROAS Total', zorder=5)
    for i, v in enumerate(roas_tot):
        if not np.isnan(v):
            ax2.annotate(f'{v:.2f}x', (i, v), textcoords='offset points', xytext=(0, 10),
                         ha='center', fontsize=8.5, color=LARANJA, fontweight='bold')

    ax1.set_xticks(x_all)
    ax1.set_xticklabels(lancamentos, fontsize=11)
    ax1.set_ylabel('Gasto (R$ mil)', fontsize=11)
    ax2.set_ylabel('ROAS Total', fontsize=11, color=LARANJA)
    ax2.tick_params(axis='y', labelcolor=LARANJA)
    valid_roas = [v for v in roas_tot if not np.isnan(v)]
    if valid_roas:
        ax2.set_ylim(0, max(valid_roas) * 1.5)

    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, fontsize=10, loc='upper left')
    ax1.set_title('Alocação de Budget (ML vs Controle) e ROAS Total do Lançamento', fontsize=13, fontweight='bold', pad=10)
    ax1.grid(axis='y', alpha=0.3)
    ax1.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f'R${v:.0f}k'))
    plt.tight_layout()
    plt.savefig(GRAFICOS_DIR / '04_budget_ml_vs_roas_total.png', dpi=180, bbox_inches='tight')
    plt.close()

    # ── Gráfico 5: Margem Total empilhada ML + Controle ──────────────────────
    fig, ax = plt.subplots(figsize=(13, 6))

    margem_ml_k   = [max(data[n]['margem_ml']   or 0, 0) / 1000 for n in lancamentos]
    margem_ctrl_k = [max(data[n]['margem_ctrl'] or 0, 0) / 1000 for n in lancamentos]
    margem_neg_k  = [min(data[n]['margem_total'] or 0, 0) / 1000 for n in lancamentos]

    ax.bar(x_all, margem_ml_k,   0.55, label='Margem ML', color=AZUL, alpha=0.90)
    ax.bar(x_all, margem_ctrl_k, 0.55, bottom=margem_ml_k, label='Margem Controle', color=CINZA, alpha=0.80)
    ax.bar(x_all, margem_neg_k,  0.55, color=VERM, alpha=0.70, label='Margem negativa')

    for i, n in enumerate(lancamentos):
        tot = (data[n]['margem_total'] or 0) / 1000
        ax.text(i, max(tot, 0) + 1, f'R${tot:.0f}k',
                ha='center', va='bottom', fontsize=8.5, fontweight='bold', color='#333')

    ax.axhline(0, color='black', linewidth=0.8)
    ax.set_xticks(x_all)
    ax.set_xticklabels(lancamentos, fontsize=11)
    ax.set_ylabel('Margem de Contribuição (R$ mil)', fontsize=11)
    ax.set_title('Evolução da Margem Total do Negócio por Lançamento\n(ML + Controle, rastreada)', fontsize=13, fontweight='bold', pad=10)
    ax.legend(fontsize=10)
    ax.grid(axis='y', alpha=0.3)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f'R${v:.0f}k'))
    plt.tight_layout()
    plt.savefig(GRAFICOS_DIR / '05_margem_total_evolucao.png', dpi=180, bbox_inches='tight')
    plt.close()

    print("Gráficos gerados:")
    for f in sorted(GRAFICOS_DIR.glob('0*.png')):
        print(f"  {f.name}")


# ─────────────────────────────────────────────────────────────────────────────
# 6. SÍNTESE EXECUTIVA
# ─────────────────────────────────────────────────────────────────────────────

# Critérios de flag de outlier (incluídos nos totais, mas identificados)
_OUTLIER_GASTO_MIN   = 35_000   # lançamentos muito pequenos → ruído
_OUTLIER_CTRL_ROAS   = 2.5      # controle sazonalmente alto → distorce baseline
_OUTLIER_GASTO_MAX   = 150_000  # lançamentos de escala atípica (ex: DEV19)
_OUTLIER_CONV_CTRL   = 10       # grupo Ctrl com < 10 conversões → estimativa instável


def _flags(m: dict) -> list[str]:
    flags = []
    gasto = m.get('gasto_total') or 0
    if gasto < _OUTLIER_GASTO_MIN:
        flags.append('baixo gasto')
    if gasto > _OUTLIER_GASTO_MAX:
        flags.append('escala atípica')
    ctrl = m.get('roas_ctrl')
    if ctrl and ctrl > _OUTLIER_CTRL_ROAS:
        flags.append('ctrl sazonal')
    conv_ctrl = m.get('conv_ctrl') or 0
    if ctrl and 0 < conv_ctrl < _OUTLIER_CONV_CTRL:
        flags.append('ctrl insuficiente')
    return flags


def add_sintese_sheet(xlsx_path: Path, data: dict, lancamentos: list) -> None:
    """
    Adiciona/atualiza sheet 'Síntese Executiva' no xlsx de evolução.

    Separa dois regimes:
      - A/B Verificado: lançamentos com grupo Controle simultâneo
      - ML-only Estimado: sem Controle; usa mediana histórica como baseline

    Args:
        xlsx_path:    caminho para o arquivo xlsx de evolução
        data:         dict {name: metrics} já calculado por add_margem_sheet
        lancamentos:  lista ordenada de nomes de lançamentos
    """
    if not data or not lancamentos:
        print("  Síntese: sem dados disponíveis")
        return

    # ── 1. Classificar lançamentos ────────────────────────────────────────────
    ab_launches   = []   # tem Controle real
    ml_only       = []   # sem Controle

    for name in lancamentos:
        m = data[name]
        has_ctrl = (m.get('roas_ctrl') or 0) > 0 and (m.get('gasto_ctrl') or 0) > 0
        if has_ctrl:
            ab_launches.append(name)
        else:
            ml_only.append(name)

    # ── 2. Baseline Control: mediana dos A/B sem nenhum flag ─────────────────────
    baseline_roas_vals = []
    for name in ab_launches:
        m = data[name]
        ctrl = m.get('roas_ctrl')
        if ctrl and not _flags(m):  # exclui baixo gasto, ctrl sazonal e escala atípica
            baseline_roas_vals.append(ctrl)

    baseline = float(np.median(baseline_roas_vals)) if baseline_roas_vals else 1.356

    # ── 3. Calcular ganho para ML-only usando baseline ─────────────────────────
    for name in ml_only:
        m = data[name]
        gt = m.get('gasto_total') or 0
        if gt > 0:
            m['roas_ctrl_proxy']  = baseline
            m['margem_cf']        = gt * baseline - gt
            m['ganho_margem']     = (m.get('margem_total') or 0) - m['margem_cf']
        else:
            m['roas_ctrl_proxy']  = None
            m['margem_cf']        = None
            m['ganho_margem']     = None

    # ── 4. Totais ─────────────────────────────────────────────────────────────
    def _sum(names, key):
        return sum(data[n].get(key) or 0 for n in names)

    def _agg(names):
        gasto  = _sum(names, 'gasto_total')
        ganho  = _sum(names, 'ganho_margem')
        wins   = sum(1 for n in names if (data[n].get('ganho_margem') or 0) > 0)
        return gasto, ganho, wins

    ab_gasto,  ab_ganho,  ab_wins  = _agg(ab_launches)
    ml_gasto,  ml_ganho,  ml_wins  = _agg(ml_only)
    tot_gasto = ab_gasto + ml_gasto
    tot_ganho = ab_ganho + ml_ganho

    # A/B sem outliers
    ab_clean = [n for n in ab_launches if not _flags(data[n])]
    ab_clean_gasto, ab_clean_ganho, ab_clean_wins = _agg(ab_clean)

    def _median_roas_ml(names):
        vals = [data[n].get('roas_ml') for n in names if data[n].get('roas_ml')]
        return float(np.median(vals)) if vals else None

    def _median_roas_ctrl(names):
        vals = [data[n].get('roas_ctrl') for n in names if data[n].get('roas_ctrl')]
        return float(np.median(vals)) if vals else None

    # ── 5. Estilos ─────────────────────────────────────────────────────────────
    AZUL_ESC   = PatternFill('solid', fgColor='1E3A5F')
    AZUL_CLAR  = PatternFill('solid', fgColor='D0E4F7')
    CINZA      = PatternFill('solid', fgColor='E8EAED')
    VERDE      = PatternFill('solid', fgColor='C6EFCE')
    AMARELO    = PatternFill('solid', fgColor='FFEB9C')
    VERMELHO   = PatternFill('solid', fgColor='FFCCCC')
    LARANJA    = PatternFill('solid', fgColor='FCE4D6')

    fnt_titulo  = Font(bold=True, size=13, color='1E3A5F')
    fnt_header  = Font(bold=True, size=10, color='FFFFFF')
    fnt_section = Font(bold=True, size=10, color='1E3A5F')
    fnt_normal  = Font(size=10)
    fnt_small   = Font(size=9, italic=True, color='666666')
    fnt_bold    = Font(bold=True, size=10)
    fnt_kpi     = Font(bold=True, size=14, color='1E3A5F')

    al_ctr = Alignment(horizontal='center', vertical='center')
    al_lft = Alignment(horizontal='left',   vertical='center')

    def _write(ws, row, col, val, font=None, fill=None, align=None):
        c = ws.cell(row, col, val)
        if font:  c.font  = font
        if fill:  c.fill  = fill
        if align: c.alignment = align
        return c

    def _header_row(ws, row, labels, fills=None):
        for i, lbl in enumerate(labels, 1):
            f = fills[i-1] if fills else AZUL_ESC
            _write(ws, row, i, lbl, font=fnt_header, fill=f, align=al_ctr)

    def _section_title(ws, row, title, ncols=8):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row, 1, title)
        c.font  = fnt_section
        c.fill  = AZUL_CLAR
        c.alignment = al_lft
        return row + 1

    def _brl(v):
        if v is None: return '—'
        sign = '+' if v > 0 else ''
        return f"{sign}R$ {v:,.0f}".replace(',', 'X').replace('.', ',').replace('X', '.')

    def _roas(v):
        return f'{v:.2f}x' if v else '—'

    def _cents(ganho, gasto):
        if not gasto: return '—'
        return f'{ganho/gasto*100:.0f}¢/R$1'

    def _pct(ganho, cf):
        if not cf or cf == 0: return '—'
        return f'+{ganho/cf*100:.0f}%'

    # ── 6. Criar/Recriar sheet ─────────────────────────────────────────────────
    wb = load_workbook(xlsx_path)
    if 'Síntese Executiva' in wb.sheetnames:
        del wb['Síntese Executiva']
    ws = wb.create_sheet('Síntese Executiva', 0)  # primeira aba

    # Larguras de coluna
    col_widths = [22, 10, 14, 12, 16, 14, 10, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ── Título ────────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    _write(ws, row, 1, 'SÍNTESE EXECUTIVA — Performance ML', font=fnt_titulo, align=al_lft)
    ws.row_dimensions[row].height = 24
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    _write(ws, row, 1,
           f'Atualizado em {datetime.now().strftime("%d/%m/%Y %H:%M")}  |  '
           f'{len(lancamentos)} lançamentos  |  Baseline Controle: {baseline:.3f}x',
           font=fnt_small, align=al_lft)
    row += 2

    # ── KPI Cards ─────────────────────────────────────────────────────────────
    row = _section_title(ws, row, 'RESUMO GERAL', ncols=8)
    kpis = [
        ('Ganho Total\n(A/B + estimado)', _brl(tot_ganho), VERDE),
        ('¢ extras/R$1\n(total)', _cents(tot_ganho, tot_gasto), VERDE),
        ('Ganho A/B\nverificado', _brl(ab_ganho), AZUL_CLAR),
        ('¢ extras/R$1\n(A/B)', _cents(ab_ganho, ab_gasto), AZUL_CLAR),
        ('Ganho A/B\n(sem outliers)', _brl(ab_clean_ganho), AZUL_CLAR),
        ('¢ extras/R$1\n(sem outliers)', _cents(ab_clean_ganho, ab_clean_gasto), AZUL_CLAR),
        ('Win rate\ngeral', f'{ab_wins + ml_wins}/{len(lancamentos)}', CINZA),
        ('Baseline\nControle', _roas(baseline), CINZA),
    ]
    ws.row_dimensions[row].height = 36
    ws.row_dimensions[row+1].height = 36
    for i, (lbl, val, fill) in enumerate(kpis, 1):
        _write(ws, row,   i, lbl, font=fnt_small, fill=fill, align=al_ctr)
        _write(ws, row+1, i, val, font=fnt_kpi,   fill=fill, align=al_ctr)
    row += 3

    # ── Tabela por Lançamento ─────────────────────────────────────────────────
    row = _section_title(ws, row, 'RESULTADO POR LANÇAMENTO', ncols=8)
    _header_row(ws, row, ['Lançamento', 'Tipo', 'Gasto', 'ROAS ML', 'ROAS Ctrl / Base', 'Ganho ML', 'Ganho %', 'Flags'])
    row += 1

    for name in lancamentos:
        m = data[name]
        flags = _flags(m)
        is_ab = name in ab_launches
        tipo  = 'A/B' if is_ab else 'ML-only'

        roas_ref = m.get('roas_ctrl') if is_ab else m.get('roas_ctrl_proxy')
        ganho    = m.get('ganho_margem')
        cf       = m.get('margem_cf')

        # Cor da linha
        if ganho is None:
            row_fill = CINZA
        elif ganho > 0:
            row_fill = VERDE if not flags else AMARELO
        else:
            row_fill = VERMELHO

        vals = [
            name,
            tipo,
            _brl(m.get('gasto_total')).replace('+', ''),
            _roas(m.get('roas_ml')),
            (_roas(roas_ref) + ('' if is_ab else ' *')),
            _brl(ganho),
            _pct(ganho, cf) if ganho and cf else '—',
            ', '.join(flags) if flags else '✓',
        ]
        for col, val in enumerate(vals, 1):
            _write(ws, row, col, val, font=fnt_normal, fill=row_fill, align=al_ctr if col > 1 else al_lft)
        row += 1

    # Nota proxy
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    _write(ws, row, 1, f'* ML-only: sem Controle simultâneo; contrafactual estimado usando baseline {baseline:.3f}x (mediana histórica)',
           font=fnt_small, align=al_lft)
    row += 2

    # ── Detalhe Regime A/B ────────────────────────────────────────────────────
    row = _section_title(ws, row, f'REGIME A/B — VERIFICADO  ({len(ab_launches)} lançamentos com Controle simultâneo)', ncols=8)
    ab_rows = [
        ('Gasto total',            _brl(ab_gasto).replace('+', '')),
        ('Margem CF total',        _brl(_sum(ab_launches, "margem_cf"))),
        ('Margem real total',      _brl(_sum(ab_launches, "margem_total"))),
        ('Ganho incremental',      _brl(ab_ganho)),
        ('¢ extras por R$1',       _cents(ab_ganho, ab_gasto)),
        ('ROAS ML mediana',        _roas(_median_roas_ml(ab_launches))),
        ('ROAS Ctrl mediana',      _roas(_median_roas_ctrl(ab_launches))),
        ('Razão ML/Ctrl mediana',  f'{_median_roas_ml(ab_launches)/_median_roas_ctrl(ab_launches):.2f}x'
                                   if _median_roas_ml(ab_launches) and _median_roas_ctrl(ab_launches) else '—'),
        ('Win rate',               f'{ab_wins}/{len(ab_launches)}'),
        ('— sem outliers ({}) —'.format(len(ab_clean)), ''),
        ('Gasto (sem outliers)',   _brl(ab_clean_gasto).replace('+', '')),
        ('Ganho (sem outliers)',   _brl(ab_clean_ganho)),
        ('¢ extras/R$1 (limpo)',   _cents(ab_clean_ganho, ab_clean_gasto)),
    ]
    for lbl, val in ab_rows:
        fill = AZUL_CLAR if lbl.startswith('—') else None
        _write(ws, row, 1, lbl,  font=fnt_bold if 'extras' in lbl else fnt_normal, fill=fill, align=al_lft)
        _write(ws, row, 2, val,  font=fnt_bold if 'extras' in lbl else fnt_normal, fill=fill, align=al_lft)
        row += 1
    row += 1

    # ── Detalhe Regime ML-only ────────────────────────────────────────────────
    if ml_only:
        row = _section_title(ws, row, f'REGIME ML-ONLY — ESTIMADO  ({len(ml_only)} lançamentos, baseline {baseline:.3f}x)', ncols=8)
        ml_rows = [
            ('Gasto total',        _brl(ml_gasto).replace('+', '')),
            ('Ganho estimado',     _brl(ml_ganho)),
            ('¢ extras por R$1',   _cents(ml_ganho, ml_gasto)),
            ('Win rate',           f'{ml_wins}/{len(ml_only)}'),
        ]
        for lbl, val in ml_rows:
            _write(ws, row, 1, lbl,  font=fnt_bold if 'extras' in lbl else fnt_normal, align=al_lft)
            _write(ws, row, 2, val,  font=fnt_bold if 'extras' in lbl else fnt_normal, align=al_lft)
            row += 1
        row += 1

    # ── Baseline detalhe ──────────────────────────────────────────────────────
    row = _section_title(ws, row, 'BASELINE CONTROLE (usado para estimativas ML-only)', ncols=8)
    _header_row(ws, row, ['Lançamento', 'ROAS Ctrl', 'Flag', '', '', '', '', ''],
                fills=[AZUL_ESC]*3 + [CINZA]*5)
    row += 1
    for name in ab_launches:
        m     = data[name]
        flags = _flags(m)
        ctrl  = m.get('roas_ctrl')
        used  = not _flags(m)  # sem nenhum flag → usado no baseline
        fill  = VERDE if used else AMARELO
        _write(ws, row, 1, name,           font=fnt_normal, fill=fill, align=al_lft)
        _write(ws, row, 2, _roas(ctrl),    font=fnt_normal, fill=fill, align=al_ctr)
        _write(ws, row, 3, ', '.join(flags) if flags else '✓ usado no baseline',
               font=fnt_normal, fill=fill, align=al_lft)
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    _write(ws, row, 1,
           f'Baseline = mediana dos ROAS Controle sem flags = {baseline:.4f}x  '
           f'(lançamentos usados: {", ".join(n for n in ab_launches if not _flags(data[n]))})',
           font=fnt_small, align=al_lft)
    row += 2

    # ── Detalhe Financeiro ML vs Controle ────────────────────────────────────
    row = _section_title(ws, row, 'DETALHE FINANCEIRO — ML vs CONTROLE', ncols=8)
    _header_row(ws, row, ['Lançamento', 'Receita ML', 'Receita Ctrl', 'Gasto ML', 'Gasto Ctrl', '% Budget ML', 'Margem ML', 'Margem Ctrl'])
    row += 1

    for name in lancamentos:
        m = data[name]
        has_ctrl = (m.get('roas_ctrl') or 0) > 0 and (m.get('gasto_ctrl') or 0) > 0
        rec_ml  = m.get('receita_ml') or 0
        rec_ct  = m.get('receita_ctrl') or 0
        g_ml    = m.get('gasto_ml') or 0
        g_ct    = m.get('gasto_ctrl') or 0
        mar_ml  = m.get('margem_ml') or 0
        mar_ct  = m.get('margem_ctrl') or 0
        pct_ml  = g_ml / (g_ml + g_ct) * 100 if (g_ml + g_ct) > 0 else 100.0

        def _brln(v): return brl(v) if v else '—'

        vals = [
            name,
            _brln(rec_ml),
            _brln(rec_ct) if has_ctrl else '—',
            brl(g_ml).replace('+', ''),
            brl(g_ct).replace('+', '') if has_ctrl else '—',
            f'{pct_ml:.0f}%',
            brl(mar_ml),
            brl(mar_ct) if has_ctrl else '—',
        ]
        for col, val in enumerate(vals, 1):
            _write(ws, row, col, val, font=fnt_normal,
                   align=al_ctr if col > 1 else al_lft)
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    _write(ws, row, 1,
           'Receita e Margem = valores rastreados (base matched). '
           'Tracking rate parcial → valores conservadores.',
           font=fnt_small, align=al_lft)

    wb.save(xlsx_path)
    print(f"  Sheet 'Síntese Executiva' salva em {xlsx_path.name}")


# ─────────────────────────────────────────────────────────────────────────────
# 7. MAIN (standalone)
# ─────────────────────────────────────────────────────────────────────────────

def _discover_periods_from_validation() -> list:
    """Auto-detecta períodos das pastas de validação (sem hardcode de datas)."""
    # PERIOD_NAMES: mapeamento vendas_start → nome do lançamento.
    # Carregado de configs/launches.yaml — adicionar novo LF lá, não aqui.
    def _load_period_names_local() -> dict:
        launches_path = BASE / 'configs' / 'launches.yaml'
        if not launches_path.exists():
            return {}
        with open(launches_path, 'r') as f:
            launches = yaml.safe_load(f)
        return {
            v['vendas_start']: name
            for name, v in launches.items()
            if isinstance(v, dict) and 'vendas_start' in v
        }
    PERIOD_NAMES = _load_period_names_local()
    validation_dir = BASE / 'outputs' / 'validation'
    SKIP = {'historico', 'cache', 'leads', 'arquivos_leads', 'feedback_loop', 'meta_features_test'}
    periods = []
    for folder in sorted(validation_dir.iterdir()):
        if not folder.is_dir() or folder.name in SKIP:
            continue
        if ':' in folder.name:
            reports = sorted(folder.glob('validation_report_*.xlsx'))
        else:
            reports = sorted(folder.glob('*.xlsx'))
        if not reports:
            continue
        for report in reports:
            try:
                pg = pd.read_excel(report, sheet_name='Performance Geral', header=None)
                rows = {
                    str(r[0]).strip(): str(r.iloc[1]).strip()
                    for _, r in pg.iterrows()
                    if pd.notna(r[0]) and len(r) > 1 and pd.notna(r.iloc[1])
                }
                cap_str = rows.get('Período de Captação', '')
                ven_str = rows.get('Período de Vendas', '')
                if ' a ' not in cap_str or ' a ' not in ven_str:
                    continue
                cap_start, cap_end       = [s.strip() for s in cap_str.split(' a ')]
                vendas_start, vendas_end = [s.strip() for s in ven_str.split(' a ')]
                name = PERIOD_NAMES.get(vendas_start, report.stem.split(' ')[0])
                periods.append({
                    'name':         name,
                    'cap_start':    cap_start,
                    'cap_end':      cap_end,
                    'vendas_start': vendas_start,
                    'vendas_end':   vendas_end,
                })
            except Exception as e:
                print(f"  Aviso: falha ao ler {report.name}: {e}")
    periods.sort(key=lambda p: p['vendas_start'])
    return periods


def update_sintese(xlsx_path: Path, periods: list) -> None:
    """
    Adiciona/atualiza a sheet 'Síntese Executiva' num xlsx já existente.

    Pensado para ser chamado diretamente de ml_evolution_report.run() logo
    após o build_excel(), sem precisar rodar gerar_evolucao_margem standalone.

    Args:
        xlsx_path: caminho para o arquivo evolucao_ml_devclub_*.xlsx recém-criado
        periods:   lista de dicts com keys name, cap_start, cap_end,
                   vendas_start, vendas_end (mesmo formato de PERIODS)
    """
    # Remover aba legada se existir
    try:
        from openpyxl import load_workbook as _lw
        _wb = _lw(xlsx_path)
        for _sheet in ['Margem & Contrafactual']:
            if _sheet in _wb.sheetnames:
                del _wb[_sheet]
        _wb.save(xlsx_path)
    except Exception as _e:
        print(f"  Aviso ao limpar abas legadas: {_e}")

    # Coletar dados de cada período
    data = {}
    for p in periods:
        vs = pd.Timestamp(p['vendas_start'])
        ve = pd.Timestamp(p['vendas_end'])
        folder = f"{vs.day:02d}:{vs.month:02d} - {ve.day:02d}:{ve.month:02d}"
        xlsx = get_latest_report(folder)
        if not xlsx:
            continue
        try:
            m = parse_comparacao_ml(xlsx)
            m['name'] = p['name']
            data[p['name']] = m
        except Exception:
            pass

    lancamentos = [p['name'] for p in periods if p['name'] in data]
    add_sintese_sheet(xlsx_path, data, lancamentos)
    _gerar_graficos(data, lancamentos)


def main():
    historico = BASE / 'outputs' / 'validation' / 'historico'
    candidates = sorted(historico.glob('evolucao_ml_devclub_*.xlsx'))
    if not candidates:
        print("Nenhum arquivo evolucao_ml_devclub_*.xlsx encontrado em historico/")
        sys.exit(1)
    xlsx_path = candidates[-1]
    print(f"Usando: {xlsx_path.name}\n")

    periods = _discover_periods_from_validation()
    print(f"Períodos auto-detectados: {[p['name'] for p in periods]}\n")

    update_sintese(xlsx_path, periods)

    print("\nConcluído.")


if __name__ == '__main__':
    main()
