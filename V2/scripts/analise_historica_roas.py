#!/usr/bin/env python3
"""
Análise histórica de ROAS e Margem de Contribuição — DevClub.

Compara métricas pré-ML (LF20–LF34, jul/2025+) vs pós-ML (LF40–LF47, dez/2025+).

Fontes de dados:
  - Meta Ads API          → gasto por período de captação
  - Guru + Hotmart + ASAS + TMB → receita por período de vendas (email+fone matched)
  - Relatórios validação  → métricas pós-ML (já calculadas)

Metodologia idêntica ao pipeline de validação:
  match_leads_to_sales_unified: email → telefone completo → últimos 6 dígitos

Uso:
  python V2/scripts/analise_historica_roas.py [--no-cache] [--output caminho.xlsx]
"""

import argparse
import json
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import pandas as pd
from dotenv import load_dotenv

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / 'src'))
load_dotenv(ROOT / '.env')

DATA_DIR       = ROOT / 'data' / 'devclub'
VALIDATION_DIR = ROOT / 'outputs' / 'validation'
LEADS_DIR      = ROOT / 'outputs' / 'validation' / 'arquivos_leads'
OUTPUT_DIR     = ROOT / 'outputs' / 'validation' / 'historico'
CACHE_DIR      = ROOT / 'files' / 'validation' / 'cache' / 'historico'

# Arquivo TMB Pedidos histórico (cobre 2022→mar/2026)
TMB_PEDIDOS_FILE = DATA_DIR / 'pedidos_03032026_1433.xlsx'

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
CACHE_DIR.mkdir(parents=True, exist_ok=True)

# ─── CONTAS META ──────────────────────────────────────────────────────────────
META_ACCOUNTS = [
    'act_188005769808959',  # Rodolfo Mori
    'act_786790755803474',  # Gestor de IA
]

# ─── LANÇAMENTOS PRÉ-ML ───────────────────────────────────────────────────────
# Descobertos automaticamente de outputs/validation/arquivos_leads/
# Nomes pós-ML (com relatório de validação) são excluídos.
POST_ML_LAUNCH_NAMES = {'LF40','LF41','LF42','DEV19','LF43','LF44','LF45','LF46','LF47'}

def _discover_pre_ml_launches() -> list[dict]:
    """
    Descobre lançamentos pré-ML a partir dos arquivos em LEADS_DIR.
    Deriva cap_start/cap_end do min/max da coluna DATA de cada arquivo.
    """
    launches = []
    for f in sorted(LEADS_DIR.glob('*LF*.xlsx')):
        name = f.stem.strip('[]').split(']')[0].strip('[ ')
        # Extrair nome limpo, ex: "[LF20] Leads" -> "LF20"
        import re
        m = re.match(r'\[?(LF\d+)\]?', f.stem)
        if not m:
            continue
        name = m.group(1)
        if name in POST_ML_LAUNCH_NAMES:
            continue
        try:
            df = pd.read_excel(f)
            df.columns = [c.strip() for c in df.columns]
            date_col = next((c for c in df.columns if c.upper() == 'DATA'), None)
            if date_col is None:
                continue
            dates = pd.to_datetime(df[date_col], errors='coerce').dropna()
            if dates.empty:
                continue
            cap_start = dates.min().strftime('%Y-%m-%d')
            cap_end   = dates.max().strftime('%Y-%m-%d')
            launches.append({'name': name, 'cap_start': cap_start, 'cap_end': cap_end})
        except Exception as e:
            print(f"  [AVISO] Não foi possível ler {f.name}: {e}")
    launches.sort(key=lambda x: x['cap_start'])
    return launches

# Mapeamento vendas_start → nome do lançamento pós-ML (mesmo de ml_evolution_report.py)
POST_ML_NAMES = {
    '2025-12-08': 'LF40',
    '2025-12-15': 'LF41',
    '2025-12-22': 'LF42',
    '2026-01-19': 'DEV19',
    '2026-02-02': 'LF43',
    '2026-02-09': 'LF44',
    '2026-03-02': 'LF45',
    '2026-03-09': 'LF46',
    '2026-03-16': 'LF47',
}

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def _vendas_range(cap_end: str) -> tuple[str, str]:
    """Período de vendas: cap_end + 7 dias → cap_end + 14 dias (padrão DevClub)."""
    end = datetime.strptime(cap_end, '%Y-%m-%d')
    v_start = end + timedelta(days=7)
    v_end   = end + timedelta(days=14)
    return v_start.strftime('%Y-%m-%d'), v_end.strftime('%Y-%m-%d')


CACHE_VERSION = 'v4'  # incrementar quando mudar fontes de dados


def _cache_path(prefix: str, start: str, end: str) -> Path:
    return CACHE_DIR / f"{prefix}_{start}_{end}_{CACHE_VERSION}.parquet"


def _load_leads(name: str) -> pd.DataFrame:
    """
    Carrega leads de outputs/validation/arquivos_leads/[{name}] Leads.xlsx.
    Retorna DataFrame com colunas normalizadas: email, telefone.
    """
    path = LEADS_DIR / f'[{name}] Leads.xlsx'
    if not path.exists():
        print(f"  [AVISO] Arquivo não encontrado: {path.name}")
        return pd.DataFrame(columns=['email', 'telefone'])
    df = pd.read_excel(path)
    df.columns = [c.strip() for c in df.columns]
    # Detectar colunas email e telefone (case-insensitive)
    email_col = next((c for c in df.columns if c.upper() in ('E-MAIL', 'EMAIL')), None)
    fone_col  = next((c for c in df.columns if c.upper() == 'TELEFONE'), None)
    if email_col is None:
        print(f"  [AVISO] Coluna de email não encontrada em {path.name}")
        return pd.DataFrame(columns=['email', 'telefone'])
    result = pd.DataFrame()
    result['email']    = df[email_col].dropna().str.lower().str.strip()
    result['telefone'] = df[fone_col].astype(str).str.strip() if fone_col else None
    result = result.dropna(subset=['email']).reset_index(drop=True)
    return result


def _norm_email(s) -> Optional[str]:
    if not s or (isinstance(s, float)):
        return None
    return str(s).lower().strip()


# ─── META SPEND ───────────────────────────────────────────────────────────────

def _get_meta_spend(cap_start: str, cap_end: str, use_cache: bool = True) -> float:
    """Soma o gasto Meta (campanhas CAP) para o período de captação."""
    cache = _cache_path('meta_spend', cap_start, cap_end)
    if use_cache and cache.exists():
        df = pd.read_parquet(cache)
        total = float(df['spend'].sum())
        print(f"  [Meta cache] {cap_start}→{cap_end}: R$ {total:,.0f}")
        return total

    from validation.meta_api_client import MetaAPIClient

    SPEND_COL = 'Valor usado (BRL)'
    total_spend = 0.0
    for account_id in META_ACCOUNTS:
        try:
            client = MetaAPIClient(account_id=account_id)
            df = client.get_campaigns(cap_start, cap_end, apply_filters=True)
            if not df.empty and SPEND_COL in df.columns:
                acct_spend = pd.to_numeric(df[SPEND_COL], errors='coerce').sum()
                total_spend += acct_spend
                print(f"  [Meta] {account_id}: R$ {acct_spend:,.0f}")
        except Exception as e:
            print(f"  [Meta ERRO] {account_id}: {e}")

    # Salvar cache como DataFrame simples
    cache_df = pd.DataFrame([{'spend': total_spend, 'cap_start': cap_start, 'cap_end': cap_end}])
    cache_df.to_parquet(cache)
    return total_spend


# ─── TMB PEDIDOS (histórico completo) ────────────────────────────────────────

_tmb_pedidos_df: Optional[pd.DataFrame] = None  # cache em memória


def _load_tmb_pedidos_for_period(vendas_start: str, vendas_end: str) -> pd.DataFrame:
    """
    Filtra vendas TMB do arquivo histórico para o período de vendas.

    O arquivo pedidos_03032026_1433.xlsx cobre 2022→mar/2026.
    Status válido: qualquer status exceto 'Cancelado' / 'Cancelamento solicitado'.
    """
    global _tmb_pedidos_df

    if not TMB_PEDIDOS_FILE.exists():
        return pd.DataFrame()

    # Carregar arquivo inteiro apenas uma vez (cache em memória)
    if _tmb_pedidos_df is None:
        raw = pd.read_excel(
            TMB_PEDIDOS_FILE,
            usecols=['E-mail do Cliente', 'Telefone do Cliente', 'Ticket do pedido',
                     'Data Efetivado', 'Criado em', 'Status Pedido'],
        )
        raw['email']     = raw['E-mail do Cliente'].apply(_norm_email)
        raw['telefone']  = raw['Telefone do Cliente'].astype(str).str.strip()
        raw['sale_value'] = pd.to_numeric(raw['Ticket do pedido'], errors='coerce')
        raw['sale_date']  = pd.to_datetime(raw['Data Efetivado'], errors='coerce')
        raw['sale_date']  = raw['sale_date'].fillna(
            pd.to_datetime(raw['Criado em'], errors='coerce')
        )
        # Excluir cancelados
        raw = raw[~raw['Status Pedido'].isin(['Cancelado', 'Cancelamento solicitado'])]
        raw = raw.dropna(subset=['email', 'sale_value', 'sale_date'])
        raw['origem'] = 'tmb'
        _tmb_pedidos_df = raw[['email', 'telefone', 'sale_value', 'sale_date', 'origem']].copy()

    start = pd.Timestamp(vendas_start)
    end   = pd.Timestamp(vendas_end) + pd.Timedelta(days=1)
    period = _tmb_pedidos_df[
        (_tmb_pedidos_df['sale_date'] >= start) &
        (_tmb_pedidos_df['sale_date'] <  end)
    ].copy()

    return period[['email', 'telefone', 'sale_value', 'sale_date', 'origem']]


# ─── RECEITA (GURU + HOTMART + ASAS) ─────────────────────────────────────────

def _get_all_sales(vendas_start: str, vendas_end: str, use_cache: bool = True) -> pd.DataFrame:
    """
    Busca todas as vendas de Guru, Hotmart, ASAS e TMB para o período.

    Retorna DataFrame com colunas: email, telefone, sale_value, sale_date, origem.
    Sem deduplicação — match_leads_to_sales_unified trata múltiplas vendas por lead.
    """
    cache = _cache_path('sales_all', vendas_start, vendas_end)
    if use_cache and cache.exists():
        df = pd.read_parquet(cache)
        print(f"  [Sales cache] {vendas_start}→{vendas_end}: {len(df)} vendas")
        return df

    frames = []

    # Guru
    try:
        from validation.guru_sales_extractor import fetch_guru_sales_from_api
        guru_raw = fetch_guru_sales_from_api(vendas_start, vendas_end)
        if not guru_raw.empty:
            mask_ok = pd.Series([True] * len(guru_raw), index=guru_raw.index)
            if 'status' in guru_raw.columns:
                mask_ok = guru_raw['status'].isin(['Aprovada', 'approved'])
            guru_raw = guru_raw[mask_ok]
            # Combinar DDD + número para telefone completo
            ddd = guru_raw.get('codigo telefone contato', pd.Series(dtype=str)).fillna('').astype(str).str.strip()
            num = guru_raw.get('telefone contato',        pd.Series(dtype=str)).fillna('').astype(str).str.strip()
            guru = pd.DataFrame({
                'email':      guru_raw['email contato'].apply(_norm_email),
                'telefone':   (ddd + num).str.strip().replace('', None),
                'sale_value': pd.to_numeric(guru_raw['valor venda'], errors='coerce'),
                'sale_date':  pd.to_datetime(guru_raw.get('data aprovacao'), errors='coerce'),
                'origem':     'guru',
            })
            guru = guru.dropna(subset=['email', 'sale_value'])
            frames.append(guru)
            print(f"  [Guru] {len(guru)} vendas")
    except Exception as e:
        print(f"  [Guru ERRO] {e}")

    # Hotmart (telefone=None — API não fornece)
    try:
        from validation.data_loader import SalesDataLoader
        loader = SalesDataLoader()
        hotmart = loader.load_hotmart_sales_from_api(vendas_start, vendas_end)
        if not hotmart.empty:
            hotmart = hotmart[['email', 'telefone', 'sale_value', 'sale_date', 'origem']].copy()
            hotmart['email']      = hotmart['email'].apply(_norm_email)
            hotmart['sale_value'] = pd.to_numeric(hotmart['sale_value'], errors='coerce')
            hotmart = hotmart.dropna(subset=['email', 'sale_value'])
            frames.append(hotmart)
            print(f"  [Hotmart] {len(hotmart)} vendas")
    except Exception as e:
        print(f"  [Hotmart ERRO] {e}")

    # ASAS
    try:
        from validation.asaas_sales_extractor import fetch_asaas_sales
        asaas = fetch_asaas_sales(vendas_start, vendas_end)
        if not asaas.empty:
            asaas = asaas[['email', 'telefone', 'sale_value', 'sale_date', 'origem']].copy()
            asaas['email']      = asaas['email'].apply(_norm_email)
            asaas['sale_value'] = pd.to_numeric(asaas['sale_value'], errors='coerce')
            asaas = asaas.dropna(subset=['email', 'sale_value'])
            frames.append(asaas)
            print(f"  [ASAS] {len(asaas)} vendas")
    except Exception as e:
        print(f"  [ASAS ERRO] {e}")

    # TMB Pedidos (arquivo histórico cobre 2022→mar/2026)
    try:
        tmb = _load_tmb_pedidos_for_period(vendas_start, vendas_end)
        if not tmb.empty:
            frames.append(tmb)
            print(f"  [TMB] {len(tmb)} vendas")
    except Exception as e:
        print(f"  [TMB ERRO] {e}")

    if not frames:
        return pd.DataFrame(columns=['email', 'telefone', 'sale_value', 'sale_date', 'origem'])

    all_sales = pd.concat(frames, ignore_index=True)
    # Garantir coluna telefone mesmo se alguma fonte não a tiver
    if 'telefone' not in all_sales.columns:
        all_sales['telefone'] = None
    if 'sale_date' not in all_sales.columns:
        all_sales['sale_date'] = pd.NaT

    all_sales.to_parquet(cache)
    return all_sales


# ─── MÉTRICAS PRÉ-ML ─────────────────────────────────────────────────────────

def calc_pre_ml(launch: dict, use_cache: bool = True) -> dict:
    """Calcula ROAS e Margem para um lançamento pré-ML via email+telefone matching."""
    from core.matching import match_leads_to_sales_unified

    name      = launch['name']
    cap_start = launch['cap_start']
    cap_end   = launch['cap_end']
    v_start, v_end = _vendas_range(cap_end)

    print(f"\n{'─'*60}")
    print(f"[{name}] cap {cap_start}→{cap_end} | vendas {v_start}→{v_end}")

    # 1. Leads captados
    leads_df = _load_leads(name)
    n_leads  = len(leads_df)
    print(f"  Leads captados: {n_leads:,}")

    # 2. Gasto Meta
    gasto = _get_meta_spend(cap_start, cap_end, use_cache)
    print(f"  Gasto Meta: R$ {gasto:,.0f}")

    # 3. Vendas brutas do período (todas as fontes, sem dedup)
    all_sales = _get_all_sales(v_start, v_end, use_cache)
    n_vendas_bruto = len(all_sales)
    print(f"  Vendas brutas no período: {n_vendas_bruto}")

    # 4. Matching leads → vendas (email → telefone completo → últimos 6 dígitos)
    if leads_df.empty or all_sales.empty:
        receita  = 0.0
        n_vendas = 0
        pct_tracking = 0.0
    else:
        matched = match_leads_to_sales_unified(leads_df, all_sales, mode='validation')
        converted = matched[matched['converted'] == True]
        receita   = float(converted['sale_value'].sum())
        n_vendas  = len(converted)
        pct_tracking = n_vendas / n_vendas_bruto if n_vendas_bruto > 0 else 0.0
        print(f"  Matched: {n_vendas} vendas | Receita: R$ {receita:,.0f} | Tracking: {pct_tracking:.1%}")

    roas   = receita / gasto if gasto > 0 else None
    margem = receita - gasto if gasto > 0 else None

    return {
        'name':         name,
        'cap_start':    cap_start,
        'cap_end':      cap_end,
        'vendas_start': v_start,
        'vendas_end':   v_end,
        'n_leads':      n_leads,
        'gasto':        gasto,
        'receita':      receita,
        'n_vendas':     n_vendas,
        'pct_tracking': pct_tracking,
        'roas':         roas,
        'margem':       margem,
        'is_ml':        False,
    }


# ─── MÉTRICAS PÓS-ML (validação existente) ───────────────────────────────────

def _load_post_ml_launches() -> list[dict]:
    """
    Extrai métricas dos relatórios de validação existentes.

    Usa parse_comparacao_ml (de gerar_evolucao_margem.py) para ler
    gasto_total, receita_total, roas_total, margem_total de cada report.
    """
    import importlib.util

    # Importar parse_comparacao_ml do script irmão
    margem_script = Path(__file__).parent / 'gerar_evolucao_margem.py'
    spec = importlib.util.spec_from_file_location('gerar_evolucao_margem', margem_script)
    mod  = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    parse_fn = mod.parse_comparacao_ml

    results = []
    for folder in sorted(VALIDATION_DIR.iterdir()):
        if not folder.is_dir() or ':' not in folder.name:
            continue
        reports = sorted(folder.glob('validation_report_*.xlsx'))
        if not reports:
            continue
        xlsx = reports[-1]  # mais recente

        # Ler datas e estatísticas da aba Performance Geral
        try:
            pg   = pd.read_excel(xlsx, sheet_name='Performance Geral', header=None)
            rows = {}
            for _, r in pg.iterrows():
                if pd.notna(r.iloc[0]) and len(r) > 1 and pd.notna(r.iloc[1]):
                    rows[str(r.iloc[0]).strip()] = r.iloc[1]
            cap_str = str(rows.get('Período de Captação', ''))
            ven_str = str(rows.get('Período de Vendas', ''))
            if ' a ' not in cap_str or ' a ' not in ven_str:
                continue
            cap_start, cap_end       = [s.strip() for s in cap_str.split(' a ')]
            vendas_start, vendas_end = [s.strip() for s in ven_str.split(' a ')]
            # Estatísticas de leads/vendas
            n_leads_val      = rows.get('Leads Meta') or rows.get('Leads')
            n_vendas_bruto   = rows.get('Vendas no Período') or rows.get('Vendas no período')
            n_vendas_matched = rows.get('Vendas identificadas')
            pct_track_raw    = rows.get('% de trackeamento') or rows.get('% Trackeamento')
            n_leads_int      = int(n_leads_val)      if pd.notna(n_leads_val)      else None
            n_vendas_int     = int(n_vendas_matched) if pd.notna(n_vendas_matched) else None
            pct_track_val    = float(pct_track_raw)  if pd.notna(pct_track_raw)    else None
        except Exception:
            continue

        # Ler métricas de negócio
        try:
            m = parse_fn(xlsx)
        except Exception as e:
            print(f"  [AVISO] parse_comparacao_ml falhou para {xlsx.name}: {e}")
            continue

        if not m.get('gasto_total'):
            continue

        name = POST_ML_NAMES.get(vendas_start, folder.name)
        results.append({
            'name':         name,
            'cap_start':    cap_start,
            'cap_end':      cap_end,
            'vendas_start': vendas_start,
            'vendas_end':   vendas_end,
            'n_leads':      n_leads_int,
            'gasto':        m.get('gasto_total', 0),
            'receita':      m.get('receita_total', 0),
            'n_vendas':     n_vendas_int,
            'pct_tracking': pct_track_val,
            'roas':         m.get('roas_total'),
            'margem':       m.get('margem_total'),
            'is_ml':        True,
        })
        print(f"  [{name}] ROAS={m.get('roas_total', 0):.2f} | Margem=R$ {m.get('margem_total', 0):,.0f}")

    results.sort(key=lambda r: r['vendas_start'])
    return results


# ─── SAÍDA EXCEL ──────────────────────────────────────────────────────────────

def _write_excel(df: pd.DataFrame, output_path: Path):
    """Gera planilha com tabela + gráficos de evolução."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ROAS & Margem Histórico'

    # Paleta
    ML_FILL   = PatternFill('solid', fgColor='1F4E79')   # azul escuro = ML
    PRE_FILL  = PatternFill('solid', fgColor='C6EFCE')   # verde claro  = pré-ML
    HDR_FILL  = PatternFill('solid', fgColor='2F5496')
    HDR_FONT  = Font(color='FFFFFF', bold=True)
    BOLD      = Font(bold=True)
    CENTER    = Alignment(horizontal='center')
    thin      = Side(style='thin', color='AAAAAA')
    BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Cabeçalho
    headers = [
        'Lançamento', 'Captação Início', 'Captação Fim',
        'Vendas Início', 'Vendas Fim',
        'Leads', 'Gasto (R$)', 'Receita (R$)', 'Vendas',
        'Tracking%', 'ROAS', 'Margem (R$)', 'ML?',
    ]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill    = HDR_FILL
        cell.font    = HDR_FONT
        cell.alignment = CENTER
        cell.border  = BORDER

    # Dados
    for _, row in df.iterrows():
        pct = row.get('pct_tracking')
        data_row = [
            row['name'],
            row['cap_start'],
            row['cap_end'],
            row['vendas_start'],
            row['vendas_end'],
            int(row['n_leads']) if pd.notna(row.get('n_leads')) else '',
            round(row['gasto'], 2)   if pd.notna(row.get('gasto'))   else '',
            round(row['receita'], 2) if pd.notna(row.get('receita')) else '',
            int(row['n_vendas'])     if pd.notna(row.get('n_vendas')) else '',
            round(float(pct), 4)     if pd.notna(pct)                else '',
            round(row['roas'], 3)    if pd.notna(row.get('roas'))     else '',
            round(row['margem'], 2)  if pd.notna(row.get('margem'))   else '',
            'Sim' if row['is_ml'] else 'Não',
        ]
        ws.append(data_row)
        row_idx = ws.max_row
        fill = ML_FILL if row['is_ml'] else PRE_FILL
        font = Font(color='FFFFFF', bold=True) if row['is_ml'] else Font()
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill   = fill
            cell.font   = font
            cell.border = BORDER
            cell.alignment = CENTER

    # Formatação de colunas
    col_widths = [10, 16, 14, 16, 14, 10, 16, 16, 10, 12, 10, 16, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Formatar coluna Tracking% como percentual
    TRACKING_COL = 10  # coluna J (0-indexed: posição 10 na lista de headers)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=TRACKING_COL)
        if cell.value != '':
            cell.number_format = '0.0%'

    # ── Legenda ──────────────────────────────────────────────────────────────
    n_data = len(df)
    ws.cell(row=n_data + 3, column=1,  value='Legenda:').font = BOLD
    ws.cell(row=n_data + 4, column=1,  value='Pré-ML').fill  = PRE_FILL
    ws.cell(row=n_data + 4, column=2,  value='Lançamentos sem sistema de ML (ROAS calculado via email+fone match)')
    ws.cell(row=n_data + 5, column=1,  value='Pós-ML').fill  = ML_FILL
    ws.cell(row=n_data + 5, column=1).font = Font(color='FFFFFF')
    ws.cell(row=n_data + 5, column=2,  value='Lançamentos com ML ativo (dados dos relatórios de validação)')

    # ── Aba de gráficos ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Gráficos')

    # Preparar dados para charts (só lançamentos com ROAS)
    df_chart = df[df['roas'].notna()].reset_index(drop=True)
    n_chart = len(df_chart)

    if n_chart > 0:
        # Gravar dados auxiliares para referência dos charts
        ws2.cell(1, 1, 'Lançamento').font = BOLD
        ws2.cell(1, 2, 'ROAS').font       = BOLD
        ws2.cell(1, 3, 'Margem (R$)').font = BOLD
        ws2.cell(1, 4, 'ML').font          = BOLD

        for i, (_, r) in enumerate(df_chart.iterrows(), 2):
            ws2.cell(i, 1, r['name'])
            ws2.cell(i, 2, round(r['roas'], 3)   if pd.notna(r['roas'])   else None)
            ws2.cell(i, 3, round(r['margem'], 0) if pd.notna(r['margem']) else None)
            ws2.cell(i, 4, 1 if r['is_ml'] else 0)

        # Gráfico ROAS
        roas_chart = LineChart()
        roas_chart.title  = 'Evolução do ROAS por Lançamento (Pré-ML vs Pós-ML)'
        roas_chart.y_axis.title = 'ROAS'
        roas_chart.x_axis.title = 'Lançamento'
        roas_chart.width  = 30
        roas_chart.height = 16

        roas_data = Reference(ws2, min_col=2, min_row=1, max_row=n_chart + 1)
        cats      = Reference(ws2, min_col=1, min_row=2, max_row=n_chart + 1)
        roas_chart.add_data(roas_data, titles_from_data=True)
        roas_chart.set_categories(cats)
        roas_chart.series[0].graphicalProperties.line.solidFill   = '2196F3'
        roas_chart.series[0].graphicalProperties.line.width       = 20000
        roas_chart.series[0].marker.symbol = 'circle'
        roas_chart.series[0].marker.size   = 6
        ws2.add_chart(roas_chart, 'F1')

        # Gráfico Margem
        margem_chart = BarChart()
        margem_chart.title   = 'Margem de Contribuição por Lançamento'
        margem_chart.y_axis.title = 'Margem (R$)'
        margem_chart.x_axis.title = 'Lançamento'
        margem_chart.width   = 30
        margem_chart.height  = 16
        margem_chart.type    = 'col'

        margem_data = Reference(ws2, min_col=3, min_row=1, max_row=n_chart + 1)
        margem_chart.add_data(margem_data, titles_from_data=True)
        margem_chart.set_categories(cats)
        # Colorir barras pós-ML de azul escuro
        for i, (_, r) in enumerate(df_chart.iterrows()):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = '1F4E79' if r['is_ml'] else '70AD47'
            margem_chart.series[0].dPt.append(pt)
        ws2.add_chart(margem_chart, 'F32')

    # ── Aba Resumo ───────────────────────────────────────────────────────────
    ws3 = wb.create_sheet('Resumo Comparativo')

    ws3.cell(1, 1, 'RESUMO: PRÉ-ML vs PÓS-ML').font = Font(bold=True, size=14)

    for group, label, col in [('pre_ml', 'Pré-ML (sem sistema)', 1), ('post_ml', 'Pós-ML (com ML ativo)', 4)]:
        subset = df[df['is_ml'] == (group == 'post_ml')]
        subset = subset[subset['roas'].notna() & subset['gasto'].notna() & subset['gasto'] > 0]

        ws3.cell(3, col, label).font = Font(bold=True)
        ws3.cell(4, col, 'Lançamentos')
        ws3.cell(4, col + 1, len(subset))
        ws3.cell(5, col, 'ROAS Médio')
        ws3.cell(5, col + 1, round(subset['roas'].mean(), 2) if len(subset) > 0 else '-')
        ws3.cell(6, col, 'ROAS Mediano')
        ws3.cell(6, col + 1, round(subset['roas'].median(), 2) if len(subset) > 0 else '-')
        ws3.cell(7, col, 'Margem Total (R$)')
        ws3.cell(7, col + 1, round(subset['margem'].sum(), 0) if len(subset) > 0 else '-')
        ws3.cell(8, col, 'Margem Média / lançamento (R$)')
        ws3.cell(8, col + 1, round(subset['margem'].mean(), 0) if len(subset) > 0 else '-')
        ws3.cell(9, col, 'Gasto Total (R$)')
        ws3.cell(9, col + 1, round(subset['gasto'].sum(), 0) if len(subset) > 0 else '-')
        ws3.cell(10, col, 'Receita Total (R$)')
        ws3.cell(10, col + 1, round(subset['receita'].sum(), 0) if len(subset) > 0 else '-')

    # Delta
    pre  = df[df['is_ml'] == False]
    post = df[df['is_ml'] == True]
    pre  = pre[pre['roas'].notna()  & (pre['gasto'] > 0)]
    post = post[post['roas'].notna() & (post['gasto'] > 0)]

    if len(pre) > 0 and len(post) > 0:
        delta_roas   = post['roas'].mean()   - pre['roas'].mean()
        delta_margem = post['margem'].mean() - pre['margem'].mean()
        ws3.cell(12, 1, 'Delta ROAS médio (pós - pré)').font = Font(bold=True)
        ws3.cell(12, 2, round(delta_roas, 2))
        ws3.cell(13, 1, 'Delta Margem média/lançamento (R$)').font = Font(bold=True)
        ws3.cell(13, 2, round(delta_margem, 0))

    for col_idx in [1, 2, 4, 5]:
        ws3.column_dimensions[get_column_letter(col_idx)].width = 35 if col_idx % 3 == 1 else 18

    wb.save(output_path)
    print(f"\n Planilha salva em: {output_path}")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def run(use_cache: bool = True, output_path: Optional[Path] = None):
    print("=" * 60)
    print("ANÁLISE HISTÓRICA ROAS & MARGEM — DevClub")
    print("=" * 60)

    # 1. Lançamentos pré-ML
    print("\n[1/3] Calculando métricas pré-ML...")
    pre_ml_launches = _discover_pre_ml_launches()
    print(f"  {len(pre_ml_launches)} lançamentos pré-ML encontrados: "
          f"{pre_ml_launches[0]['name']} → {pre_ml_launches[-1]['name']}")
    pre_ml_rows = []
    for launch in pre_ml_launches:
        row = calc_pre_ml(launch, use_cache=use_cache)
        pre_ml_rows.append(row)

    # 2. Lançamentos pós-ML
    print("\n[2/3] Carregando métricas pós-ML dos relatórios de validação...")
    post_ml_rows = _load_post_ml_launches()

    # 3. Combinar e exportar
    print("\n[3/3] Gerando relatório...")
    all_rows = pre_ml_rows + post_ml_rows
    df = pd.DataFrame(all_rows).sort_values('cap_start').reset_index(drop=True)

    if output_path is None:
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = OUTPUT_DIR / f'historico_roas_devclub_{ts}.xlsx'

    _write_excel(df, output_path)

    # Resumo no console
    print("\n" + "=" * 60)
    print("RESUMO")
    print("=" * 60)
    pre  = df[df['is_ml'] == False].query('roas.notna() and gasto > 0', engine='python')
    post = df[df['is_ml'] == True].query('roas.notna() and gasto > 0', engine='python')
    if len(pre) > 0:
        print(f"Pré-ML  ({len(pre)} lançamentos): ROAS médio = {pre['roas'].mean():.2f} | "
              f"Margem média = R$ {pre['margem'].mean():,.0f}")
    if len(post) > 0:
        print(f"Pós-ML  ({len(post)} lançamentos): ROAS médio = {post['roas'].mean():.2f} | "
              f"Margem média = R$ {post['margem'].mean():,.0f}")

    return df


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Análise histórica ROAS & Margem DevClub')
    parser.add_argument('--no-cache',  action='store_true', help='Ignorar cache e re-buscar todas as APIs')
    parser.add_argument('--output',    type=str,            help='Caminho do arquivo de saída (.xlsx)')
    args = parser.parse_args()

    output = Path(args.output) if args.output else None
    run(use_cache=not args.no_cache, output_path=output)
