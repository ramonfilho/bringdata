"""
compute_top5_roas_attributable.py — TOP 5 ROAS atribuível 60d por janela de
captação. Fonte canônica do projeto.

DEFINIÇÃO OPERACIONAL
─────────────────────
Para cada lançamento (LF) definido em configs/launches.yaml:

  ROAS_60d(LF) = receita_atribuível(LF) / spend(LF)

  receita_atribuível(LF) = soma do valor das vendas em que:
    - O comprador (matching por email ou telefone normalizado) coincide com
      um lead capturado pertencente a esse LF
    - sale_date ∈ (lead.capture_date, lead.capture_date + 60 dias]
    - O lead pertence a esse LF segundo a regra de **primeiro LF**: cada
      identidade (email ou telefone) é atribuída ao LF de captação mais
      antigo em que aparece. Evita double-count entre lançamentos.

  spend(LF) = "Gasto Total" da aba Performance Geral do relatório
  individual mais recente em outputs/validation/YYYY-MM/LFxx*.xlsx
  (gerado pelo time DevClub via Meta Insights API, filtro proprietário).

FONTES DE DADOS
───────────────
  Leads (preferência por LF):
    - Se outputs/validation/arquivos_leads/[LFxx] Leads.xlsx existe → usa esse
      (cobertura completa do LF, exportado pelo time DevClub).
    - Senão → files/validation/cache/railway_leads_*.parquet (LF49+ atualmente).
    O parquet do Railway só tem dados ≥ 2026-02-18, então LF43-LF44 e parte
    do LF45 dependem do legacy. Verificado em 2026-05-14.
  Vendas (4 plataformas consolidadas):
    - Guru:    outputs/cache/raw_data_latest.pkl  (status='Aprovada')
    - Hotmart: idem                                (status='Aprovada')
    - Asaas:   outputs/analysis/asaas_realized.parquet
    - TMB:     data/devclub/pedidos_*.xlsx        (Data Efetivado not null
                                                   AND Data Cancelado null,
                                                   valor = Ticket × 1/12)
  Spend:
    - outputs/validation/YYYY-MM/LFxx*.xlsx  (Performance Geral · Gasto Total)
    - Pega o arquivo mais recente por LF (max mtime)

NORMALIZAÇÃO PRA MATCHING
─────────────────────────
  Email:    lowercase, strip
  Telefone: dígitos-só; se começa com 55 e tem ≥12 dígitos, remove 55;
            mantém últimos 10-11 dígitos

CAVEATS
───────
  - Univariado: cada (lead, venda) é match por email OU telefone — não usa
    fuzzy matching de nome. Lead com email/telefone errado fica unmatched.
  - Janela 60d: hardcoded. Lançamentos com ciclo de venda mais longo são
    sub-estimados; mais curto, super-estimados. Default escolhido por
    cobrir 1 lançamento + 1 follow-up típico DevClub.
  - Spend filtrado vs raw Meta: o "Gasto Total" do xlsx individual passa
    por filtros do time (campanha de marca, criativos descontinuados,
    etc.). Não é equivalente à query raw da Meta Insights API. Confirmado
    pelo usuário em 2026-05-14 como a fonte oficial.
  - LFs sem dado completo (LF54: só composição, sem report individual)
    são excluídos do ranking.

OUTPUT
──────
  outputs/analysis/roas_attributable_60d.csv  (canonical)
  Stdout: tabela ordenada por ROAS desc + Top 5 destacado

USO
───
  python -m scripts.compute_top5_roas_attributable
  python -m scripts.compute_top5_roas_attributable --window-days 90
"""
from __future__ import annotations

import argparse
import logging
import pickle
import re
import sys
from pathlib import Path

import pandas as pd
import yaml

REPO_ROOT = Path(__file__).resolve().parents[1]

LAUNCHES_YAML = REPO_ROOT / 'configs' / 'launches.yaml'
RAILWAY_LEADS = REPO_ROOT / 'files' / 'validation' / 'cache' / 'railway_leads_2024-12-30_2026-05-08.parquet'
LEGACY_LEADS_DIR = REPO_ROOT / 'outputs' / 'validation' / 'arquivos_leads'
GURU_HOTMART_PKL = REPO_ROOT / 'outputs' / 'cache' / 'raw_data_latest.pkl'
ASAAS_PARQUET = REPO_ROOT / 'outputs' / 'analysis' / 'asaas_realized.parquet'
TMB_GLOB = 'data/devclub/pedidos_*.xlsx'
INDIVIDUAL_REPORTS_GLOB = 'outputs/validation/20*/LF*.xlsx'

OUTPUT_CSV_TPL = 'outputs/analysis/roas_attributable_{window}d.csv'

TMB_PRIMEIRA_PARCELA_FRAC = 1 / 12

logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# NORMALIZAÇÃO
# ─────────────────────────────────────────────────────────────────────────────

def _norm_email(s):
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return None
    s = str(s).strip().lower()
    return s if '@' in s and '.' in s else None


_DIGITS = re.compile(r'\D+')


def _norm_phone(s):
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return None
    s = _DIGITS.sub('', str(s))
    if len(s) >= 12 and s.startswith('55'):
        s = s[2:]
    return s if 10 <= len(s) <= 11 else None


def _norm_id_col(df: pd.DataFrame, email_col: str, phone_col: str) -> pd.DataFrame:
    out = df.copy()
    out['_email'] = out[email_col].map(_norm_email) if email_col in out.columns else None
    out['_phone'] = out[phone_col].map(_norm_phone) if phone_col in out.columns else None
    return out


# ─────────────────────────────────────────────────────────────────────────────
# LEADS
# ─────────────────────────────────────────────────────────────────────────────

def load_leads_railway() -> pd.DataFrame:
    df = pd.read_parquet(RAILWAY_LEADS)
    df['Data'] = pd.to_datetime(df['Data'], errors='coerce')
    df = _norm_id_col(df, 'E-mail', 'Telefone')
    return df[['Data', '_email', '_phone']].rename(columns={'Data': 'capture_date'})


def load_leads_legacy_lf(lf: str) -> pd.DataFrame:
    """LF40-LF42: lê outputs/validation/arquivos_leads/[LFxx] Leads.xlsx"""
    path = LEGACY_LEADS_DIR / f'[{lf}] Leads.xlsx'
    if not path.exists():
        return pd.DataFrame(columns=['capture_date', '_email', '_phone'])
    df = pd.read_excel(path)
    df['DATA'] = pd.to_datetime(df['DATA'], errors='coerce')
    df = _norm_id_col(df, 'E-MAIL', 'TELEFONE')
    return df[['DATA', '_email', '_phone']].rename(columns={'DATA': 'capture_date'})


def assign_leads_to_lf(leads: pd.DataFrame, launches: dict) -> pd.DataFrame:
    """Atribui cada lead ao LF cuja janela cap_start/cap_end contém capture_date.
       Lead que cai em múltiplas janelas (sobreposição) → primeira (cronológica)."""
    out = leads.copy()
    out['lf'] = None
    out['_cap_dt'] = pd.to_datetime(out['capture_date'])
    # Itera em ordem cronológica de cap_start
    items = sorted(launches.items(), key=lambda kv: kv[1].get('cap_start', '9999'))
    for name, cfg in items:
        cs = pd.to_datetime(cfg.get('cap_start'))
        ce = pd.to_datetime(cfg.get('cap_end'))
        if pd.isna(cs) or pd.isna(ce):
            continue
        # Apenas leads ainda sem LF
        mask = out['lf'].isna() & (out['_cap_dt'] >= cs) & (out['_cap_dt'] <= ce + pd.Timedelta(days=1))
        out.loc[mask, 'lf'] = name
    return out.drop(columns='_cap_dt')


def dedup_first_lf(leads: pd.DataFrame) -> pd.DataFrame:
    """Cada identidade (email ou telefone) só conta no LF mais antigo (capture_date mín).
       Critério: ordena por capture_date asc, dedup por email primeiro, depois por
       phone (preservando o registro mais antigo)."""
    out = leads.sort_values('capture_date').copy()
    out['_keep'] = True
    # Dedup por email
    seen_email = set()
    for idx, row in out[out['_email'].notna()].iterrows():
        if row['_email'] in seen_email:
            out.at[idx, '_keep'] = False
        else:
            seen_email.add(row['_email'])
    # Dedup por phone (entre os que sobraram)
    survivors = out[out['_keep']]
    seen_phone = set()
    for idx, row in survivors[survivors['_phone'].notna()].iterrows():
        if row['_phone'] in seen_phone:
            out.at[idx, '_keep'] = False
        else:
            seen_phone.add(row['_phone'])
    return out[out['_keep']].drop(columns='_keep').copy()


# ─────────────────────────────────────────────────────────────────────────────
# VENDAS
# ─────────────────────────────────────────────────────────────────────────────

def load_sales_guru_hotmart() -> pd.DataFrame:
    with open(GURU_HOTMART_PKL, 'rb') as f:
        raw = pickle.load(f)
    out = []
    for key, source in [('[API] Vendas Guru', 'guru'), ('[API] Vendas Hotmart', 'hotmart')]:
        df = raw[key]['Sheet1'].copy()
        df = df[df['status'].astype(str).str.lower() == 'aprovada']
        df['sale_date'] = pd.to_datetime(df['data'], errors='coerce')
        df['sale_value'] = pd.to_numeric(df['valor'], errors='coerce')
        df = df[df['sale_date'].notna() & df['sale_value'].notna()]
        df = _norm_id_col(df, 'email', 'telefone')
        df['source'] = source
        out.append(df[['sale_date', 'sale_value', '_email', '_phone', 'source']])
    return pd.concat(out, ignore_index=True)


def load_sales_asaas() -> pd.DataFrame:
    df = pd.read_parquet(ASAAS_PARQUET)
    df['sale_date'] = pd.to_datetime(df['sale_date'], errors='coerce')
    df['sale_value'] = pd.to_numeric(df['sale_value'], errors='coerce')
    df = df[df['sale_date'].notna() & df['sale_value'].notna()]
    df = _norm_id_col(df, 'email', 'telefone')
    df['source'] = 'asaas'
    return df[['sale_date', 'sale_value', '_email', '_phone', 'source']]


def load_sales_tmb() -> pd.DataFrame:
    paths = sorted(REPO_ROOT.glob(TMB_GLOB), key=lambda p: p.stat().st_mtime, reverse=True)
    if not paths:
        logger.warning('  Sem TMB pedidos_*.xlsx')
        return pd.DataFrame(columns=['sale_date', 'sale_value', '_email', '_phone', 'source'])
    df = pd.read_excel(paths[0])
    df['Data Efetivado'] = pd.to_datetime(df['Data Efetivado'], errors='coerce')
    df['Data Cancelado'] = pd.to_datetime(df['Data Cancelado'], errors='coerce')
    df = df[df['Data Efetivado'].notna() & df['Data Cancelado'].isna()]
    df['Ticket do pedido'] = pd.to_numeric(df['Ticket do pedido'], errors='coerce')
    df = df[df['Ticket do pedido'].notna() & (df['Ticket do pedido'] > 0)]
    df = _norm_id_col(df, 'E-mail do Cliente', 'Telefone do Cliente')
    out = pd.DataFrame({
        'sale_date': df['Data Efetivado'].values,
        'sale_value': df['Ticket do pedido'].values * TMB_PRIMEIRA_PARCELA_FRAC,
        '_email': df['_email'].values,
        '_phone': df['_phone'].values,
        'source': 'tmb_p1',
    })
    return out


# ─────────────────────────────────────────────────────────────────────────────
# SPEND (xlsx individual mais recente)
# ─────────────────────────────────────────────────────────────────────────────

def load_spend_per_lf() -> dict:
    import openpyxl
    from collections import defaultdict
    files_by_lf = defaultdict(list)
    for p in REPO_ROOT.glob(INDIVIDUAL_REPORTS_GLOB):
        m = re.match(r'^(LF\d+)\s', p.name)
        if m:
            files_by_lf[m.group(1)].append(p)
    spend = {}
    for lf, files in files_by_lf.items():
        latest = max(files, key=lambda p: p.stat().st_mtime)
        wb = openpyxl.load_workbook(latest, data_only=True, read_only=True)
        if 'Performance Geral' not in wb.sheetnames:
            wb.close()
            continue
        ws = wb['Performance Geral']
        for row in ws.iter_rows(values_only=True):
            if row and row[0] and 'Gasto Total' in str(row[0]):
                spend[lf] = (float(row[1]) if row[1] is not None else None, latest.name)
                break
        wb.close()
    return spend


# ─────────────────────────────────────────────────────────────────────────────
# MATCHING lead × venda
# ─────────────────────────────────────────────────────────────────────────────

def match_sales_to_leads(leads: pd.DataFrame, sales: pd.DataFrame, window_days: int) -> pd.DataFrame:
    """Pra cada venda, acha o lead matching (email OU phone). Se houver match
       E sale_date ∈ (capture_date, capture_date + window_days], conta.
       Resolve duplicidade: 1 venda só pode ser atribuída a 1 lead (o mais
       antigo cuja janela contém sale_date — ie. primeiro LF)."""
    # Index por email e phone pra lookup O(1)
    leads_by_email = (
        leads.dropna(subset=['_email'])
        .sort_values('capture_date')
        .drop_duplicates(subset=['_email'], keep='first')
        .set_index('_email')[['capture_date', 'lf']]
        .to_dict('index')
    )
    leads_by_phone = (
        leads.dropna(subset=['_phone'])
        .sort_values('capture_date')
        .drop_duplicates(subset=['_phone'], keep='first')
        .set_index('_phone')[['capture_date', 'lf']]
        .to_dict('index')
    )

    win = pd.Timedelta(days=window_days)
    out = []
    matched_email = matched_phone = matched_neither = 0
    in_window = out_of_window = 0
    for _, s in sales.iterrows():
        sd = s['sale_date']
        lead_info = None
        if s['_email'] and s['_email'] in leads_by_email:
            lead_info = leads_by_email[s['_email']]
            matched_email += 1
        elif s['_phone'] and s['_phone'] in leads_by_phone:
            lead_info = leads_by_phone[s['_phone']]
            matched_phone += 1
        else:
            matched_neither += 1
            continue
        cd = lead_info['capture_date']
        if pd.isna(cd) or pd.isna(sd):
            continue
        if cd < sd <= cd + win:
            in_window += 1
            out.append({
                'lf': lead_info['lf'],
                'sale_date': sd,
                'capture_date': cd,
                'sale_value': s['sale_value'],
                'source': s['source'],
                'days_to_sale': (sd - cd).days,
            })
        else:
            out_of_window += 1

    logger.info(
        f'  Vendas: {len(sales):,} → matched_email={matched_email:,}, matched_phone={matched_phone:,}, '
        f'sem_match={matched_neither:,} | in_window={in_window:,}, out_of_window={out_of_window:,}'
    )
    return pd.DataFrame(out)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument('--window-days', type=int, default=60,
                        help='Janela de atribuição em dias após captação (default 60)')
    parser.add_argument('--top-n', type=int, default=5)
    args = parser.parse_args()

    launches = yaml.safe_load(LAUNCHES_YAML.read_text())

    logger.info('=== LEADS ===')
    # 1) Pra cada LF de launches.yaml, prefere legacy xlsx se existir.
    #    Atribui o LF diretamente (sem inferir por data — confia no arquivo).
    logger.info('[1/4] Carregando leads legacy ([LFxx] Leads.xlsx)...')
    legacy_parts = []
    legacy_lfs_covered = set()
    for lf in launches.keys():
        path = LEGACY_LEADS_DIR / f'[{lf}] Leads.xlsx'
        if not path.exists():
            continue
        df_l = load_leads_legacy_lf(lf)
        df_l['lf'] = lf  # atribui direto ao LF do arquivo
        legacy_parts.append(df_l)
        legacy_lfs_covered.add(lf)
        logger.info(f'  {lf}: {len(df_l):,} leads (legacy)')
    leads_legacy = pd.concat(legacy_parts, ignore_index=True) if legacy_parts else pd.DataFrame()

    # 2) Pra LFs SEM legacy, usa Railway e atribui por janela cap_start..cap_end
    logger.info('[2/4] Carregando Railway pra LFs sem legacy...')
    leads_rail_raw = load_leads_railway()
    logger.info(f'  Railway bruto: {len(leads_rail_raw):,} leads')
    # Atribui apenas a LFs não cobertos pelo legacy
    launches_railway_only = {n: c for n, c in launches.items() if n not in legacy_lfs_covered}
    leads_rail = assign_leads_to_lf(leads_rail_raw, launches_railway_only)
    leads_rail = leads_rail.dropna(subset=['lf'])
    logger.info(f'  Railway atribuídos (LFs não-legacy): {len(leads_rail):,}')

    leads_lf = pd.concat([leads_legacy, leads_rail], ignore_index=True)
    logger.info(f'[3/4] Total leads atribuídos: {len(leads_lf):,} '
                f'(legacy: {len(leads_legacy):,}, railway: {len(leads_rail):,})')

    logger.info('[4/4] Dedup primeiro-LF (email/phone)...')
    leads_dedup = dedup_first_lf(leads_lf.dropna(subset=['lf']))
    logger.info(f'  Após dedup: {len(leads_dedup):,}')
    leads_per_lf = leads_dedup.groupby('lf').size().to_dict()

    logger.info('\n=== VENDAS ===')
    logger.info('[1/3] Guru + Hotmart...')
    df_gh = load_sales_guru_hotmart()
    logger.info(f'  {len(df_gh):,}')
    logger.info('[2/3] Asaas...')
    df_as = load_sales_asaas()
    logger.info(f'  {len(df_as):,}')
    logger.info('[3/3] TMB...')
    df_tmb = load_sales_tmb()
    logger.info(f'  {len(df_tmb):,}')
    sales_all = pd.concat([df_gh, df_as, df_tmb], ignore_index=True)
    logger.info(f'  Total combinado: {len(sales_all):,} vendas aprovadas/efetivadas')

    logger.info(f'\n=== MATCHING (janela {args.window_days}d) ===')
    matched = match_sales_to_leads(leads_dedup, sales_all, args.window_days)
    logger.info(f'  Vendas atribuíveis: {len(matched):,}')

    receita_per_lf = matched.groupby('lf')['sale_value'].sum().to_dict()
    n_compradores_per_lf = matched.groupby('lf').size().to_dict()
    receita_por_source = matched.groupby(['lf', 'source'])['sale_value'].sum().unstack(fill_value=0).to_dict('index')

    logger.info('\n=== SPEND ===')
    spend_data = load_spend_per_lf()
    logger.info(f'  {len(spend_data)} LFs com Gasto Total')

    rows = []
    for lf, cfg in launches.items():
        if lf not in spend_data:
            continue
        sp, sp_file = spend_data[lf]
        if sp is None or sp == 0:
            continue
        n_leads = leads_per_lf.get(lf, 0)
        rec = receita_per_lf.get(lf, 0)
        n_v = n_compradores_per_lf.get(lf, 0)
        src = receita_por_source.get(lf, {})
        rows.append({
            'lf': lf,
            'cap_start': cfg.get('cap_start'),
            'cap_end': cfg.get('cap_end'),
            'n_leads': n_leads,
            'n_compradores_60d': n_v,
            'taxa_conv_60d_pct': (n_v / n_leads * 100) if n_leads else 0,
            'receita_60d_brl': rec,
            'spend_meta_brl': sp,
            'roas_atribuivel_60d': rec / sp if sp else 0,
            'receita_guru': src.get('guru', 0),
            'receita_hotmart': src.get('hotmart', 0),
            'receita_asaas': src.get('asaas', 0),
            'receita_tmb_p1': src.get('tmb_p1', 0),
            'spend_source_file': sp_file,
        })
    df_out = pd.DataFrame(rows).sort_values('roas_atribuivel_60d', ascending=False).reset_index(drop=True)
    df_out.insert(0, 'rank', range(1, len(df_out) + 1))

    output_csv = REPO_ROOT / OUTPUT_CSV_TPL.format(window=args.window_days)
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    df_out.to_csv(output_csv, index=False)

    # Print tabela
    print()
    print('═' * 110)
    print(f'  ROAS ATRIBUÍVEL {args.window_days}d — fonte canônica')
    print('═' * 110)
    print(f"  {'#':>2} {'LF':<6} {'cap':<23} {'leads':>7} {'compr':>5} {'conv%':>6} {'receita':>11} {'spend':>11} {'ROAS':>6}")
    print('  ' + '─' * 88)
    for _, r in df_out.iterrows():
        rank = r['rank']
        mark = ' ★' if rank <= args.top_n else '  '
        print(f"  {rank:>2}{mark}{r['lf']:<6} {r['cap_start']} → {r['cap_end']}"
              f" {r['n_leads']:>7,} {r['n_compradores_60d']:>5,} {r['taxa_conv_60d_pct']:>5.2f}%"
              f" {r['receita_60d_brl']:>10,.0f}  {r['spend_meta_brl']:>10,.0f} {r['roas_atribuivel_60d']:>5.2f}x")
    print('═' * 110)
    print(f'\n  → CSV: {output_csv.relative_to(REPO_ROOT)}')
    print(f"  → Top {args.top_n}: {', '.join(df_out.head(args.top_n)['lf'].tolist())}")


if __name__ == '__main__':
    main()
