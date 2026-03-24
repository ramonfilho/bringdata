#!/usr/bin/env python3
"""
Série Temporal Mensal de ROAS e Margem de Contribuição — DevClub.

Pergunta central: depois que ML entrou em produção (dez/2025),
o DevClub passou a ganhar mais dinheiro em termos absolutos?

Metodologia — agregação mensal, sem matching de leads:
  ROAS    = receita_total_mês / gasto_meta_mês
  Margem  = receita_total_mês − gasto_meta_mês

  Receita total = TODAS as vendas DevClub no mês calendário
                  (Guru + Hotmart + ASAS + TMB), sem filtro por lead.
                  Vendas > MAX_SALE_VALUE excluídas (B2B corporativo).
  Gasto Meta    = soma de todas as campanhas no mês calendário.

ML entrou em produção: dezembro/2025 (LF40).

Uso:
  python V2/scripts/analise_serie_temporal.py [--no-cache] [--output caminho.xlsx]
"""

import argparse
import os
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import pandas as pd
from dotenv import load_dotenv

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / 'src'))
load_dotenv(ROOT / '.env')

VALIDATION_DIR = ROOT / 'outputs' / 'validation'
LEADS_DIR      = ROOT / 'outputs' / 'validation' / 'arquivos_leads'
OUTPUT_DIR     = ROOT / 'outputs' / 'validation' / 'historico'
DATA_DIR       = ROOT / 'data' / 'devclub'
CACHE_DIR      = ROOT / 'files' / 'validation' / 'cache' / 'serie_temporal'
TMB_PEDIDOS_FILE = DATA_DIR / 'pedidos_03032026_1433.xlsx'

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
CACHE_DIR.mkdir(parents=True, exist_ok=True)

META_ACCOUNTS = [
    'act_188005769808959',
    'act_786790755803474',
]

# Mês a partir do qual ML está ativo
ML_START_MONTH = '2025-12'  # dezembro 2025 = LF40

# Teto de valor por venda — acima disso é B2B corporativo (contratos únicos)
MAX_SALE_VALUE = 10_000.0

# Período de análise
ANALYSIS_START = '2025-04'  # primeiro mês com dados completos
ANALYSIS_END   = '2026-03'  # último mês fechado

CACHE_VERSION = 'v1m'  # 'm' = mensal

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def _vendas_range(cap_end: str) -> tuple[str, str]:
    end     = datetime.strptime(cap_end, '%Y-%m-%d')
    v_start = end + timedelta(days=7)
    v_end   = end + timedelta(days=14)
    return v_start.strftime('%Y-%m-%d'), v_end.strftime('%Y-%m-%d')


def _cache_path(prefix: str, start: str, end: str) -> Path:
    return CACHE_DIR / f"{prefix}_{start}_{end}_{CACHE_VERSION}.parquet"


def _norm_email(s) -> Optional[str]:
    if not s or isinstance(s, float):
        return None
    return str(s).lower().strip()


# ─── MESES DE ANÁLISE ─────────────────────────────────────────────────────────

def _months_range(start: str, end: str) -> list[tuple[str, str, str]]:
    """
    Retorna lista de (label, primeiro_dia, ultimo_dia) para cada mês no intervalo.
    start/end no formato 'YYYY-MM'.
    """
    months = []
    y, m = int(start[:4]), int(start[5:7])
    ey, em = int(end[:4]), int(end[5:7])
    while (y, m) <= (ey, em):
        import calendar
        last_day = calendar.monthrange(y, m)[1]
        label      = f"{y}-{m:02d}"
        first      = f"{y}-{m:02d}-01"
        last       = f"{y}-{m:02d}-{last_day:02d}"
        is_ml      = label >= ML_START_MONTH
        months.append({'label': label, 'start': first, 'end': last, 'is_ml': is_ml})
        m += 1
        if m > 12:
            m = 1; y += 1
    return months


# ─── META SPEND ───────────────────────────────────────────────────────────────

def _get_meta_spend(cap_start: str, cap_end: str, use_cache: bool) -> float:
    cache = _cache_path('meta_spend', cap_start, cap_end)
    if use_cache and cache.exists():
        df    = pd.read_parquet(cache)
        total = float(df['spend'].sum())
        print(f"  [Meta cache] R$ {total:,.0f}")
        return total

    from validation.meta_api_client import MetaAPIClient
    SPEND_COL   = 'Valor usado (BRL)'
    total_spend = 0.0
    for account_id in META_ACCOUNTS:
        try:
            client     = MetaAPIClient(account_id=account_id)
            df         = client.get_campaigns(cap_start, cap_end, apply_filters=True)
            if not df.empty and SPEND_COL in df.columns:
                acct   = pd.to_numeric(df[SPEND_COL], errors='coerce').sum()
                total_spend += acct
                print(f"  [Meta] {account_id}: R$ {acct:,.0f}")
        except Exception as e:
            print(f"  [Meta ERRO] {account_id}: {e}")

    pd.DataFrame([{'spend': total_spend}]).to_parquet(cache)
    return total_spend


# ─── RECEITA TOTAL (sem matching) ─────────────────────────────────────────────

_tmb_df: Optional[pd.DataFrame] = None


def _load_tmb() -> pd.DataFrame:
    global _tmb_df
    if _tmb_df is not None:
        return _tmb_df
    if not TMB_PEDIDOS_FILE.exists():
        return pd.DataFrame()
    raw = pd.read_excel(
        TMB_PEDIDOS_FILE,
        usecols=['E-mail do Cliente', 'Ticket do pedido', 'Data Efetivado', 'Criado em', 'Status Pedido'],
    )
    raw['sale_value'] = pd.to_numeric(raw['Ticket do pedido'], errors='coerce')
    raw['sale_date']  = pd.to_datetime(raw['Data Efetivado'], errors='coerce').fillna(
        pd.to_datetime(raw['Criado em'], errors='coerce')
    )
    raw = raw[~raw['Status Pedido'].isin(['Cancelado', 'Cancelamento solicitado'])]
    raw = raw.dropna(subset=['sale_value', 'sale_date'])
    raw['origem'] = 'tmb'
    _tmb_df = raw[['sale_value', 'sale_date', 'origem']].copy()
    return _tmb_df


def _get_total_revenue(vendas_start: str, vendas_end: str, use_cache: bool) -> tuple[float, int]:
    """
    Retorna (receita_total, n_vendas) — soma de TODAS as vendas no período,
    sem filtro por lead ou campanha. Exclui vendas > MAX_SALE_VALUE (B2B).
    """
    cache = _cache_path('revenue_total', vendas_start, vendas_end)
    if use_cache and cache.exists():
        df = pd.read_parquet(cache)
        print(f"  [Revenue cache] R$ {df['sale_value'].sum():,.0f} ({len(df)} vendas)")
        return float(df['sale_value'].sum()), len(df)

    frames = []

    # Guru
    try:
        from validation.guru_sales_extractor import fetch_guru_sales_from_api
        raw = fetch_guru_sales_from_api(vendas_start, vendas_end)
        if not raw.empty:
            mask = pd.Series([True] * len(raw), index=raw.index)
            if 'status' in raw.columns:
                mask = raw['status'].isin(['Aprovada', 'approved'])
            guru = pd.DataFrame({
                'sale_value': pd.to_numeric(raw[mask]['valor venda'], errors='coerce'),
                'origem':     'guru',
            }).dropna(subset=['sale_value'])
            frames.append(guru)
            print(f"  [Guru] {len(guru)} vendas R$ {guru['sale_value'].sum():,.0f}")
    except Exception as e:
        print(f"  [Guru ERRO] {e}")

    # Hotmart
    try:
        from validation.data_loader import SalesDataLoader
        hm = SalesDataLoader().load_hotmart_sales_from_api(vendas_start, vendas_end)
        if not hm.empty:
            hm = pd.DataFrame({
                'sale_value': pd.to_numeric(hm['sale_value'], errors='coerce'),
                'origem':     'hotmart',
            }).dropna(subset=['sale_value'])
            frames.append(hm)
            print(f"  [Hotmart] {len(hm)} vendas R$ {hm['sale_value'].sum():,.0f}")
    except Exception as e:
        print(f"  [Hotmart ERRO] {e}")

    # ASAS
    try:
        from validation.asaas_sales_extractor import fetch_asaas_sales
        asaas = fetch_asaas_sales(vendas_start, vendas_end)
        if not asaas.empty:
            asaas = pd.DataFrame({
                'sale_value': pd.to_numeric(asaas['sale_value'], errors='coerce'),
                'origem':     'asaas',
            }).dropna(subset=['sale_value'])
            frames.append(asaas)
            print(f"  [ASAS] {len(asaas)} vendas R$ {asaas['sale_value'].sum():,.0f}")
    except Exception as e:
        print(f"  [ASAS ERRO] {e}")

    # TMB
    try:
        tmb_all = _load_tmb()
        if not tmb_all.empty:
            start = pd.Timestamp(vendas_start)
            end   = pd.Timestamp(vendas_end) + pd.Timedelta(days=1)
            tmb   = tmb_all[
                (tmb_all['sale_date'] >= start) & (tmb_all['sale_date'] < end)
            ][['sale_value', 'origem']].copy()
            if not tmb.empty:
                frames.append(tmb)
                print(f"  [TMB] {len(tmb)} vendas R$ {tmb['sale_value'].sum():,.0f}")
    except Exception as e:
        print(f"  [TMB ERRO] {e}")

    if not frames:
        empty = pd.DataFrame(columns=['sale_value', 'origem'])
        empty.to_parquet(cache)
        return 0.0, 0

    combined = pd.concat(frames, ignore_index=True)

    # Remover B2B corporativo
    n_before = len(combined)
    combined = combined[combined['sale_value'] <= MAX_SALE_VALUE]
    n_removed = n_before - len(combined)
    if n_removed > 0:
        print(f"  [Filtro B2B] {n_removed} vendas acima de R$ {MAX_SALE_VALUE:,.0f} removidas")

    combined.to_parquet(cache)
    return float(combined['sale_value'].sum()), len(combined)


# ─── CÁLCULO MENSAL ───────────────────────────────────────────────────────────

def calc_month(month: dict, use_cache: bool) -> dict:
    label  = month['label']
    start  = month['start']
    end    = month['end']
    is_ml  = month['is_ml']

    print(f"\n{'─'*60}")
    print(f"[{label}] {start} → {end} | {'Pós-ML' if is_ml else 'Pré-ML'}")

    gasto             = _get_meta_spend(start, end, use_cache)
    receita, n_vendas = _get_total_revenue(start, end, use_cache)

    print(f"  Gasto: R$ {gasto:,.0f} | Receita: R$ {receita:,.0f} | Vendas: {n_vendas}")

    roas   = receita / gasto if gasto > 0 else None
    margem = receita - gasto if gasto > 0 else None

    return {
        'mes':      label,
        'start':    start,
        'end':      end,
        'gasto':    gasto,
        'receita':  receita,
        'n_vendas': n_vendas,
        'roas':     roas,
        'margem':   margem,
        'is_ml':    is_ml,
    }


# ─── EXCEL ────────────────────────────────────────────────────────────────────

def _write_excel(df: pd.DataFrame, output_path: Path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    ML_FILL  = PatternFill('solid', fgColor='1F4E79')
    PRE_FILL = PatternFill('solid', fgColor='C6EFCE')
    HDR_FILL = PatternFill('solid', fgColor='2F5496')
    HDR_FONT = Font(color='FFFFFF', bold=True)
    ML_FONT  = Font(color='FFFFFF', bold=True)
    CENTER   = Alignment(horizontal='center')
    thin     = Side(style='thin', color='AAAAAA')
    BORDER   = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Aba dados ─────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Série Temporal'

    headers = [
        'Mês', 'Início', 'Fim',
        'Gasto Meta (R$)', 'Receita Total (R$)', 'Nº Vendas',
        'ROAS', 'Margem (R$)', 'Fase',
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = CENTER; cell.border = BORDER

    for _, row in df.iterrows():
        data_row = [
            row['mes'],
            row['start'], row['end'],
            round(row['gasto'], 2)   if pd.notna(row.get('gasto'))   else '',
            round(row['receita'], 2) if pd.notna(row.get('receita')) else '',
            int(row['n_vendas'])     if pd.notna(row.get('n_vendas')) else '',
            round(row['roas'], 3)    if pd.notna(row.get('roas'))     else '',
            round(row['margem'], 2)  if pd.notna(row.get('margem'))   else '',
            'Pós-ML' if row['is_ml'] else 'Pré-ML',
        ]
        ws.append(data_row)
        ri   = ws.max_row
        fill = ML_FILL if row['is_ml'] else PRE_FILL
        font = ML_FONT if row['is_ml'] else Font()
        for c in range(1, len(headers) + 1):
            cell = ws.cell(ri, c)
            cell.fill = fill; cell.font = font
            cell.alignment = CENTER; cell.border = BORDER

    col_widths = [12, 14, 14, 18, 18, 12, 10, 16, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Legenda
    n = len(df)
    ws.cell(n+3, 1, 'Legenda:').font = Font(bold=True)
    ws.cell(n+4, 1, 'Pré-ML').fill = PRE_FILL
    ws.cell(n+4, 2, 'Lançamentos sem sistema de ML')
    ws.cell(n+5, 1, 'Pós-ML').fill = ML_FILL
    ws.cell(n+5, 1).font = ML_FONT
    ws.cell(n+5, 2, 'Lançamentos com ML ativo')
    ws.cell(n+7, 1, 'Nota:').font = Font(bold=True)
    ws.cell(n+7, 2, 'Receita = TODAS as vendas DevClub no período (Guru+Hotmart+ASAS+TMB), sem filtro por campanha.')
    ws.cell(n+8, 2, 'Representa receita total do negócio na semana de vendas de cada lançamento.')

    # ── Aba gráficos ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Gráficos')

    df_c = df[df['roas'].notna() & df['gasto'].notna() & (df['gasto'] > 0)].reset_index(drop=True)
    n_c  = len(df_c)

    if n_c > 0:
        ws2.cell(1, 1, 'Mês').font = Font(bold=True)
        ws2.cell(1, 2, 'ROAS').font       = Font(bold=True)
        ws2.cell(1, 3, 'Margem (R$)').font = Font(bold=True)
        ws2.cell(1, 4, 'Receita (R$)').font = Font(bold=True)
        ws2.cell(1, 5, 'Gasto (R$)').font  = Font(bold=True)
        ws2.cell(1, 6, 'ML').font          = Font(bold=True)

        for i, (_, r) in enumerate(df_c.iterrows(), 2):
            ws2.cell(i, 1, r['mes'])
            ws2.cell(i, 2, round(r['roas'],   3) if pd.notna(r['roas'])   else None)
            ws2.cell(i, 3, round(r['margem'], 0) if pd.notna(r['margem']) else None)
            ws2.cell(i, 4, round(r['receita'], 0))
            ws2.cell(i, 5, round(r['gasto'],   0))
            ws2.cell(i, 6, 1 if r['is_ml'] else 0)

        cats = Reference(ws2, min_col=1, min_row=2, max_row=n_c+1)

        # Gráfico ROAS
        roas_chart = LineChart()
        roas_chart.title = 'Evolução Mensal do ROAS — Pré-ML vs Pós-ML (dez/2025+)'
        roas_chart.y_axis.title = 'ROAS (Receita Total / Gasto Meta)'
        roas_chart.x_axis.title = 'Mês'
        roas_chart.width = 32; roas_chart.height = 16
        roas_data = Reference(ws2, min_col=2, min_row=1, max_row=n_c+1)
        roas_chart.add_data(roas_data, titles_from_data=True)
        roas_chart.set_categories(cats)
        roas_chart.series[0].graphicalProperties.line.solidFill = '2196F3'
        roas_chart.series[0].graphicalProperties.line.width     = 20000
        roas_chart.series[0].marker.symbol = 'circle'
        roas_chart.series[0].marker.size   = 6
        ws2.add_chart(roas_chart, 'H1')

        # Gráfico Margem
        margem_chart = BarChart()
        margem_chart.title = 'Margem de Contribuição Mensal'
        margem_chart.y_axis.title = 'Margem (R$)'
        margem_chart.x_axis.title = 'Mês'
        margem_chart.type  = 'col'
        margem_chart.width = 32; margem_chart.height = 16
        margem_data = Reference(ws2, min_col=3, min_row=1, max_row=n_c+1)
        margem_chart.add_data(margem_data, titles_from_data=True)
        margem_chart.set_categories(cats)
        for i, (_, r) in enumerate(df_c.iterrows()):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = '1F4E79' if r['is_ml'] else '70AD47'
            margem_chart.series[0].dPt.append(pt)
        ws2.add_chart(margem_chart, 'H32')

        # Gráfico Receita + Gasto (barras empilhadas)
        combo_chart = BarChart()
        combo_chart.title = 'Receita Total vs Gasto Meta por Mês'
        combo_chart.y_axis.title = 'R$'
        combo_chart.type  = 'col'
        combo_chart.width = 32; combo_chart.height = 16
        rec_data   = Reference(ws2, min_col=4, min_row=1, max_row=n_c+1)
        gasto_data = Reference(ws2, min_col=5, min_row=1, max_row=n_c+1)
        combo_chart.add_data(rec_data,   titles_from_data=True)
        combo_chart.add_data(gasto_data, titles_from_data=True)
        combo_chart.set_categories(cats)
        combo_chart.series[0].graphicalProperties.solidFill = '4CAF50'  # receita verde
        combo_chart.series[1].graphicalProperties.solidFill = 'F44336'  # gasto vermelho
        ws2.add_chart(combo_chart, 'H63')

    # ── Aba Resumo ────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet('Resumo')
    ws3.cell(1, 1, 'RESUMO: PRÉ-ML vs PÓS-ML — Série Temporal').font = Font(bold=True, size=14)
    ws3.cell(2, 1, 'Receita = total de vendas na semana de vendas de cada lançamento (sem matching)').font = Font(italic=True)

    pre  = df[~df['is_ml'] & df['roas'].notna() & (df['gasto'] > 0)]
    post = df[ df['is_ml'] & df['roas'].notna() & (df['gasto'] > 0)]

    for subset, label, col in [(pre, 'Pré-ML', 1), (post, 'Pós-ML', 4)]:
        ws3.cell(4, col, label).font = Font(bold=True)
        ws3.cell(5, col, 'Lançamentos');       ws3.cell(5, col+1, len(subset))
        ws3.cell(6, col, 'ROAS Médio');        ws3.cell(6, col+1, round(subset['roas'].mean(), 2)    if len(subset) else '-')
        ws3.cell(7, col, 'ROAS Mediano');      ws3.cell(7, col+1, round(subset['roas'].median(), 2)  if len(subset) else '-')
        ws3.cell(8, col, 'Margem Total (R$)'); ws3.cell(8, col+1, round(subset['margem'].sum(), 0)   if len(subset) else '-')
        ws3.cell(9, col, 'Margem Média/lançamento (R$)'); ws3.cell(9, col+1, round(subset['margem'].mean(), 0) if len(subset) else '-')
        ws3.cell(10, col, 'Receita Total (R$)'); ws3.cell(10, col+1, round(subset['receita'].sum(), 0) if len(subset) else '-')
        ws3.cell(11, col, 'Gasto Total (R$)');  ws3.cell(11, col+1, round(subset['gasto'].sum(), 0)   if len(subset) else '-')

    if len(pre) > 0 and len(post) > 0:
        ws3.cell(13, 1, 'Delta ROAS médio (pós − pré)').font = Font(bold=True)
        ws3.cell(13, 2, round(post['roas'].mean() - pre['roas'].mean(), 2))
        ws3.cell(14, 1, 'Delta Margem média/lançamento (R$)').font = Font(bold=True)
        ws3.cell(14, 2, round(post['margem'].mean() - pre['margem'].mean(), 0))
        ws3.cell(15, 1, 'Uplift ROAS (%)').font = Font(bold=True)
        uplift = (post['roas'].mean() / pre['roas'].mean() - 1) * 100 if pre['roas'].mean() > 0 else None
        ws3.cell(15, 2, f"{round(uplift, 1)}%" if uplift else '-')

    for c in [1, 2, 4, 5]:
        ws3.column_dimensions[get_column_letter(c)].width = 35 if c % 3 == 1 else 18

    wb.save(output_path)
    print(f"\n Planilha salva em: {output_path}")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def run(use_cache: bool = True, output_path: Optional[Path] = None):
    print("=" * 60)
    print("SÉRIE TEMPORAL MENSAL — ROAS & MARGEM — DevClub")
    print(f"Período: {ANALYSIS_START} → {ANALYSIS_END} | ML a partir de: {ML_START_MONTH}")
    print(f"Filtro B2B: vendas > R$ {MAX_SALE_VALUE:,.0f} excluídas")
    print("=" * 60)

    months = _months_range(ANALYSIS_START, ANALYSIS_END)
    print(f"\n{len(months)} meses: {months[0]['label']} → {months[-1]['label']}")

    rows = []
    for month in months:
        row = calc_month(month, use_cache=use_cache)
        rows.append(row)

    df = pd.DataFrame(rows)

    if output_path is None:
        ts          = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = OUTPUT_DIR / f'serie_temporal_mensal_{ts}.xlsx'

    _write_excel(df, output_path)

    # Resumo console
    print("\n" + "=" * 60)
    print("RESUMO")
    print("=" * 60)
    pre  = df[~df['is_ml'] & df['roas'].notna() & (df['gasto'] > 0)]
    post = df[ df['is_ml'] & df['roas'].notna() & (df['gasto'] > 0)]
    if len(pre):
        print(f"Pré-ML  ({len(pre)} meses): ROAS médio={pre['roas'].mean():.2f} | "
              f"Receita total=R$ {pre['receita'].sum():,.0f} | Margem total=R$ {pre['margem'].sum():,.0f}")
    if len(post):
        print(f"Pós-ML  ({len(post)} meses): ROAS médio={post['roas'].mean():.2f} | "
              f"Receita total=R$ {post['receita'].sum():,.0f} | Margem total=R$ {post['margem'].sum():,.0f}")

    return df


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Série temporal ROAS DevClub — sem matching')
    parser.add_argument('--no-cache',   action='store_true')
    parser.add_argument('--output',     type=str)
    parser.add_argument('--start',      type=str, default=ANALYSIS_START, help='Mês inicial YYYY-MM')
    parser.add_argument('--end',        type=str, default=ANALYSIS_END,   help='Mês final YYYY-MM')
    parser.add_argument('--max-sale',   type=float, default=MAX_SALE_VALUE, help='Teto por venda (B2B filter)')
    args = parser.parse_args()

    # Sobrescrever constantes se passado via CLI
    import sys as _sys, importlib as _il
    _mod = _il.import_module(__name__) if __name__ != '__main__' else _il.import_module('__main__')
    globals()['ANALYSIS_START'] = args.start
    globals()['ANALYSIS_END']   = args.end
    globals()['MAX_SALE_VALUE'] = args.max_sale

    output = Path(args.output) if args.output else None
    run(use_cache=not args.no_cache, output_path=output)
