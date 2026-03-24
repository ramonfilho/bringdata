"""
export_leads_comercial.py

Exporta leads dos decis selecionados de um lançamento para Excel.
Fonte: Railway (leads_capi), filtrado por data de captação.

Uso:
    # LF48 — D8, D9, D10
    python V2/scripts/export_leads_comercial.py \
        --lancamento LF48 \
        --cap-start 2026-03-10 \
        --cap-end   2026-03-16

    # Decis customizados
    python V2/scripts/export_leads_comercial.py \
        --lancamento LF48 \
        --cap-start 2026-03-10 \
        --cap-end   2026-03-16 \
        --decis D10

    # Salvar em pasta específica
    python V2/scripts/export_leads_comercial.py \
        --lancamento LF48 \
        --cap-start 2026-03-10 \
        --cap-end   2026-03-16 \
        --output /tmp/leads_lf48.xlsx
"""

import argparse
import os
import sys
from pathlib import Path
from datetime import datetime

import pandas as pd
import pg8000.native
from dotenv import load_dotenv

ROOT = Path(__file__).parent.parent
load_dotenv(ROOT / ".env")

# ---------------------------------------------------------------------------
# Colunas exportadas para o time comercial
# ---------------------------------------------------------------------------

COLUNAS_EXPORT = {
    "decil":               "Decil",
    "leadScore":           "Score",
    "nomeCompleto":        "Nome",
    "email":               "Email",
    "telefone":            "Telefone",
    "genero":              "Gênero",
    "idade":               "Idade",
    "ocupacao":            "Ocupação",
    "faixaSalarial":       "Faixa Salarial",
    "estudouProgramacao":  "Estudou Programação",
    "computador":          "Tem Computador",
    "cartaoCredito":       "Tem Cartão de Crédito",
    "source":              "Fonte",
    "campaign":            "Campanha",
    "medium":              "Medium",
    "data":                "Data Captação",
}

# ---------------------------------------------------------------------------
# Railway
# ---------------------------------------------------------------------------

def load_railway(cap_start: str, cap_end: str, decis: list[str]) -> pd.DataFrame:
    """Busca leads da tabela Lead filtrando por período e decil (inteiro 1-10)."""
    end_excl = (pd.to_datetime(cap_end) + pd.Timedelta(days=1)).strftime("%Y-%m-%d")

    # Converter "D8", "D9", "D10" → 8, 9, 10 (decil é inteiro na tabela Lead)
    decis_int = [int(d.lstrip("D")) for d in decis]

    conn = pg8000.native.Connection(
        host=os.environ["RAILWAY_DB_HOST"],
        port=int(os.environ["RAILWAY_DB_PORT"]),
        user=os.environ["RAILWAY_DB_USER"],
        password=os.environ["RAILWAY_DB_PASSWORD"],
        database=os.environ["RAILWAY_DB_NAME"],
        ssl_context=True,
    )

    placeholders = ", ".join(f":d{i}" for i in range(len(decis_int)))
    decil_params = {f"d{i}": d for i, d in enumerate(decis_int)}

    rows = conn.run(
        f"""
        SELECT
            'D' || decil::text          AS decil,
            "leadScore",
            "nomeCompleto",
            LOWER(email)                AS email,
            telefone,
            pesquisa->>'genero'         AS genero,
            pesquisa->>'idade'          AS idade,
            pesquisa->>'ocupacao'       AS ocupacao,
            pesquisa->>'faixaSalarial'  AS "faixaSalarial",
            pesquisa->>'estudouProgramacao' AS "estudouProgramacao",
            pesquisa->>'computador'     AS computador,
            pesquisa->>'cartaoCredito'  AS "cartaoCredito",
            source,
            campaign,
            medium,
            data
        FROM "Lead"
        WHERE data >= :start_date
          AND data <  :end_date_excl
          AND decil IN ({placeholders})
          AND email IS NOT NULL
        ORDER BY "leadScore" DESC NULLS LAST
        """,
        start_date=cap_start,
        end_date_excl=end_excl,
        **decil_params,
    )
    conn.close()

    cols = list(COLUNAS_EXPORT.keys())
    df = pd.DataFrame(rows, columns=cols)
    return df


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def export(df: pd.DataFrame, output_path: Path, lancamento: str, decis: list[str]):
    df_export = df.copy()
    df_export.rename(columns=COLUNAS_EXPORT, inplace=True)

    # Score: formatar como decimal BR
    if "Score" in df_export.columns:
        df_export["Score"] = pd.to_numeric(df_export["Score"], errors="coerce").round(4)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Leads")

        ws = writer.sheets["Leads"]

        # Larguras de coluna
        col_widths = {
            "Decil": 8, "Score": 10, "Nome": 28, "Email": 32, "Telefone": 16,
            "Cidade": 16, "Gênero": 10, "Idade": 12, "Ocupação": 24,
            "Faixa Salarial": 22, "Estudou Programação": 20, "Tem Computador": 16,
            "Tem Cartão de Crédito": 20, "Fonte": 14, "Campanha": 40,
            "Medium": 14, "Data Captação": 20,
        }
        for col_idx, col_name in enumerate(df_export.columns, start=1):
            from openpyxl.utils import get_column_letter
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = col_widths.get(col_name, 14)

        # Header colorido por decil
        from openpyxl.styles import PatternFill, Font, Alignment
        FILLS = {
            "D10": PatternFill("solid", fgColor="1E3A5F"),
            "D9":  PatternFill("solid", fgColor="2563EB"),
            "D8":  PatternFill("solid", fgColor="60A5FA"),
        }
        WHITE_FONT = Font(bold=True, color="FFFFFF")
        DARK_FONT  = Font(bold=True, color="1E3A5F")

        for cell in ws[1]:
            cell.fill    = PatternFill("solid", fgColor="1E3A5F")
            cell.font    = WHITE_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20

        # Colorir linhas por decil
        decil_col_idx = list(df_export.columns).index("Decil") + 1
        for row_idx in range(2, len(df_export) + 2):
            decil_val = ws.cell(row_idx, decil_col_idx).value
            fill = FILLS.get(str(decil_val))
            if fill:
                for cell in ws[row_idx]:
                    cell.fill = fill
                    if decil_val == "D8":
                        cell.font = DARK_FONT
                    else:
                        cell.font = Font(color="FFFFFF")

        ws.freeze_panes = "A2"

    # Resumo
    print(f"\n{'='*50}")
    print(f"Lançamento : {lancamento}")
    print(f"Decis      : {', '.join(decis)}")
    print(f"Total      : {len(df_export):,} leads")
    for d in sorted(decis, reverse=True):
        n = (df_export["Decil"] == d).sum()
        print(f"  {d}: {n:,}")
    print(f"\nSalvo em   : {output_path}")
    print("=" * 50)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Exporta leads por decil para Excel.")
    parser.add_argument("--lancamento",  required=True, help="Nome do lançamento (ex: LF48)")
    parser.add_argument("--cap-start",   required=True, metavar="YYYY-MM-DD", help="Início da captação")
    parser.add_argument("--cap-end",     required=True, metavar="YYYY-MM-DD", help="Fim da captação")
    parser.add_argument("--decis",       default="D8,D9,D10",
                        help="Decis a exportar, separados por vírgula (padrão: D8,D9,D10)")
    parser.add_argument("--output",      default=None, metavar="CAMINHO",
                        help="Caminho do arquivo de saída (padrão: outputs/comercial/<lancamento>_decis_<ts>.xlsx)")
    args = parser.parse_args()

    decis = [d.strip().upper() for d in args.decis.split(",")]

    if args.output:
        output_path = Path(args.output)
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = ROOT / "outputs" / "comercial" / f"{args.lancamento}_leads_{'_'.join(decis)}_{ts}.xlsx"

    print(f"Buscando leads {', '.join(decis)} no Railway ({args.cap_start} → {args.cap_end})...")
    df = load_railway(args.cap_start, args.cap_end, decis)

    if df.empty:
        print(f"Nenhum lead encontrado para os decis {decis} no período informado.")
        sys.exit(1)

    export(df, output_path, args.lancamento, decis)

    import subprocess
    subprocess.run(["open", str(output_path)], check=False)


if __name__ == "__main__":
    main()
