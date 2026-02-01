"""
Script para analisar inadimplência TMB e calcular valor real de vendas

Análises realizadas:
1. Taxa de inadimplência geral (parcelas recebidas, vencidas, aguardando)
2. Valor real de uma venda TMB considerando histórico de pagamento
3. Análise por grau de risco (Alto vs Médio)
4. Evolução temporal da inadimplência

Gera um único arquivo Excel com 4 abas contendo todas as análises.

Uso:
    python analyze_tmb_inadimplencia.py --contas-receber files/validation/vendas/contas_a_receber.xlsx
    python analyze_tmb_inadimplencia.py --contas-receber files/validation/vendas/contas_a_receber.xlsx --output files/analises/relatorio.xlsx
"""

import pandas as pd
import numpy as np
from datetime import datetime
import argparse
import logging
from pathlib import Path
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def analyze_inadimplencia_geral(df_efetivado: pd.DataFrame) -> dict:
    """
    Calcula taxa de inadimplência geral TMB.

    Args:
        df_efetivado: DataFrame com parcelas de pedidos efetivados

    Returns:
        Dict com estatísticas de inadimplência
    """
    logger.info("📊 ANÁLISE 1: Taxa de Inadimplência Geral")

    # Distribuição por status
    status_counts = df_efetivado['Status Parcela'].value_counts()
    status_valores = df_efetivado.groupby('Status Parcela')['Valor da Parcela (R$)'].sum()

    logger.info(f"   Total de parcelas: {len(df_efetivado):,}")

    # Calcular totais
    valor_total = status_valores.sum()
    valor_recebido = status_valores.get('Recebido', 0) + status_valores.get('Antecipado', 0)
    valor_vencido = status_valores.get('Vencido', 0)
    valor_aguardando = status_valores.get('Aguardando Pagamento', 0)

    taxa_inadimplencia = (valor_vencido / valor_total * 100) if valor_total > 0 else 0
    taxa_recebido = (valor_recebido / valor_total * 100) if valor_total > 0 else 0
    taxa_aguardando = (valor_aguardando / valor_total * 100) if valor_total > 0 else 0

    logger.info(f"   Valor Total Contratado: R$ {valor_total:,.2f}")
    logger.info(f"   ✅ Recebido: R$ {valor_recebido:,.2f} ({taxa_recebido:.1f}%)")
    logger.info(f"   ⚠️ Aguardando: R$ {valor_aguardando:,.2f} ({taxa_aguardando:.1f}%)")
    logger.info(f"   ❌ Vencido: R$ {valor_vencido:,.2f} ({taxa_inadimplencia:.1f}%)")

    return {
        'valor_total': valor_total,
        'valor_recebido': valor_recebido,
        'valor_vencido': valor_vencido,
        'valor_aguardando': valor_aguardando,
        'taxa_inadimplencia': taxa_inadimplencia,
        'taxa_recebido': taxa_recebido,
        'taxa_aguardando': taxa_aguardando,
        'status_counts': status_counts.to_dict(),
        'status_valores': status_valores.to_dict()
    }


def analyze_evolucao_temporal(df_efetivado: pd.DataFrame) -> pd.DataFrame:
    """
    Analisa evolução da inadimplência ao longo do tempo.

    Args:
        df_efetivado: DataFrame com parcelas de pedidos efetivados

    Returns:
        DataFrame com taxa de inadimplência por período
    """
    logger.info("📈 ANÁLISE: Evolução Temporal da Inadimplência")

    # Agrupar por ano/mês da venda
    df_efetivado['ano_mes_venda'] = pd.to_datetime(df_efetivado['Data Efetivado']).dt.to_period('M')

    evolucao = []
    periodos_unicos = sorted(df_efetivado['ano_mes_venda'].dropna().unique().tolist())

    for periodo in periodos_unicos:
        df_periodo = df_efetivado[df_efetivado['ano_mes_venda'] == periodo]

        total = df_periodo['Valor da Parcela (R$)'].sum()
        vencido = df_periodo[df_periodo['Status Parcela'] == 'Vencido']['Valor da Parcela (R$)'].sum()

        if total > 0:
            taxa = (vencido / total * 100)
            pedidos = df_periodo['Pedido'].nunique()
            evolucao.append({
                'Periodo': str(periodo),
                'Pedidos': pedidos,
                'Valor_Total': total,
                'Valor_Vencido': vencido,
                'Taxa_Inadimplencia_%': round(taxa, 1)
            })

    df_evolucao = pd.DataFrame(evolucao)

    logger.info(f"   Períodos analisados: {len(df_evolucao)}")
    logger.info(f"   Últimos 6 meses (média): {df_evolucao.tail(6)['Taxa_Inadimplencia_%'].mean():.1f}%")

    return df_evolucao


def analyze_por_grau_risco(df_efetivado: pd.DataFrame) -> pd.DataFrame:
    """
    Analisa inadimplência por grau de risco.

    Args:
        df_efetivado: DataFrame com parcelas de pedidos efetivados

    Returns:
        DataFrame com estatísticas por grau de risco
    """
    logger.info("🎯 ANÁLISE: Inadimplência por Grau de Risco")

    risco_stats = []

    # Obter todos os graus de risco únicos e ordenar
    graus_risco = sorted(df_efetivado['Grau de risco'].unique(),
                         key=lambda x: {'Alto': 1, 'Médio': 2, 'Baixo': 3, '-': 4}.get(x, 5))

    for risco in graus_risco:
        df_risco = df_efetivado[df_efetivado['Grau de risco'] == risco]

        if len(df_risco) > 0:
            total = df_risco['Valor da Parcela (R$)'].sum()
            vencido = df_risco[df_risco['Status Parcela'] == 'Vencido']['Valor da Parcela (R$)'].sum()
            taxa = (vencido / total * 100) if total > 0 else 0
            pedidos = df_risco['Pedido'].nunique()

            # Usar nome mais descritivo para sem classificação
            risco_display = 'Sem Classificação' if risco == '-' else risco

            risco_stats.append({
                'Grau_Risco': risco_display,
                'Pedidos': pedidos,
                'Valor_Total': total,
                'Valor_Vencido': vencido,
                'Taxa_Inadimplencia_%': round(taxa, 1)
            })

            logger.info(f"   {risco_display}: {taxa:.1f}% inadimplência ({pedidos:,} pedidos)")

    return pd.DataFrame(risco_stats)


def calculate_valor_real_venda(df_efetivado: pd.DataFrame, valor_nominal: float = 2200.40) -> dict:
    """
    Calcula o valor real de uma venda TMB baseado em histórico de pagamento.

    Args:
        df_efetivado: DataFrame com parcelas de pedidos efetivados
        valor_nominal: Valor nominal do contrato (padrão R$ 2.200,40)

    Returns:
        Dict com estatísticas de valor real
    """
    logger.info("💰 ANÁLISE 2: Valor Real de uma Venda TMB")

    # Data de referência (hoje)
    data_hoje = datetime.now()

    # Adicionar maturidade
    df_efetivado['data_venda'] = pd.to_datetime(df_efetivado['Data Efetivado'])
    df_efetivado['meses_maturidade'] = ((data_hoje - df_efetivado['data_venda']).dt.days / 30).astype(int)

    # Agrupar por pedido
    pedido_stats = df_efetivado.groupby('Pedido').agg({
        'Valor da Parcela (R$)': 'sum',
        'data_venda': 'first',
        'meses_maturidade': 'first',
        'Grau de risco': lambda x: x.mode()[0] if len(x.mode()) > 0 else 'Desconhecido'
    }).reset_index()

    pedido_stats.columns = ['Pedido', 'Valor_Total_Contrato', 'Data_Venda', 'Meses_Maturidade', 'Grau_Risco']

    # Calcular valor recebido por pedido
    valor_recebido_por_pedido = df_efetivado[
        df_efetivado['Status Parcela'].isin(['Recebido', 'Antecipado'])
    ].groupby('Pedido')['Valor da Parcela (R$)'].sum()

    pedido_stats['Valor_Recebido'] = pedido_stats['Pedido'].map(valor_recebido_por_pedido).fillna(0)
    pedido_stats['Percentual_Recebido'] = (pedido_stats['Valor_Recebido'] / pedido_stats['Valor_Total_Contrato'] * 100).round(1)

    # Filtrar vendas maduras (>= 6 meses)
    vendas_maduras = pedido_stats[pedido_stats['Meses_Maturidade'] >= 6].copy()

    logger.info(f"   Total de pedidos: {len(pedido_stats):,}")
    logger.info(f"   Pedidos maduros (>= 6 meses): {len(vendas_maduras):,}")

    if len(vendas_maduras) == 0:
        logger.warning("   ⚠️ Nenhuma venda madura encontrada (< 6 meses)")
        return {
            'valor_nominal': valor_nominal,
            'valor_real_medio': 0,
            'valor_real_mediana': 0,
            'pct_recebido_medio': 0,
            'pct_recebido_mediana': 0,
            'perda_esperada': 0
        }

    pct_recebido_medio = vendas_maduras['Percentual_Recebido'].mean()
    pct_recebido_mediana = vendas_maduras['Percentual_Recebido'].median()

    valor_real_medio = valor_nominal * (pct_recebido_medio / 100)
    valor_real_mediana = valor_nominal * (pct_recebido_mediana / 100)

    logger.info(f"   Valor Nominal: R$ {valor_nominal:,.2f}")
    logger.info(f"   Valor Real (média): R$ {valor_real_medio:,.2f} ({pct_recebido_medio:.1f}% recebido)")
    logger.info(f"   Valor Real (mediana): R$ {valor_real_mediana:,.2f} ({pct_recebido_mediana:.1f}% recebido)")
    logger.info(f"   Perda esperada: R$ {valor_nominal - valor_real_medio:,.2f} ({100 - pct_recebido_medio:.1f}%)")

    # Análise por grau de risco - incluir todos os graus
    graus_risco = sorted(vendas_maduras['Grau_Risco'].unique(),
                         key=lambda x: {'Alto': 1, 'Médio': 2, 'Baixo': 3, '-': 4}.get(x, 5))

    valor_por_risco = {}
    for risco in graus_risco:
        vendas_risco = vendas_maduras[vendas_maduras['Grau_Risco'] == risco]

        if len(vendas_risco) > 0:
            pct_recebido = vendas_risco['Percentual_Recebido'].mean()
            valor_real = valor_nominal * (pct_recebido / 100)

            # Usar nome mais descritivo para sem classificação
            risco_display = 'Sem Classificação' if risco == '-' else risco

            valor_por_risco[risco_display] = {
                'pedidos': len(vendas_risco),
                'pct_recebido': round(pct_recebido, 1),
                'valor_real': round(valor_real, 2),
                'perda': round(valor_nominal - valor_real, 2)
            }
            logger.info(f"   Risco {risco_display}: R$ {valor_real:,.2f} ({pct_recebido:.1f}% recebido)")

    return {
        'valor_nominal': valor_nominal,
        'valor_real_medio': round(valor_real_medio, 2),
        'valor_real_mediana': round(valor_real_mediana, 2),
        'pct_recebido_medio': round(pct_recebido_medio, 1),
        'pct_recebido_mediana': round(pct_recebido_mediana, 1),
        'perda_esperada': round(valor_nominal - valor_real_medio, 2),
        'valor_por_risco': valor_por_risco,
        'pedidos_analisados': len(vendas_maduras),
        'vendas_maduras_df': vendas_maduras
    }


def format_excel_brazilian_style(file_path: Path):
    """
    Aplica formatação brasileira ao Excel: moeda, porcentagem, largura de colunas, etc.

    Args:
        file_path: Caminho do arquivo Excel
    """
    wb = load_workbook(file_path)

    # Estilos
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_alignment = Alignment(horizontal="left", vertical="center")
    number_alignment = Alignment(horizontal="right", vertical="center")

    border_side = Side(style='thin', color='D3D3D3')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    # Formatos brasileiros
    money_format = 'R$ #,##0.00'
    percentage_format = '0.0"%"'
    number_format = '#,##0'
    date_format = 'DD/MM/YYYY HH:MM:SS'

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Formatar cabeçalho (primeira linha)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Ajustar largura das colunas e formatar células
        for idx, column in enumerate(ws.columns, 1):
            column_letter = get_column_letter(idx)
            column_cells = list(column)
            header_cell = column_cells[0]
            header_value = str(header_cell.value) if header_cell.value else ""

            # Determinar largura ideal
            max_length = len(header_value)
            for cell in column_cells[1:]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            # Ajustar largura (mínimo 12, máximo 50)
            adjusted_width = min(max(max_length + 2, 12), 50)
            ws.column_dimensions[column_letter].width = adjusted_width

            # Aplicar formatação específica por tipo de coluna
            for cell in column_cells[1:]:
                cell.border = border

                # Valores monetários (coluna contém "Valor" ou "Perda")
                if any(keyword in header_value for keyword in ['Valor', 'Perda']):
                    cell.number_format = money_format
                    cell.alignment = number_alignment

                # Porcentagens (coluna contém "%" ou "Pct")
                elif '%' in header_value or 'Pct' in header_value:
                    cell.number_format = percentage_format
                    cell.alignment = number_alignment

                # Números inteiros (Pedidos)
                elif 'Pedidos' in header_value:
                    cell.number_format = number_format
                    cell.alignment = number_alignment

                # Data
                elif 'Data' in header_value:
                    cell.number_format = date_format
                    cell.alignment = cell_alignment

                # Texto
                else:
                    cell.alignment = cell_alignment

        # Congelar primeira linha (cabeçalho)
        ws.freeze_panes = 'A2'

    wb.save(file_path)
    logger.info(f"   ✨ Formatação brasileira aplicada com sucesso")


def gerar_cenarios_precificacao(taxa_realizacao_ponderada: float, preco_atual: float) -> pd.DataFrame:
    """
    Gera cenários de precificação considerando inadimplência.

    Args:
        taxa_realizacao_ponderada: Taxa de realização média ponderada (ex: 0.629 = 62.9%)
        preco_atual: Preço atual do produto

    Returns:
        DataFrame com cenários de precificação
    """
    valor_liquido_atual = preco_atual * taxa_realizacao_ponderada

    # Lista de preços a testar (até R$ 3.500 para não ultrapassar R$ 2.200 líquido)
    precos = [
        preco_atual,  # Atual
        2300.00,
        2400.00,
        2500.00,
        2600.00,
        2700.00,
        2800.00,
        2861.84,  # Para receber R$ 1.800
        2900.00,
        3000.00,
        3100.00,
        3200.00,
        3300.00,
        3400.00,
        3497.53,  # Para receber R$ 2.200 (máximo)
    ]

    cenarios = []

    for preco in precos:
        # Valor líquido recebido (após inadimplência)
        valor_liquido = preco * taxa_realizacao_ponderada

        # Aumentos vs preço atual
        aumento_absoluto = preco - preco_atual
        aumento_percentual = (aumento_absoluto / preco_atual) * 100

        # Ganhos vs valor líquido atual
        ganho_absoluto = valor_liquido - valor_liquido_atual
        ganho_percentual = (ganho_absoluto / valor_liquido_atual) * 100

        cenarios.append({
            'Preço Cobrado': preco,
            'Aumento (%)': aumento_percentual,
            'Aumento (R$)': aumento_absoluto,
            'Valor Líquido Recebido': valor_liquido,
            'Ganho vs Atual (%)': ganho_percentual,
            'Ganho vs Atual (R$)': ganho_absoluto
        })

    return pd.DataFrame(cenarios)


def generate_summary_report(
    inadimplencia: dict,
    valor_real: dict,
    evolucao: pd.DataFrame,
    risco: pd.DataFrame,
    output_path: Path
):
    """
    Gera relatório único em Excel com 5 abas formatadas.

    Args:
        inadimplencia: Dict com estatísticas de inadimplência
        valor_real: Dict com estatísticas de valor real
        evolucao: DataFrame com evolução temporal
        risco: DataFrame com análise por risco
        output_path: Caminho do arquivo Excel de saída
    """
    logger.info("💾 Gerando relatório Excel com 5 abas...")

    # 1. Resumo geral
    resumo = pd.DataFrame([{
        'Data Análise': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
        'Valor Nominal TMB': valor_real['valor_nominal'],
        'Valor Real Médio': valor_real['valor_real_medio'],
        'Valor Real Mediana': valor_real['valor_real_mediana'],
        'Pct Recebido Médio (%)': valor_real['pct_recebido_medio'],
        'Perda Esperada': valor_real['perda_esperada'],
        'Taxa Inadimplência (%)': inadimplencia['taxa_inadimplencia'],
        'Valor Total Contratado': inadimplencia['valor_total'],
        'Valor Vencido': inadimplencia['valor_vencido'],
        'Pedidos Analisados': valor_real['pedidos_analisados']
    }])

    # 2. Valor por risco
    valor_risco_df = pd.DataFrame()
    if 'valor_por_risco' in valor_real:
        valor_risco_data = []
        for risco_nome, dados in valor_real['valor_por_risco'].items():
            valor_risco_data.append({
                'Grau de Risco': risco_nome,
                'Pedidos': dados['pedidos'],
                'Pct Recebido (%)': dados['pct_recebido'],
                'Valor Real': dados['valor_real'],
                'Perda': dados['perda']
            })
        valor_risco_df = pd.DataFrame(valor_risco_data)

    # 3. Renomear colunas da evolução temporal
    evolucao_formatted = evolucao.copy()
    evolucao_formatted.columns = ['Período', 'Pedidos', 'Valor Total', 'Valor Vencido', 'Taxa Inadimplência (%)']

    # 4. Renomear colunas do risco
    risco_formatted = risco.copy()
    risco_formatted.columns = ['Grau de Risco', 'Pedidos', 'Valor Total', 'Valor Vencido', 'Taxa Inadimplência (%)']

    # 5. Gerar cenários de precificação
    # Calcular taxa de realização ponderada
    taxa_realizacao = valor_real['pct_recebido_medio'] / 100  # Converter % para decimal
    preco_nominal = valor_real['valor_nominal']
    cenarios_precificacao = gerar_cenarios_precificacao(taxa_realizacao, preco_nominal)

    # Criar arquivo Excel com múltiplas abas
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        resumo.to_excel(writer, sheet_name='Resumo', index=False)
        evolucao_formatted.to_excel(writer, sheet_name='Evolucao_Temporal', index=False)
        risco_formatted.to_excel(writer, sheet_name='Por_Grau_Risco', index=False)
        if not valor_risco_df.empty:
            valor_risco_df.to_excel(writer, sheet_name='Valor_Real_Risco', index=False)
        cenarios_precificacao.to_excel(writer, sheet_name='Cenarios_Precificacao', index=False)

    # Aplicar formatação brasileira
    format_excel_brazilian_style(output_path)

    logger.info(f"   ✅ Relatório gerado: {output_path}")
    logger.info(f"   📑 Abas: Resumo, Evolucao_Temporal, Por_Grau_Risco, Valor_Real_Risco, Cenarios_Precificacao")


def main():
    parser = argparse.ArgumentParser(description='Analisa inadimplência TMB e calcula valor real de vendas')
    parser.add_argument('--contas-receber', required=True, help='Caminho do arquivo de contas a receber (Excel)')
    parser.add_argument('--output', default=None, help='Caminho do arquivo Excel de saída (padrão: files/analises/tmb_inadimplencia_relatorio.xlsx)')
    parser.add_argument('--valor-nominal', type=float, default=2200.40, help='Valor nominal do contrato TMB (padrão: 2200.40)')

    args = parser.parse_args()

    # Determinar caminho de saída
    if args.output:
        output_path = Path(args.output)
    else:
        project_root = Path(__file__).parent.parent.parent
        output_path = project_root / 'files' / 'analises' / 'tmb_inadimplencia_relatorio.xlsx'

    # Criar diretório se não existir
    output_path.parent.mkdir(parents=True, exist_ok=True)

    logger.info("="*80)
    logger.info("🚀 ANÁLISE DE INADIMPLÊNCIA TMB")
    logger.info("="*80)
    logger.info(f"   Arquivo: {args.contas_receber}")
    logger.info(f"   Output: {output_path}")

    # Carregar dados
    logger.info("\n📂 Carregando arquivo de contas a receber...")
    df = pd.read_excel(args.contas_receber)
    logger.info(f"   ✅ {len(df):,} parcelas carregadas")

    # Filtrar apenas pedidos efetivados
    df_efetivado = df[df['Status Pedido'] == 'Efetivado'].copy()
    logger.info(f"   ✅ {len(df_efetivado):,} parcelas de pedidos efetivados")

    # Executar análises
    print("\n" + "="*80)
    inadimplencia = analyze_inadimplencia_geral(df_efetivado)

    print("\n" + "="*80)
    evolucao = analyze_evolucao_temporal(df_efetivado)

    print("\n" + "="*80)
    risco = analyze_por_grau_risco(df_efetivado)

    print("\n" + "="*80)
    valor_real = calculate_valor_real_venda(df_efetivado, args.valor_nominal)

    # Gerar relatórios
    print("\n" + "="*80)
    generate_summary_report(inadimplencia, valor_real, evolucao, risco, output_path)

    # Conclusão
    print("\n" + "="*80)
    logger.info("✅ ANÁLISE CONCLUÍDA")
    logger.info("="*80)
    logger.info(f"\n📊 RESUMO:")
    logger.info(f"   Valor Nominal TMB: R$ {valor_real['valor_nominal']:,.2f}")
    logger.info(f"   Valor Real (esperado): R$ {valor_real['valor_real_medio']:,.2f}")
    logger.info(f"   Taxa de Inadimplência: {inadimplencia['taxa_inadimplencia']:.1f}%")
    logger.info(f"   Perda por venda: R$ {valor_real['perda_esperada']:,.2f}")

    if 'valor_por_risco' in valor_real:
        logger.info(f"\n   Por Grau de Risco:")
        for risco_nome, dados in valor_real['valor_por_risco'].items():
            logger.info(f"      {risco_nome}: R$ {dados['valor_real']:,.2f} ({dados['pct_recebido']}% recebido)")


if __name__ == '__main__':
    main()
