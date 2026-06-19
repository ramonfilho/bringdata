#!/usr/bin/env python3
"""
Extrator pós-batch: lê os relatórios de validação (janela 60d) dos 18 lançamentos
e monta DUAS tabelas tidy:

  1. sales_tidy.parquet  — uma linha por venda:
       lf, trackeado, grupo, data_captura, data_venda, contratado, recebido, fonte, meio
  2. spend_grupo.csv     — gasto de captação por (lf, grupo)  [do bloco Comparação ML]

A partir daí toda a análise (janelas week/60d, corte de coorte por data_captura,
banda piso/teto das órfãs, ROAS, dinheiro novo) é pós-processamento puro.

Fonte de revenue = aba "Detalhes das Conversões" (venda-a-venda, 2 bases, com datas).
Fonte de spend   = aba "Comparação ML" (1º bloco = lançamento cheio, gasto por grupo).
"""
import sys
from pathlib import Path
from datetime import datetime
import openpyxl
import pandas as pd
import yaml

BASE = Path("/Users/ramonmoreira/bring_data.worktrees/ab-arm/V2")
VALID = BASE / "outputs/validation"
OUT = BASE / "analise_valor_ml"
LAUNCHES = yaml.safe_load(open(BASE / "configs/launches.yaml"))

LFS = ['LF40','LF41','LF42','LF43','LF44','LF45','LF46','LF47','LF48','LF49',
       'LF50','LF51','LF52','LF53','LF54','LF55','LF56','DEV19']

# Detalhes (0-idx): 0 Trackeado, 7 Grupo, 8 DataCaptura, 9 DataVenda,
#                   10 Contratado, 11 Recebido, 12 Fonte, 13 Meio
def _date(x):
    if x is None:
        return None
    if isinstance(x, datetime):
        return x
    s = str(x).strip()
    if not s:
        return None
    try:
        return datetime.fromisoformat(s[:19].replace("Z", ""))
    except ValueError:
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(s[:len(fmt)], fmt)
            except ValueError:
                pass
    return None


def find_file(lf):
    """Pega o xlsx 60d mais recente do LF (re-rodada do batch)."""
    cands = []
    for p in VALID.glob(f"*/{lf} - *.xlsx"):
        cands.append((p.stat().st_mtime, p))
    if not cands:
        return None
    return max(cands)[1]


def read_detalhes(path, lf):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Detalhes das Conversões"]
    out = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        if r is None or all(c is None for c in r):
            continue
        out.append(dict(
            lf=lf,
            trackeado=(str(r[0]).strip() if r[0] else ""),
            grupo=(str(r[7]).strip() if r[7] else "None"),
            data_captura=_date(r[8]),
            data_venda=_date(r[9]),
            contratado=float(r[10] or 0),
            recebido=float(r[11] or 0),
            fonte=(str(r[12]).strip().lower() if r[12] else ""),
            meio=(str(r[13]).strip() if r[13] else ""),
        ))
    wb.close()
    return out


def read_spend(path):
    """Gasto de captação por grupo: 1º bloco da aba Comparação ML (linha 'Gasto')."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Comparação ML"]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    col_map, hdr_idx = {}, None
    for i, row in enumerate(grid):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if "Métrica" in cells and any(c.startswith("Champion") for c in cells):
            hdr_idx = i
            for j, c in enumerate(cells):
                lc = c.lower()
                if lc.startswith("champion"):
                    col_map[j] = "Champion"
                elif lc.startswith("challenger"):
                    col_map[j] = "Challenger"
                elif lc.startswith("controle"):
                    col_map[j] = "Controle"
            break
    spend = {}
    if hdr_idx is not None:
        for row in grid[hdr_idx + 1: hdr_idx + 15]:
            c0 = str(row[0]).strip().lower() if row and row[0] is not None else ""
            if c0 == "gasto":
                for j, grp in col_map.items():
                    try:
                        v = float(row[j])
                        spend[grp] = v
                    except (TypeError, ValueError):
                        pass
                break
    return spend


def total_spend(path):
    """Gasto total do lançamento (Performance Geral -> 'Gasto Total')."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Performance Geral"]
    val = None
    for row in ws.iter_rows(values_only=True):
        if row and row[0] and str(row[0]).strip().lower() == "gasto total":
            try:
                val = float(str(row[1]).replace("R$", "").replace(".", "").replace(",", ".").strip())
            except (TypeError, ValueError, AttributeError):
                val = row[1]
            break
    wb.close()
    return val


def main():
    sales_rows, spend_rows, manifest = [], [], []
    for lf in LFS:
        p = find_file(lf)
        if p is None:
            print(f"  FALTA {lf}: nenhum xlsx encontrado")
            manifest.append(dict(lf=lf, arquivo=None, n_vendas=0))
            continue
        det = read_detalhes(p, lf)
        sales_rows.extend(det)
        sp = read_spend(p)
        cfg = LAUNCHES.get(lf, {})
        for grp, val in sp.items():
            spend_rows.append(dict(lf=lf, grupo=grp, gasto=val))
        manifest.append(dict(lf=lf, arquivo=p.name, n_vendas=len(det),
                             grupos_spend=",".join(sorted(sp)) or "-",
                             cap_start=str(cfg.get("cap_start")),
                             cap_end=str(cfg.get("cap_end")),
                             vendas_start=str(cfg.get("vendas_start"))))
        print(f"  {lf:6} {p.name[:34]:36} vendas={len(det):4}  spend_grp={sorted(sp)}")

    sales = pd.DataFrame(sales_rows)
    spend = pd.DataFrame(spend_rows)
    man = pd.DataFrame(manifest)
    OUT.mkdir(parents=True, exist_ok=True)
    sales.to_parquet(OUT / "sales_tidy.parquet", index=False)
    spend.to_csv(OUT / "spend_grupo.csv", index=False)
    man.to_csv(OUT / "manifest.csv", index=False)
    print(f"\n  -> sales_tidy.parquet ({len(sales)} vendas, {sales['lf'].nunique()} LFs)")
    print(f"  -> spend_grupo.csv ({len(spend)} linhas)")
    print(f"  -> manifest.csv")


if __name__ == "__main__":
    main()
